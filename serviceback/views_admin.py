import base64, datetime, json, requests, os, xlrd
from django.http import HttpResponse
from rest_framework.decorators import api_view
from urllib3.exceptions import InsecureRequestWarning  # added
from .models import *
from django.db import connection
from django.contrib.auth.models import User, Group
from django.core.mail import BadHeaderError
from django.db.models import Max
import pandas as pd
import psycopg2 as ps
import shutil
import xlsxwriter as xlsxwriter

import locale
import string
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# *******************************************
# Системные функции для работы приложения
# *******************************************
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)
basepath = '//192.168.5.29/ssd/'


def get_org_id(username):
    usrobj = ProfileUser.objects.all()
    usrobj = usrobj.filter(name=username)
    for itemorg in usrobj:
        id_org = itemorg.id_org
    return id_org


def statusfromDBbyOrg(id_org, id_group, status, datestatus, iin):
    
    # if not (status == '99' and status == '10'): 
    # status = '%' + status + '%'
    id_group = '%' + id_group + '%'
    iin = '%' + iin  + '%'
    
    query = f"""
                WITH visit as(
                    SELECT
                        *
                    FROM
                        serviceback_visits 
                    WHERE
                        id_org = '{id_org}' 
                        and datestatus = '{datestatus}' 
                        and id_group like '{id_group}' 
                        and iin like '{iin}'),
                tempgroup as(
                    SELECT
                        id_org,
						id_group,
						group_name
                    FROM
                        serviceback_groups
                    WHERE
                        (id_org, id_group) in (SELECT
											  	id_org,
											  	id_group
											  FROM
											  	visit)
                        and is_delete = false),
                tempusers as(
                    SELECT
                        username,
						first_name
                    FROM
                        auth_user
					WHERE
						username in (
								SELECT
									username
								FROM
									visit)),
                tmptbl_1 as(
                    SELECT
                        iin,
                        max(timestatus) as timestatus
                    FROM
                        visit
                    GROUP BY
                        iin),                                
                tmptbl as(
                    SELECT
                        iin, 
                        id_group, 
                        timestatus
                    FROM
                        visit
                    WHERE 
                        (iin,
                        timestatus) in 
                            (SELECT
                                iin, 
                                timestatus
                            FROM
                                tmptbl_1)),
                resvisit as (
                    SELECT
                        *
                    FROM
                        visit
					WHERE (iin, id_group, timestatus) in (
												SELECT
													tmptbl.iin,
													tmptbl.id_group,
													tmptbl.timestatus
												FROM
													tmptbl as tmptbl)),
				childs as (
					SELECT
						iin,
						name as child_name,
						image_url,
						registered
					FROM
						serviceback_childs
					WHERE
						id_org in (
									SELECT
										id_org
									FROM
										resvisit)),
				readytable as (
                        SELECT 
                            childs.iin as child_iin, 
                            childs.child_name as child_name,
                            childs.image_url as image_url,
                            CASE
                                WHEN NOT childs.registered  THEN '0'
                                WHEN visits.status IS NULL  THEN '1'
                                ELSE visits.status
                            END as status,
                            CASE
                                WHEN NOT childs.registered  THEN 'Первичная идентификация'
                                WHEN visits.status IS NULL or visits.status = '1' THEN 'Ожидают сканирования'
                                WHEN visits.status = '2'    THEN 'Посещён'
                                WHEN visits.status = '3'    THEN 'Больничный'
                                WHEN visits.status = '4'    THEN 'Отпуск'
                                WHEN visits.status = '5'    THEN 'Не посетил'
                                WHEN visits.status = '9'    THEN 'Не распознано'
                                WHEN visits.status = '10' and NOT comments IS NULL THEN 'Ложное фото. Действие зафиксировано!'
                                WHEN visits.status = '10' and comments IS NULL THEN 'Фото проверяется администратором!'
                            END as statusname,
                            CASE
                                WHEN visits.datestatus IS NULL THEN '{datestatus}'
                                ELSE visits.datestatus
                            END as datestatus,
                            visits.image_url as visit_photo,
                            visits.comments as vis_comments,
                            visits.id_group,
                            gr.group_name,
                            visits.username,
                            users.first_name
                        FROM
							resvisit as visits		
                        FULL JOIN
							childs as childs
						ON
                            visits.iin = childs.iin
                        LEFT JOIN
							tempgroup as gr
						ON
                            gr.id_org = visits.id_org
                            and gr.id_group = visits.id_group
                        LEFT JOIN
							tempusers as users 
						ON
                            visits.username = users.username)
					SELECT  
						child_iin as iin, 
                        child_name as name, 
                        status, 
                        statusname, 
                        to_char(datestatus,'dd.mm.yyyy') as datestatus, 
                        id_group, 
                        group_name, 
                        username, 
                        first_name,
                        replace(image_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as image_url,
                        replace(visit_photo, 'FilesArchiv', 'https://face11.qazna24.kz/media') as visit_photo 
                FROM 
                    readytable
                WHERE 
                    case when '{status}' = '' then true
                         when '{status}' = '10' then status = '10' and vis_comments IS NOT NULL
                         when '{status}' = '11' then status = '10' and vis_comments IS NULL
                         else status = '{status}' end
                ORDER BY
                    group_name,
                    child_name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
               for row in cursor.fetchall()]
    return result

#нужно удалить
# Сервис для обновления данных в базе данных
@api_view(['POST'])
def updateChilds(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # Структура должна быть как показано ниже
    datajson = """[
    {
        "org_name": "JSOFT",
        "id_org": 99999,
        "group_mass": [
            {
                "id_group": "10001716",
                "group_name": "АДМ",
                "child_mass": [
                    {
                        "child_name": "Акжигитова Индира Сансызбаевна",
                        "child_iin": "730101403130",
                        "child_present": "false",
                        "status": "",
                        "child_image": "child_image"
                    },
                    {
                        "child_name": "Ашимов Танат Канатович",
                        "child_iin": "910511301086",
                        "child_present": "false",
                        "status": "",
                        "child_image": "child_image"
                    }
                ]
            },
            {
                "id_group": "10001717",
                "group_name": "Методист",
                "child_mass": [
                    {
                        "child_name": "Аманжол Бексұлтан Батырханұлы",
                        "child_iin": "940910301873",
                        "child_present": "false",
                        "status": "",
                        "child_image": "child_image"
                    }
                ]
            }
        ]
    }
]"""

    # Получаем данные с тела запроса
    data = json.loads(request.body)
    for orgitem in data:
        id_org = orgitem['org_ID']
        org_name = orgitem['org_name']
        # Получаем имеющиеся данные с БД (таблица организации)
        findorg = Organizations.objects.all()
        # Фильтруем по коду организации
        findorg = findorg.filter(id_org=id_org)

        # Проверяем, есть ли в БД такая организация
        existorg = False
        for item in findorg:
            existorg = True
            break

        # Если нет, то создаем организацию, если есть, то ничего не делаем
        if existorg == False:
            itemOrg = Organizations(id_org=id_org, org_name=org_name)
            itemOrg.save()


        # Получаем массив групп в запросе и также проверяем на существование
        groups = orgitem['group_mass']
        for itemgr in groups:
            id_group = itemgr['group_id']
            # Получаем из БД все группы и фильтруем по коду орг и группы
            findgroup = Groups.objects.all()
            findgroup = findgroup.filter(id_group=id_group)
            findgroup = findgroup.filter(id_org=id_org)

            # Проверим, есть ли такая группа
            existgr = False
            for item in findgroup:
                existgr = True
                break

            # Если нет в БД, то создаем группу
            if existgr == False:
                itemGr = Groups(id_group=id_group, group_name=itemgr['group_name'], id_org=id_org)
                itemGr.save()


            # Получаем детей в группе из запроса
            childs = itemgr['child_mass']
            for itemch in childs:
                iin = itemch['child_iin']
                name = itemch['child_name']

                # Получаем детей в БД, фильтруем по орг, группе, и коду детей
                findchild = Childs.objects.all()
                findchild = findchild.filter(id_org=id_org)
                findchild = findchild.filter(id_group=id_group)
                findchild = findchild.filter(iin=iin)

                # Проверяем есть ли в БД ребенок с таким кодом
                existchild = False
                for item in findchild:
                    existchild = True
                    break

                # Если нет в БД, то создаем запись
                if existchild == False:
                    itemChld = Childs(id_group=id_group,
                                      id_org=id_org, iin=iin, name=name)
                    itemChld.save()

                # Если нет папки ребенка, то создаем папку (формат: С910114301692)
                # path_child = path_group + '/C' + str(iin)
                # direxist = os.path.isdir(path_child)
                # if not direxist:
                    # os.mkdir(path_child)

    return HttpResponse('{"status": "success"}', content_type="application/json")

# Сервис авторизациии пользователя. Если только админ флаг установлен в Истина
@api_view(['GET'])
def authuser(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    is_admin_org = False
    is_admin = False
    username = request.user
    
    user = User.objects.get(pk=username.id)
    usrobj = ProfileUser.objects.filter(name=username)

    existuser = False
    for itemorg in usrobj:
        existuser = True
        if itemorg.is_adm_org:
            is_admin_org = True
            break

    if not existuser and user.is_staff:
        newprofile = ProfileUser()
        newprofile.id_org = '910114301692'
        newprofile.name = username
        newprofile.is_adm_org = True
        newprofile.is_pass_chek = False
        newprofile.is_delete = False
        newprofile.save()
        is_admin_org = True

    if not is_admin_org:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    is_admin = user.is_staff

    for itemorg in usrobj:
        return HttpResponse('{"status": "Авторизован", "change_pass": "'+ str(not itemorg.is_pass_chek) +'", "is_admin": "'+ str(is_admin) +'"}', content_type="application/json", status=200)

#нужно удалить
@api_view(['GET'])
def getinfo(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)

    id_org = get_org_id(request.user)

    fullusername = ''
    editfakephoto = False
    username = request.user
    user_org_id = ''
    user_org_name = ''
    is_metodist = False

    gr = Group.objects.filter(user = username.id)
    for itemperm in gr:
        if itemperm.name == 'editFakePhoto':
            editfakephoto = True
        if (itemperm.name == 'MetodistDDO'):
            is_metodist = True

    usernameitem = User.objects.filter(username=username)
    for itemfullname in usernameitem:
        fullusername = itemfullname.first_name

    orgitem = Organizations.objects.filter(id_org=id_org)
    for item in orgitem:
        user_org_id = item.id_org
        user_org_name = item.org_name
        user_id_region = item.id_region

    is_staff = str(request.user.is_staff)

    return HttpResponse('{"username": "'+ fullusername +'", "editfakephoto": "' + str(editfakephoto) +'", "user_org_id": "'+ user_org_id+'", "user_org_name": "'+user_org_name+'", "is_staff": "'+is_staff+'", "is_metodist": "'+str(is_metodist)+'", "user_id_region": "'+str(user_id_region)+'"}', content_type="application/json")

#нужно удалить
@api_view(['GET'])
def getinfoorg(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)
    id_org = get_org_id(request.user)

    user_org_id = ''
    user_org_name = ''
    is_metodist = False

    groups = request.user.groups.all()
    for group in groups:
        if (group.name == 'MetodistDDO'):
            is_metodist = True
            break
    
    orgitem = Organizations.objects.filter(id_org=id_org)
    for item in orgitem:
        user_org_id = item.id_org
        user_org_name = item.org_name
        user_id_region = item.id_region
        is_staff = str(request.user.is_staff)
        
    # for itemorg in usrobj:
    return HttpResponse('{"user_org_id": "'+ user_org_id+'", "user_org_name": "'+user_org_name+'", "is_staff": "'+is_staff+'", "is_metodist": "'+str(is_metodist)+'", "user_id_region": "'+str(user_id_region)+'"}', content_type="application/json")


@api_view(['GET'])
def changepass(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    username = request.user   
    password = request.GET.get("password")
  
    try:
        u = User.objects.get(username=username)
        u.set_password(password)
        u.save()

        usrobj = ProfileUser.objects.filter(name=username)
        for itemorg in usrobj:
            itemorg.is_pass_chek = True
            itemorg.save()

        return HttpResponse('{"status": "Пароль успешно изменен"}', content_type="application/json", status=200)
    except User.DoesNotExist:
        return HttpResponse('{"status": "Ошибка изменения пароля"}', content_type="application/json", status=500)

# сервис для дневного плана
@api_view(['GET'])
def childstatus(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)

    id_group = request.GET.get("id_group")
    id_org = request.GET.get("id_org")
    datestatus = request.GET.get("datestatus")

    if datetime.datetime.now().strftime("%d.%m.%Y") == datestatus: 
        # дневной план для текущей даты
        query = f"""
                    with child_info as(
                         SELECT
							 id_org,
							 id_group,
                             registered,
                             name as child_name,
                             iin
                         FROM
                             serviceback_childs
                         WHERE
                             id_org = '{id_org}'
                             and CASE WHEN '{id_group}' = '' THEN TRUE
                                	  ELSE id_group = '{id_group}'
                                 END
							and not is_delete),
                  visit as(
                        SELECT
                            id as _id,
                            id_org,
                            id_group,
                            iin,
                            datestatus,
                            timestatus,
                            status,
                            username,
                            comments
                        FROM
                            serviceback_visits
                        WHERE
                            (id_org,
                            id_group,
                            iin) in (
                                SELECT
                                    id_org,
                                    id_group,
                                    iin
                                FROM
                                    child_info)
                            and datestatus = '{datestatus}'),
                        max_vis_id as(
                                SELECT
                                    max(_id) as _id,
                                    iin
                                FROM
                                    visit
                                GROUP BY
                                    iin),
                        vis_finall as(
                                SELECT
                                    *
                                FROM
                                    visit
                                WHERE
                                    _id in(
                                        SELECT
                                            _id
                                        FROM
                                            max_vis_id)),
                        username as(
                                SELECT
                                    username,
                                    first_name
                                FROM
                                    auth_user
                                WHERE
                                    username in (
                                        SELECT
                                            username
                                        FROM
                                            vis_finall)),
                        vis_with_username as(
                                SELECT
                                    vis.*,
                                    us.first_name
                                FROM
                                    vis_finall as vis
                                LEFT JOIN
                                    username as us
                                ON
                                    vis.username = us.username),
                        all_childs as(
                                SELECT
                                    vis.id_org,
                                    vis.id_group,
                                    vis.iin,
                                    vis.status,
                                    vis.timestatus as timestatus,
                                    vis.username  as username,
                                    vis.first_name as first_name,
                                    vis.comments as com
                                FROM
                                    vis_with_username as vis),
                        group_info as(
                                SELECT
                                    id_org,
                                    id_group,
                                    group_name
                                FROM
                                    serviceback_groups
                                WHERE
                                    (id_org, id_group) in (
                                                        SELECT
                                                            id_org,
                                                            id_group
                                                        FROM
                                                            child_info)),
                        full_info as(
                                SELECT					
                                    vis.username,
									vis.first_name,
									vis.status,
									vis.com,
                                    gr.group_name,
									ch.iin,
									ch.id_org,
									ch.id_group,
                                    ch.child_name,
                                    ch.registered
                                FROM
                                    child_info as ch
                                LEFT JOIN
                                    group_info as gr
                                ON
                                    ch.id_org = gr.id_org
                                    and ch.id_group = gr.id_group
                                LEFT JOIN
                                    all_childs as vis
                                ON
                                    vis.iin = ch.iin
									and ch.id_org = vis.id_org
                                    and ch.id_group = vis.id_group),
                        tmp_full as(
                                SELECT
                                    id_org,
                                    id_group,
                                    group_name,
                                    iin,
                                    child_name as name,
                                    username,
                                    first_name,
                                    CASE 
                                        WHEN status is NULL and NOT registered THEN '0'
										WHEN status is NULL and registered THEN '1'
                                        WHEN status = '10' and com IS NULL THEN '11'
                                        ELSE status
                                    END as status,
                                    CASE
                                        WHEN status is NULL and NOT registered THEN 'Первичная идентификация'
                                        WHEN status is NULL and registered THEN 'Ожидает сканирования'
                                        WHEN status = '2' THEN 'Присутствие'
                                        WHEN status = '3' THEN 'Больничный'
                                        WHEN status = '4' THEN 'Отпуск'
                                        WHEN status = '5' THEN 'Отсутствие'
                                        WHEN status = '9' THEN 'Не распознано'
                                        WHEN status = '10' and NOT com IS NULL THEN 'Ложное фото!'
                                        WHEN status = '10' and com IS NULL THEN 'Фото проверяется администратором!'
                                    END as statusname
                                FROM
                                    full_info)
                                SELECT 
                                    * 
                                FROM 
                                    tmp_full
                                ORDER BY
									group_name,
                                    name"""
        
        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row))
                for row in cursor.fetchall()]                
    
        return HttpResponse(json.dumps(result), content_type="application/json")
    else:
        query = f"""
                    with visit as(
                        SELECT
                            id as _id,
                            id_org,
                            id_group,
                            iin,
                            datestatus,
                            timestatus,
                            status,
                            username,
                            comments
                        FROM
                            serviceback_visits
                        WHERE
                            id_org = '{id_org}'
                            and CASE WHEN '{id_group}' = '' THEN TRUE
                                ELSE id_group = '{id_group}'
                                END
                            and datestatus = '{datestatus}'),
                        max_vis_id as(
                                SELECT
                                    max(_id) as _id,
                                    iin
                                FROM
                                    visit
                                GROUP BY
                                    iin),
                        vis_finall as(
                                SELECT
                                    *
                                FROM
                                    visit
                                WHERE
                                    _id in(
                                        SELECT
                                            _id
                                        FROM
                                            max_vis_id)),
                        username as(
                                SELECT
                                    username,
                                    first_name
                                FROM
                                    auth_user
                                WHERE
                                    username in (
                                        SELECT
                                            username
                                        FROM
                                            vis_finall)),
                        vis_with_username as(
                                SELECT
                                    vis.*,
                                    us.first_name
                                FROM
                                    vis_finall as vis
                                LEFT JOIN
                                    username as us
                                ON
                                    vis.username = us.username),
                        all_childs as(
                                SELECT
                                    vis.id_org,
                                    vis.id_group,
                                    vis.iin,
                                    vis.status,
                                    vis.timestatus as timestatus,
                                    vis.username  as username,
                                    vis.first_name as first_name,
                                    vis.comments as com
                                FROM
                                    vis_with_username as vis),
                        child_info as(
                                SELECT
                                    registered,
                                    name as child_name,
                                    iin
                                FROM
                                    serviceback_childs
                                WHERE
                                    iin in(
                                        SELECT
                                            iin
                                        FROM
                                            all_childs)),
                        group_info as(
                                SELECT
                                    id_org,
                                    id_group,
                                    group_name
                                FROM
                                    serviceback_groups
                                WHERE
                                    (id_org, id_group) in (
                                                        SELECT
                                                            id_org,
                                                            id_group
                                                        FROM
                                                            all_childs)),
                        full_info as(
                                SELECT					
                                    vis.*,
                                    gr.group_name,
                                    ch.child_name,
                                    ch.registered
                                FROM
                                    all_childs as vis
                                LEFT JOIN
                                    group_info as gr
                                ON
                                    vis.id_org = gr.id_org
                                    and vis.id_group = gr.id_group
                                LEFT JOIN
                                    child_info as ch
                                ON
                                    vis.iin = ch.iin),
                        tmp_full as(
                                SELECT
                                    id_org,
                                    id_group,
                                    group_name,
                                    iin,
                                    child_name as name,
                                    username,
                                    first_name,
                                    CASE 
                                        WHEN status = '1' and NOT registered THEN '0'
                                        WHEN status = '10' and com IS NULL THEN '11'
                                        ELSE status
                                    END as status,
                                    CASE
                                        WHEN status = '1' and NOT registered THEN 'Первичная идентификация'
                                        WHEN status = '1' THEN 'Ожидает сканирования'
                                        WHEN status = '2' THEN 'Присутствие'
                                        WHEN status = '3' THEN 'Больничный'
                                        WHEN status = '4' THEN 'Отпуск'
                                        WHEN status = '5' THEN 'Отсутствие'
                                        WHEN status = '9' THEN 'Не распознано'
                                        WHEN status = '10' and NOT com IS NULL THEN 'Ложное фото!'
                                        WHEN status = '10' and com IS NULL THEN 'Фото проверяется администратором!'
                                    END as statusname
                                FROM
                                    full_info)
                                SELECT 
                                    * 
                                FROM 
                                    tmp_full
                                ORDER BY
									group_name,
                                    name"""

        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row))
                for row in cursor.fetchall()]                
    
        return HttpResponse(json.dumps(result), content_type="application/json")

# сервис для дневного плана
@api_view(['GET'])
def childstatus_1(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)

    id_group = request.GET.get("id_group")
    id_org = request.GET.get("id_org")
    
    status = request.GET.get("status")
    datestatus = request.GET.get("datestatus")
    iin = request.GET.get("iin")
    
    result = statusfromDBbyOrg(id_org, id_group, status, datestatus, iin)
    return HttpResponse(json.dumps(result), content_type="application/json")

@api_view(['GET'])
def allstatus(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)
    
    id_org = request.GET.get("id_org")
    id_group = request.GET.get("id_group")   
    datestatus = request.GET.get("datestatus") 
    
    if datetime.datetime.now().strftime("%d.%m.%Y") == datestatus: 
        # количество детей по статусам для текущей даты
        query = f"""
                    with child_info as(
                         SELECT
							 id_org,
							 id_group,
                             registered,
                             iin
                         FROM
                             serviceback_childs
                         WHERE
                             id_org = '{id_org}'
                             and CASE WHEN '{id_group}' = '' THEN TRUE
                                	  ELSE id_group = '{id_group}'
                                 END
							and not is_delete),
                  visit as(
                        SELECT
                            id as _id,
                            id_org,
                            id_group,
                            iin,
                            status,
                            comments as com
                        FROM
                            serviceback_visits
                        WHERE
                            (id_org,
                            id_group,
                            iin) in (
                                SELECT
                                    id_org,
                                    id_group,
                                    iin
                                FROM
                                    child_info)
                            and datestatus = '{datestatus}'),
                        max_vis_id as(
                                SELECT
                                    max(_id) as _id,
                                    iin
                                FROM
                                    visit
                                GROUP BY
                                    iin),
                        all_childs as(
                                SELECT
                                    *
                                FROM
                                    visit
                                WHERE
                                    _id in(
                                        SELECT
                                            _id
                                        FROM
                                            max_vis_id)),
                        full_info as(
                                SELECT					
                                    vis.status,
									vis.com,
                                    ch.iin,
									ch.id_org,
									ch.id_group,
                                    ch.registered
                                FROM
                                    child_info as ch
                                LEFT JOIN
                                    all_childs as vis
                                ON
                                    vis.iin = ch.iin
									and ch.id_org = vis.id_org
                                    and ch.id_group = vis.id_group),
                        tmp_full as(
                                SELECT
                                    id_org,
                                    id_group,
                                    iin,
                                    CASE 
                                        WHEN status is NULL and NOT registered THEN '0'
										WHEN status is NULL and registered THEN '1'
                                        WHEN status = '10' and com IS NULL THEN '11'
                                        ELSE status
                                    END as status
                                FROM
                                    full_info),
						tmp_finall as(
                                SELECT 
                                    id_group, 
									id_org, 
									CASE
										WHEN status = '0'  THEN 1
										ELSE 0
									END as notregister,
									CASE
										WHEN status = '1' THEN 1
										ELSE 0
									END as notscanned,
									CASE
										WHEN status = '2'  THEN 1
										ELSE 0
									END as visited,
									CASE
										WHEN status = '3'  THEN 1
										ELSE 0
									END as bolnich,
									CASE
										WHEN status = '4' THEN 1
										ELSE 0
									END as otpusk,
                                    CASE
										WHEN status = '5' THEN 1
										ELSE 0
									END as notvisited,
									CASE
										WHEN status = '10' THEN 1
										ELSE 0
									END as fake,
									CASE
										WHEN status = '11' THEN 1
										ELSE 0
									END as checkphoto,
									status as common
                                FROM 
                                    tmp_full)
								SELECT
									id_org, 
									sum(notregister) as notreg, 
									sum(notscanned) as notscanned,
                                    sum(notvisited) as notvis, 
									sum(visited) as vis, 
									sum(bolnich) as boln, 
									sum(otpusk) as otp,
									sum(fake) as fake,
									sum(checkphoto) as check,
									count(common) as common
								FROM
									tmp_finall
								GROUP BY
									id_org"""

        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row))
                for row in cursor.fetchall()]

        return HttpResponse(json.dumps(result), content_type="application/json")

    else:
        query = f"""with visit as(
                        SELECT
                            id as _id,
                            id_org,
                            id_group,
                            iin,
                            status,
                            comments as com
                        FROM
                            serviceback_visits
                        WHERE
                            id_org = '{id_org}'
                            and CASE WHEN '{id_group}' = '' THEN TRUE
                                ELSE id_group = '{id_group}'
                                END
                            and datestatus = '{datestatus}'),
                        max_vis_id as(
                                SELECT
                                    max(_id) as _id,
                                    iin
                                FROM
                                    visit
                                GROUP BY
                                    iin),
                        vis_finall as(
                                SELECT
                                    *
                                FROM
                                    visit
                                WHERE
                                    _id in(
                                        SELECT
                                            _id
                                        FROM
                                            max_vis_id)),
                        child_info as(
                                SELECT
                                    registered,
                                    iin
                                FROM
                                    serviceback_childs
                                WHERE
                                    iin in(
                                        SELECT
                                            iin
                                        FROM
                                            vis_finall)),
                        full_info as(
                                SELECT					
                                    vis.*,
                                    ch.registered
                                FROM
                                    vis_finall as vis
                                LEFT JOIN
                                    child_info as ch
                                ON
                                    vis.iin = ch.iin),
                        tmp_full as(
                                SELECT
                                    id_org,
                                    id_group,
                                    iin,
                                    CASE 
                                        WHEN status = '1' and NOT registered THEN '0'
                                        WHEN status = '10' and com IS NULL THEN '11'
                                        ELSE status
                                    END as status
                                FROM
                                    full_info),
						tmp_finall as(
                                SELECT 
                                    id_group, 
									id_org, 
									CASE
										WHEN status = '0'  THEN 1
										ELSE 0
									END as notregister,
									CASE
										WHEN status = '1' THEN 1
										ELSE 0
									END as notscanned,
									CASE
										WHEN status = '2'  THEN 1
										ELSE 0
									END as visited,
									CASE
										WHEN status = '3'  THEN 1
										ELSE 0
									END as bolnich,
									CASE
										WHEN status = '4' THEN 1
										ELSE 0
									END as otpusk,
                                    CASE
										WHEN status = '5' THEN 1
										ELSE 0
									END as notvisited,
									CASE
										WHEN status = '10' THEN 1
										ELSE 0
									END as fake,
									CASE
										WHEN status = '11' THEN 1
										ELSE 0
									END as checkphoto,
									status as common
                                FROM 
                                    tmp_full)
								SELECT
									id_org, 
									sum(notregister) as notreg, 
									sum(notscanned) as notscanned,
                                    sum(notvisited) as notvis, 
									sum(visited) as vis, 
									sum(bolnich) as boln, 
									sum(otpusk) as otp,
									sum(fake) as fake,
									sum(checkphoto) as check,
									count(common) as common
								FROM
									tmp_finall
								GROUP BY
									id_org"""

        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row))
                for row in cursor.fetchall()]

        return HttpResponse(json.dumps(result), content_type="application/json")

# старый метод
@api_view(['GET'])
def allstatus_1(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)
    
    id_org = request.GET.get("id_org")
    id_group = request.GET.get("id_group")   
    id_group = '%' + id_group + '%'
    datestatus = request.GET.get("datestatus") 
    
    query = f"""WITH visit as (
                    SELECT
                        *
                    FROM
                        serviceback_visits				  
                    WHERE
                        id_org = '{id_org}'
                        and id_group LIKE '{id_group}'
                        and datestatus = '{datestatus}'),
                tmptbl as (
                    SELECT
                        iin, 
                        max(timestatus) as timestatus 
                    FROM
                        visit
                    GROUP BY
                        iin),
                resvisit as (
                    SELECT
                        visit.*
                    FROM
                        visit
                    JOIN tmptbl ON
                        visit.iin = tmptbl.iin and
                        visit.timestatus = tmptbl.timestatus),
				statustable as (
                    SELECT 
                        childs.id_group,
                        childs.iin, 
                        childs.name,
                        CASE
                            WHEN NOT childs.registered  THEN '0'
                            WHEN visits.status IS NULL  THEN '1'
                            ELSE visits.status
                        END as status,
                        visits.comments as comments
                    FROM
                        serviceback_childs as childs
                    LEFT JOIN resvisit as visits ON
                        visits.id_org = childs.id_org
                        and visits.id_group = childs.id_group
                        and visits.iin = childs.iin
                        and visits.datestatus = '{datestatus}'
                    WHERE 
                        childs.id_org = '{id_org}'
                        and childs.id_group LIKE '{id_group}'
                        and childs.is_delete = false
                    ORDER BY 
                        childs.name),
                groupstatus as (
                    SELECT 
                        groupstbl.id_group, 
                        groupstbl.id_org, 
                        groupstbl.group_name,
                        CASE
                            WHEN statustable.status = '0'  THEN 1
                            ELSE 0
                        END as notregister,
                        CASE
                            WHEN statustable.status = '1'  THEN 1
                            ELSE 0
                        END as notvisited,
                        CASE
                            WHEN statustable.status = '2'  THEN 1
                            ELSE 0
                        END as visited,
                        CASE
                            WHEN statustable.status = '3'  THEN 1
                            ELSE 0
                        END as bolnich,
                        CASE
                            WHEN statustable.status = '4'  THEN 1
                            ELSE 0
                        END as otpusk,
                        CASE
                            WHEN statustable.status = '10' and NOT comments IS NULL THEN 1
                            ELSE 0
                        END as fake,
                        CASE
                            WHEN statustable.status = '10' and comments IS NULL THEN 1
                            ELSE 0
                        END as checkphoto,
                        CASE
                            WHEN statustable.status in ('2','3','4')  THEN 1
                            ELSE 0
                        END as checked,
                        statustable.status as common
                    FROM 
                        serviceback_groups as groupstbl
                    LEFT JOIN statustable ON 
                        groupstbl.id_group = statustable.id_group
                    WHERE 
                        groupstbl.id_org = '{id_org}'
                        and groupstbl.is_delete = false)
                    select id_org, 
                            sum(notregister) as notreg, 
                            sum(notvisited) as notvis, 
                            sum(visited) as vis, 
                            sum(bolnich) as boln, 
                            sum(otpusk) as otp,
                            sum(fake) as fake,
                            sum(checkphoto) as check,
                            sum(checked) as cheked,
                            count(common) as common from groupstatus
                    group by (id_org)"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
               for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result), content_type="application/json")

#нужно удалить
@api_view(['GET'])
def statusphoto(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)
    image_url = request.GET.get("image_url")

    encoded_string = ""
    pathdir = basepath + image_url
    with open(pathdir, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read())    
    resp = {'image': encoded_string.decode("utf-8")}
    return HttpResponse(json.dumps(resp, indent=2), content_type="application/json")

#нужно удалить
@api_view(['GET'])
def childphoto(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)

    username = request.user
    id_org = get_org_id(username)
    id_group = request.GET.get("id_group")
    iin = request.GET.get("iin")

    childobj = Childs.objects.all()
    childobj = childobj.filter(id_org = id_org, id_group = id_group, iin = iin)
    image_url = ''
    encoded_string = ""
    
    try:
        for item in childobj:
            image_url = basepath +  item.image_url

        with open(image_url, "rb") as image_file:
                encoded_string = base64.b64encode(image_file.read())    
        resp = {'image': encoded_string.decode("utf-8")}
        return HttpResponse(json.dumps(resp, indent=2), content_type="application/json")
    except:
        return HttpResponse('{"status": "error"}', content_type="application/json", status = 500)

# Сервисы, связанные с группой
# Сервис получения списка воспитательных групп для обычного пользователя
@api_view(['GET'])
def grouplist(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    
    id_org = request.GET.get("org_id")
    if (id_org == ''):
        id_org = get_org_id(request.user)
    groupname = request.GET.get("groupname")
    groupname = '%' + groupname + '%'
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""with gr as(
                    SELECT
                        groups.id_org,
                        groups.id_group,
                        groups.group_name,
                        groups.group_age,
                        groups.group_count,
                        groups.username,
                        groups.category
                    FROM
                        serviceback_groups as groups
                    WHERE
                        (NOT groups.is_delete)
                        and (groups.id_org = '{id_org}')
                        and (
                            (upper(groups.group_name) LIKE upper('{groupname}'))
                                or groups.id_org LIKE ('{groupname}')
                            )
                            ),
                users as (
                    SELECT
                        us.first_name,
                        us.username
                    FROM    
                        auth_user AS us)
                SELECT
                    gr.id_org,
                    gr.id_group,
                    gr.group_name,
                    gr.group_age,
                    gr.group_count,
                    gr.category,
                    users.first_name,
                    users.username
                FROM
                    gr as gr
                LEFT JOIN
                    users as users
                ON
                    gr.username = users.username
                ORDER BY
                    gr.group_name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    if lastelement > len(result):
            lastelement = len(result)
    res = {"list": result[firstelement:lastelement],
           "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res), content_type="application/json")

# Сервис получения списка воспитательных групп для админа
@api_view(['GET'])
def grouplistadmin(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    groupname = request.GET.get("groupname")
    groupname = '%' + groupname + '%'
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""with gr as(
                    SELECT
                        groups.*
                    FROM
                        serviceback_groups as groups
                    WHERE
                        NOT groups.is_delete
                        and (upper(groups.group_name) LIKE upper('{groupname}')
                        or groups.id_org LIKE ('{groupname}'))),
                users as (
                    SELECT
                        us.first_name,
                        us.username
                    FROM    
                        auth_user AS us),
				org as (
					SELECT
						org.org_name,
						org.id_org
					FROM
						serviceback_organizations as org
					WHERE
						org.id_org in (
									SELECT
										gr.id_org
									FROM
										gr as gr))
                SELECT
                    gr.id_org,
                    gr.id_group,
                    gr.group_name,
                    gr.group_age,
                    gr.group_count,
                    gr.category,
					org.org_name,
                    users.first_name,
                    users.username
                FROM
                    gr as gr
                LEFT JOIN
                    users as users
                ON
                    gr.username = users.username
				LEFT JOIN					
					org as org
				ON
					org.id_org = gr.id_org
                ORDER BY
					org.org_name,
                    gr.group_name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
    if lastelement > len(result):
            lastelement = len(result)
    res = {"list": result[firstelement:lastelement],
           "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res), content_type="application/json")

# Сервис редактирования таблицы "Воспитательные группы"
@api_view(['POST'])
def groupedit(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    param = request.GET.get("param")
    datastr = request.body
    res = json.loads(datastr)

    if param == 'edit':
        for item in res:
            if not request.user.is_staff:
                id_org = get_org_id(request.user)
                if not id_org == item['id_org']:
                    return HttpResponse('{"status": "Нет доступа для записи."}', content_type="application/json", status=401)
            else:
                id_org = item['id_org']

            try:
                zapisgroup = Groups.objects.get(id_group=item['id_group'], id_org=id_org)
                zapisgroup.group_name = item['group_name']
                zapisgroup.group_count = item['group_count']
                zapisgroup.group_age = item['group_age']
                zapisgroup.category = item['category']
                zapisgroup.username = item['username']
                zapisgroup.save()
            except:
                return HttpResponse('{"status": "Ошибка редактирования группы."}', content_type="application/json", status=500)

    if param == 'add':
        for item in res:
            if not request.user.is_staff:
                id_org = get_org_id(request.user)
                if not id_org == item['id_org']:
                    return HttpResponse('{"status": "Нет доступа для записи."}', content_type="application/json", status=401)
            else:
                id_org = item['id_org']

            try:
                zapisgroup = Groups.objects.get(group_name=item['group_name'], id_org=id_org)
                return HttpResponse('{"status": "Группа с таким названием уже существует."}', content_type="application/json", status=500)
            except:
                    zapisgroup = Groups()
                    zapisgroup.group_name = item['group_name']
                    zapisgroup.group_count = item['group_count']
                    zapisgroup.group_age = item['group_age']
                    zapisgroup.category = item['category']
                    zapisgroup.username = item['username']
                    zapisgroup.id_org = id_org
                    zapisgroup.save()
                    zapisgroup.id_group = zapisgroup.id
                    zapisgroup.save()

    if param == 'del':
        for item in res:
            if not request.user.is_staff:
                id_org = get_org_id(request.user)
                if not id_org == item['id_org']:
                    return HttpResponse('{"status": "Нет доступа для записи."}', content_type="application/json", status=401)
            else:
                id_org = item['id_org']

            zapischilds = Childs.objects.filter(
                id_org=id_org, id_group=item['id_group'], is_delete = False)
            flag = False
            for itemchild in zapischilds:
                flag = True
                break

            if flag:
                return HttpResponse('{"status": "Ошибка удаления. В группе зарегистрированы дети."}', content_type="application/json", status=500)
            else:
                zapisgroup = Groups.objects.get(id_group=item['id_group'])
                zapisgroup.is_delete = True
                zapisgroup.save()

    return HttpResponse('{"status": "Успешно выполнена."}', content_type="application/json", status=200)

#нужно удалить
# Сервис получения табеля для организации
@api_view(['GET'])
def gettabelbyday(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id_org = get_org_id(request.user)
    datestatus = request.GET.get("datestatus")

    query = """SELECT id_group, iin, datestatus, status FROM serviceback_visits as visits
                WHERE
                    id_org = %s and datestatus = %s
                ORDER BY id_group
                """

    with connection.cursor() as cursor:
        cursor.execute(query, [id_org, datestatus])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result, indent=4, sort_keys=True, default=str), content_type="application/json")


# Сервис для табеля Магжан
@api_view(['GET'])
def gettabelbymonth(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # id_org = get_org_id(request.user)
    id_org    = request.GET.get("id_org")
    id_group    = request.GET.get("id_group")
    datenachalo = request.GET.get("datenachalo")
    datekonec   = request.GET.get("datekonec")

    id_group = '%' + id_group + '%'

    query = f"""with weekday as(
		SELECT 
			* 
		FROM 
			serviceback_weekendday 
		WHERE 
			weekend >= '{datenachalo}' 
			and weekend <= '{datekonec}'),
        visits as(
            SELECT
                visits.id_org,
                visits.id_group,
                visits.iin,
                visits.datestatus,
                visits.timestatus,
                visits.status,
                visits.edited
            FROM
                serviceback_visits as visits
            WHERE 
                visits.id_org = '{id_org}'
                and visits.id_group like '{id_group}'
                and visits.datestatus >= '{datenachalo}'
                and visits.datestatus <= '{datekonec}'
                and visits.datestatus not in (
                                SELECT 
                                    weekend 
                                FROM 
                                    weekday)					
            ORDER BY
                datestatus),
        visit as (
            SELECT 
                visits.id_org, 
                visits.id_group, 
                visits.iin, 
                visits.datestatus, 
                max(visits.timestatus) as timestatus
            FROM 
                visits
            GROUP BY 
                visits.id_org, 
                visits.id_group, 
                visits.iin, 
                visits.datestatus),
        childs as(
            SELECT
                child.name as childname,
                child.image_url as image,
                child.iin as iin,
                child.id_group,
                child.id_org
            FROM
                serviceback_childs as child 
            WHERE
                child.iin in (
                            SELECT
                                v.iin
                            FROM
                                visit as v)),
        weekchild as (
                SELECT 
                    child.id_org, 
                    child.id_group, 
                    child.iin, 
                    w.weekend as datestatus, 
                    '7' as status
                FROM 
                    visits as child, 
                    weekday as w
                GROUP BY 
                    child.id_org, 
                    child.id_group, 
                    child.iin, 
                    w.weekend),
        un as(
                SELECT
                    vis.id_org, 
                    vis.id_group, 
                    vis.iin, 
                    vis.datestatus, 
                    vis.status, 
                    vis.edited
                FROM
                    visits as vis
                WHERE
                    (vis.id_org,
                    vis.id_group, 
                    vis.iin, 
                    vis.datestatus, 
                    vis.timestatus) in
                                (SELECT 
                                    id_org, 
                                    id_group, 
                                    iin, 
                                    datestatus, 
                                    timestatus 
                                FROM 
                                    visit)
                UNION
                SELECT
                    w.id_org, 
                    w.id_group, 
                    w.iin, 
                    w.datestatus, 
                    w.status, 
                    false as edited
                FROM
                    weekchild as w),
        gr as(
                SELECT
                    grr.group_name as grname,
                    grr.id_group as id_group
                FROM
                    serviceback_groups as grr
                WHERE 
                    (grr.id_org, grr.id_group) in (
                                    SELECT
                                        un.id_org,
                                        un.id_group
                                    FROM
                                        un)),
        vt as(
            SELECT
                    un.datestatus, 
                    un.status, 
                    un.edited,
                    un.iin as iin,
                    childs.childname,
                    childs.image,
                    childs.id_group as child_group,
                    un.id_group,
                    gr.grname
            FROM
                    un 
            LEFT JOIN 
                    gr
            ON
                    un.id_group = gr.id_group
            LEFT JOIN
                    childs 
            ON 
                    un.iin = childs.iin),
        vtt as(
                SELECT 
                    id_group, 
                    iin, 
                    child_group,
                    datestatus, 
                    status, 
                    childname, 
                    image, 
                    grname,
                    (case when edited then 20 else 10 end) as edited
                FROM 
                    vt 
                ORDER BY
                    iin, 
                    datestatus),
        allchilds as(
                SELECT
                    id_group, 
                    iin,
                    max(child_group) as child_group, 
                    max(childname) as childname, 
                    max(grname) as group_name,
                    max(replace(image, 
                            'FilesArchiv', 'https://face11.qazna24.kz/media')) as image_url,
                            max(case when date_part('day', datestatus) = 1 then cast(status as integer) + edited
                                else 0 end) as day_1,	
                            max(case when date_part('day', datestatus) = 2 then cast(status as integer) + edited
                                else 0 end) as day_2,
                            max(case when date_part('day', datestatus) = 3 then cast(status as integer) + edited
                                else 0 end) as day_3,
                            max(case when date_part('day', datestatus) = 4 then cast(status as integer) + edited
                                else 0 end) as day_4,
                            max(case when date_part('day', datestatus) = 5 then cast(status as integer) + edited
                                else 0 end) as day_5,
                            max(case when date_part('day', datestatus) = 6 then cast(status as integer) + edited
                                else 0 end) as day_6,
                            max(case when date_part('day', datestatus) = 7 then cast(status as integer) + edited
                                else 0 end) as day_7,
                            max(case when date_part('day', datestatus) = 8 then cast(status as integer) + edited
                                else 0 end) as day_8,
                            max(case when date_part('day', datestatus) = 9 then cast(status as integer) + edited
                                else 0 end) as day_9,	
                            max(case when date_part('day', datestatus) = 10 then cast(status as integer) + edited
                                else 0 end) as day_10,
                            max(case when date_part('day', datestatus) = 11 then cast(status as integer) + edited
                                else 0 end) as day_11,
                            max(case when date_part('day', datestatus) = 12 then cast(status as integer) + edited
                                else 0 end) as day_12,
                            max(case when date_part('day', datestatus) = 13 then cast(status as integer) + edited
                                else 0 end) as day_13,
                            max(case when date_part('day', datestatus) = 14 then cast(status as integer) + edited
                                else 0 end) as day_14,
                            max(case when date_part('day', datestatus) = 15 then cast(status as integer) + edited
                                else 0 end) as day_15,
                            max(case when date_part('day', datestatus) = 16 then cast(status as integer) + edited
                                else 0 end) as day_16,
                            max(case when date_part('day', datestatus) = 17 then cast(status as integer) + edited
                                else 0 end) as day_17,	
                            max(case when date_part('day', datestatus) = 18 then cast(status as integer) + edited
                                else 0 end) as day_18,
                            max(case when date_part('day', datestatus) = 19 then cast(status as integer) + edited
                                else 0 end) as day_19,
                            max(case when date_part('day', datestatus) = 20 then cast(status as integer) + edited
                                else 0 end) as day_20,
                            max(case when date_part('day', datestatus) = 21 then cast(status as integer) + edited
                                else 0 end) as day_21,
                            max(case when date_part('day', datestatus) = 22 then cast(status as integer) + edited
                                else 0 end) as day_22,
                            max(case when date_part('day', datestatus) = 23 then cast(status as integer) + edited
                                else 0 end) as day_23,
                            max(case when date_part('day', datestatus) = 24 then cast(status as integer) + edited
                                else 0 end) as day_24,
                            max(case when date_part('day', datestatus) = 25 then cast(status as integer) + edited
                                else 0 end) as day_25,	
                            max(case when date_part('day', datestatus) = 26 then cast(status as integer) + edited
                                else 0 end) as day_26,
                            max(case when date_part('day', datestatus) = 27 then cast(status as integer) + edited
                                else 0 end) as day_27,
                            max(case when date_part('day', datestatus) = 28 then cast(status as integer) + edited
                                else 0 end) as day_28,
                            max(case when date_part('day', datestatus) = 29 then cast(status as integer) + edited
                                else 0 end) as day_29,
                            max(case when date_part('day', datestatus) = 30 then cast(status as integer) + edited
                                else 0 end) as day_30,
                            max(case when date_part('day', datestatus) = 31 then cast(status as integer) + edited
                                else 0 end) as day_31 from vtt
                        GROUP BY
                            id_group,
                            iin)
                    select id_group, iin, childname, child_group, group_name,
                                    case when day_1 = 0 then 15 else day_1 end,
                                    case when day_2 = 0 then 15 else day_2 end,
                                    case when day_3 = 0 then 15 else day_3 end,
                                    case when day_4 = 0 then 15 else day_4 end,
                                    case when day_5 = 0 then 15 else day_5 end,
                                    case when day_6 = 0 then 15 else day_6 end,
                                    case when day_7 = 0 then 15 else day_7 end,
                                    case when day_8 = 0 then 15 else day_8 end,
                                    case when day_9 = 0 then 15 else day_9 end,
                                    case when day_10 = 0 then 15 else day_10 end,
                                    case when day_11 = 0 then 15 else day_11 end,
                                    case when day_12 = 0 then 15 else day_12 end,
                                    case when day_13 = 0 then 15 else day_13 end,
                                    case when day_14 = 0 then 15 else day_14 end,
                                    case when day_15 = 0 then 15 else day_15 end,
                                    case when day_16 = 0 then 15 else day_16 end,
                                    case when day_17 = 0 then 15 else day_17 end,
                                    case when day_18 = 0 then 15 else day_18 end,
                                    case when day_19 = 0 then 15 else day_19 end,
                                    case when day_20 = 0 then 15 else day_20 end,
                                    case when day_21 = 0 then 15 else day_21 end,
                                    case when day_22 = 0 then 15 else day_22 end,
                                    case when day_23 = 0 then 15 else day_23 end,
                                    case when day_24 = 0 then 15 else day_24 end,
                                    case when day_25 = 0 then 15 else day_25 end,
                                    case when day_26 = 0 then 15 else day_26 end,
                                    case when day_27 = 0 then 15 else day_27 end,
                                    case when day_28 = 0 then 15 else day_28 end,
                                    case when day_29 = 0 then 15 else day_29 end,
                                    case when day_30 = 0 then 15 else day_30 end,
                                    case when day_31 = 0 then 15 else day_31 end
                    FROM allchilds
                    ORDER BY group_name, childname"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result, indent=4, sort_keys=True, default=str), content_type="application/json")

#нужно удалить
@api_view(['GET'])
def gettabelbym(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # id_org = get_org_id(request.user)
    id_org    = request.GET.get("id_org")
    id_group    = request.GET.get("id_group")
    datenachalo = request.GET.get("datenachalo")
    datekonec   = request.GET.get("datekonec")

    id_group = '%' + id_group + '%'

    query = f"""with weekday as(
                    SELECT 
                        * 
                    FROM 
                        serviceback_weekendday 
                    WHERE
                        weekend >= '{datenachalo}'
                        and weekend <= '{datekonec}'),
	 visits as(
                    SELECT
                        visits.id_org,
		 				visits.id_group,
		 				visits.iin,
		 				visits.datestatus,
		 				visits.timestatus,
		 				visits.status,
		 				visits.edited
                    FROM
                        serviceback_visits as visits
                    WHERE 
                        visits.id_org = '{id_org}'
						and visits.id_group like '{id_group}'
                        and visits.datestatus >= '{datenachalo}'
                        and visits.datestatus <= '{datekonec}'
						and visits.datestatus not in (SELECT weekend FROM weekday)
					
					ORDER BY
						datestatus),
	visit as (SELECT visits.id_org, visits.id_group, visits.iin, visits.datestatus, max(visits.timestatus) as timestatus
			 	FROM visits
			 	GROUP BY visits.id_org, visits.id_group, visits.iin, visits.datestatus),
weekchild as (SELECT vis.id_org, vis.id_group, vis.iin, w.weekend as datestatus, '7' as status
				FROM visit as vis, weekday as w
				GROUP BY vis.id_org, vis.id_group, vis.iin, w.weekend),
un as(
			 select vis.id_org, vis.id_group, vis.iin, vis.datestatus, vis.status, vis.edited
				from visits as vis
				WHERE (vis.id_org, vis.id_group, vis.iin, vis.datestatus, vis.timestatus) in
						(SELECT id_org, id_group, iin, datestatus, timestatus FROM visit)
UNION
			select w.id_org, w.id_group, w.iin, w.datestatus, w.status, false as edited
				from weekchild as w),
childs as(
                    SELECT
                        child.name as childname,
                        child.image_url as image,
						child.iin as iin
                    FROM
                        serviceback_childs as child 
                    WHERE
                        child.id_org = '{id_org}'
                        and child.iin in (
										SELECT
											iin
										FROM
											un)),
gr as(
                    SELECT
                        grr.group_name as grname,
						grr.id_group as id_group
                    FROM
                        serviceback_groups as grr
                    WHERE grr.id_org = '{id_org}'
                        and grr.id_group in (
										SELECT
											un.id_group
										FROM
											un)),
vt as(
                    SELECT
                        un.*,
                        childs.childname,
                        childs.image,
                        gr.grname
                    FROM
                        un 
                    LEFT JOIN gr
                    ON
                        un.id_group = gr.id_group
                    JOIN
                        childs on un.iin = childs.iin),
vtt as(SELECT id_group, iin, datestatus, status, childname, image, grname,
                        (case when edited then 20
							 else 10 end) as edited
FROM vt order by iin, datestatus)
SELECT
                    id_group, 
                    iin, 
                    max(childname) as name, 
                    max(grname) as group_name,
                    max(replace(image, 
                    'FilesArchiv', 'https://face11.qazna24.kz/media')) as image_url,
					max(case when date_part('day', datestatus) = 1 then cast(status as integer) + edited
                        else 0 end) as day_1,	
                    max(case when date_part('day', datestatus) = 2 then cast(status as integer) + edited
                        else 0 end) as day_2,
                    max(case when date_part('day', datestatus) = 3 then cast(status as integer) + edited
                        else 0 end) as day_3,
                    max(case when date_part('day', datestatus) = 4 then cast(status as integer) + edited
                        else 0 end) as day_4,
                    max(case when date_part('day', datestatus) = 5 then cast(status as integer) + edited
                        else 0 end) as day_5,
                    max(case when date_part('day', datestatus) = 6 then cast(status as integer) + edited
                        else 0 end) as day_6,
                    max(case when date_part('day', datestatus) = 7 then cast(status as integer) + edited
                        else 0 end) as day_7,
                    max(case when date_part('day', datestatus) = 8 then cast(status as integer) + edited
                        else 0 end) as day_8,
                    max(case when date_part('day', datestatus) = 9 then cast(status as integer) + edited
                        else 0 end) as day_9,	
                    max(case when date_part('day', datestatus) = 10 then cast(status as integer) + edited
                        else 0 end) as day_10,
                    max(case when date_part('day', datestatus) = 11 then cast(status as integer) + edited
                        else 0 end) as day_11,
                    max(case when date_part('day', datestatus) = 12 then cast(status as integer) + edited
                        else 0 end) as day_12,
                    max(case when date_part('day', datestatus) = 13 then cast(status as integer) + edited
                        else 0 end) as day_13,
                    max(case when date_part('day', datestatus) = 14 then cast(status as integer) + edited
                        else 0 end) as day_14,
                    max(case when date_part('day', datestatus) = 15 then cast(status as integer) + edited
                        else 0 end) as day_15,
                    max(case when date_part('day', datestatus) = 16 then cast(status as integer) + edited
                        else 0 end) as day_16,
                    max(case when date_part('day', datestatus) = 17 then cast(status as integer) + edited
                        else 0 end) as day_17,	
                    max(case when date_part('day', datestatus) = 18 then cast(status as integer) + edited
                        else 0 end) as day_18,
                    max(case when date_part('day', datestatus) = 19 then cast(status as integer) + edited
                        else 0 end) as day_19,
                    max(case when date_part('day', datestatus) = 20 then cast(status as integer) + edited
                        else 0 end) as day_20,
                    max(case when date_part('day', datestatus) = 21 then cast(status as integer) + edited
                        else 0 end) as day_21,
                    max(case when date_part('day', datestatus) = 22 then cast(status as integer) + edited
                        else 0 end) as day_22,
                    max(case when date_part('day', datestatus) = 23 then cast(status as integer) + edited
                        else 0 end) as day_23,
                    max(case when date_part('day', datestatus) = 24 then cast(status as integer) + edited
                        else 0 end) as day_24,
                    max(case when date_part('day', datestatus) = 25 then cast(status as integer) + edited
                        else 0 end) as day_25,	
                    max(case when date_part('day', datestatus) = 26 then cast(status as integer) + edited
                        else 0 end) as day_26,
                    max(case when date_part('day', datestatus) = 27 then cast(status as integer) + edited
                        else 0 end) as day_27,
                    max(case when date_part('day', datestatus) = 28 then cast(status as integer) + edited
                        else 0 end) as day_28,
                    max(case when date_part('day', datestatus) = 29 then cast(status as integer) + edited
                        else 0 end) as day_29,
                    max(case when date_part('day', datestatus) = 30 then cast(status as integer) + edited
                        else 0 end) as day_30,
                    max(case when date_part('day', datestatus) = 31 then cast(status as integer) + edited
                        else 0 end) as day_31 from vtt
                GROUP BY
                    id_group,
                    iin
                ORDER BY
                    group_name,
                    name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result, indent=4, sort_keys=True, default=str), content_type="application/json")    

#нужно удалить
# Сервис для табеля Магжан
@api_view(['GET'])
def gettabel(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id_org = get_org_id(request.user)
    id_group    = request.GET.get("id_group")
    datenachalo = request.GET.get("datenachalo")
    datekonec   = request.GET.get("datekonec")

    if id_group=="":
        id_group = '%' + id_group + '%'

    query = f"""with visit as(
                    SELECT
                        visits.*
                    FROM
                        serviceback_visits as visits
                    WHERE 
                        visits.id_org = '{id_org}'
                        and visits.datestatus >= '{datenachalo}'
                        and visits.datestatus <= '{datekonec}'
                        and visits.id_group like '{id_group}'),
                childs as(
                    SELECT
                        child.*
                    FROM
                        serviceback_childs as child 
                    WHERE
                        child.id_org = '{id_org}'
                        and child.is_delete = false),
                gr as(
                    SELECT
                        grr.*
                    FROM
                        serviceback_groups as grr
                    WHERE grr.id_org = '{id_org}'
                        and grr.is_delete = false),
                vt as(
                    SELECT
                        visit.*,
                        childs.name as childname,
                        childs.image_url as image,
                        gr.group_name as grname
                    FROM
                        visit 
                    LEFT JOIN gr
                    ON
                        visit.id_group = gr.id_group
                    JOIN
                        childs on visit.iin = childs.iin
                        and visit.id_group = childs.id_group),
                vtr as(
                    SELECT
                        id_group as idgr, 
                        iin as iin, 
                        datestatus as datestatus, 
                        max(timestatus) as timestatus 
                    FROM
                        vt
                    GROUP BY
                        id_group,
                        iin,
                        datestatus),
                vtt as(
                    SELECT
                        id_group,
                        iin, 
                        datestatus, 
                        timestatus,
                        status, 
                        childname, 
                        image, 
                        grname
                    FROM
                        vt
                    WHERE
                        (id_group, iin, datestatus, timestatus) in (select idgr, iin, datestatus, timestatus from vtr))
                
                SELECT
                    id_group, 
                    iin, 
                    max(childname) as name, 
                    max(grname) as group_name, 
                    max(replace(image, 
                    'FilesArchiv', 'https://face11.qazna24.kz/media')) as image_url,
                    max(case when date_part('day', datestatus) = 1 then cast(status as integer)
                        else 0 end) as day_1,	
                    max(case when date_part('day', datestatus) = 2 then cast(status as integer)
                        else 0 end) as day_2,
                    max(case when date_part('day', datestatus) = 3 then cast(status as integer)
                        else 0 end) as day_3,
                    max(case when date_part('day', datestatus) = 4 then cast(status as integer)
                        else 0 end) as day_4,
                    max(case when date_part('day', datestatus) = 5 then cast(status as integer)
                        else 0 end) as day_5,
                    max(case when date_part('day', datestatus) = 6 then cast(status as integer)
                        else 0 end) as day_6,
                    max(case when date_part('day', datestatus) = 7 then cast(status as integer)
                        else 0 end) as day_7,
                    max(case when date_part('day', datestatus) = 8 then cast(status as integer)
                        else 0 end) as day_8,
                    max(case when date_part('day', datestatus) = 9 then cast(status as integer)
                        else 0 end) as day_9,	
                    max(case when date_part('day', datestatus) = 10 then cast(status as integer)
                        else 0 end) as day_10,
                    max(case when date_part('day', datestatus) = 11 then cast(status as integer)
                        else 0 end) as day_11,
                    max(case when date_part('day', datestatus) = 12 then cast(status as integer)
                        else 0 end) as day_12,
                    max(case when date_part('day', datestatus) = 13 then cast(status as integer)
                        else 0 end) as day_13,
                    max(case when date_part('day', datestatus) = 14 then cast(status as integer)
                        else 0 end) as day_14,
                    max(case when date_part('day', datestatus) = 15 then cast(status as integer)
                        else 0 end) as day_15,
                    max(case when date_part('day', datestatus) = 16 then cast(status as integer)
                        else 0 end) as day_16,
                    max(case when date_part('day', datestatus) = 17 then cast(status as integer)
                        else 0 end) as day_17,	
                    max(case when date_part('day', datestatus) = 18 then cast(status as integer)
                        else 0 end) as day_18,
                    max(case when date_part('day', datestatus) = 19 then cast(status as integer)
                        else 0 end) as day_19,
                    max(case when date_part('day', datestatus) = 20 then cast(status as integer)
                        else 0 end) as day_20,
                    max(case when date_part('day', datestatus) = 21 then cast(status as integer)
                        else 0 end) as day_21,
                    max(case when date_part('day', datestatus) = 22 then cast(status as integer)
                        else 0 end) as day_22,
                    max(case when date_part('day', datestatus) = 23 then cast(status as integer)
                        else 0 end) as day_23,
                    max(case when date_part('day', datestatus) = 24 then cast(status as integer)
                        else 0 end) as day_24,
                    max(case when date_part('day', datestatus) = 25 then cast(status as integer)
                        else 0 end) as day_25,	
                    max(case when date_part('day', datestatus) = 26 then cast(status as integer)
                        else 0 end) as day_26,
                    max(case when date_part('day', datestatus) = 27 then cast(status as integer)
                        else 0 end) as day_27,
                    max(case when date_part('day', datestatus) = 28 then cast(status as integer)
                        else 0 end) as day_28,
                    max(case when date_part('day', datestatus) = 29 then cast(status as integer)
                        else 0 end) as day_29,
                    max(case when date_part('day', datestatus) = 30 then cast(status as integer)
                        else 0 end) as day_30,
                    max(case when date_part('day', datestatus) = 31 then cast(status as integer)
                        else 0 end) as day_31 from vtt
                GROUP BY
                    id_group,
                    iin
                ORDER BY
                    group_name,
                    name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(list(result), content_type="application/json")    

# Сервис для изменения статуса ребенка
@api_view(['POST'])
def editvisit(request):
    username = request.user
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
        
    datastr = request.body
    res = json.loads(datastr)

    for item in res:
        if (item['id_org'] == '' or item['id_group'] == '' or
            item['iin'] == '' or item['datestatus'] == '' or 
            (item['status'] != '3' and item['status'] != '4' and item['status'] != '5' and item['status'] != '10')):
            return HttpResponse('{"status": "Ошибка при редактирования данных ребенка!"}', content_type="application/json", status=500)

        date_obj = datetime.datetime.strptime(item['datestatus'], '%d.%m.%Y')
        formatted_date = date_obj.strftime('%Y-%m-%d')
        iin = item['iin']
        id_org = item['id_org']
        status = item['status']

        if item['status'] == '4':
            current_date = datetime.date.today()
            
            query = f"""
                    with visit as(
                            SELECT 
                                id as _id,
                                iin,
                                datestatus,
                                status
                            FROM 
                                public.serviceback_visits
                            WHERE 
                                iin='{iin}'
                                and datestatus BETWEEN date_trunc('year', '{current_date}'::date) and '{current_date}'
                            ),
                        max_vis_id as(
                            SELECT
                                iin,
                                datestatus,
                                max(_id) as _id
                            FROM
                                visit
                            GROUP BY
                                datestatus,
                                iin),
                        readyvis as (
                            SELECT
                                count(*) as count_rec
                            FROM
                                visit
                            where (_id) in (
                                    SELECT 
                                        _id
                                    FROM 
                                        max_vis_id)
                                and status = '4')
                            SELECT 
                                * 
                            FROM 
                                readyvis"""
            with connection.cursor() as cursor:
                cursor.execute(query)
                connection.commit()
                result = cursor.fetchone()
                if len(result) > 0:
                    if result[0] > 60:
                        return HttpResponse('{"status": "Превышен лимит отпуска на этот год!"}', content_type="application/json", status=400)

        try:
            max_timestatus = Visits.objects.filter(id_org = id_org, id_group = item['id_group'], iin = iin, datestatus = formatted_date).aggregate(Max('timestatus'))['timestatus__max']
            current_time = datetime.datetime.now().time()
            time_str = current_time.strftime("%H:%M:%S")
            if max_timestatus:
                latest_visit = Visits.objects.filter(id_org = id_org, id_group = item['id_group'], iin = iin, datestatus = formatted_date, timestatus=max_timestatus).first()

                if latest_visit:
                    latest_status = latest_visit.status
                    
                    if not datetime.datetime.now().strftime('%Y-%m-%d') == date_obj:                

                        query = f"""
                            with visit as(
                                    SELECT 
                                        id as _id,
                                        id_org,
                                        iin,
                                        status
                                    FROM 
                                        public.serviceback_visits
                                    WHERE 
                                        iin='{iin}'
                                        and datestatus = '{formatted_date}'
                                        and id_org = '{id_org}'
                                    ),
                                max_vis_id as(
                                    SELECT
                                        iin,
                                        max(_id) as _id
                                    FROM
                                        visit
                                    GROUP BY
                                        iin),
                                readyvis as (
                                    SELECT
                                        *
                                    FROM
                                        visit
                                    where (_id) in (
                                            SELECT 
                                                _id
                                            FROM 
                                                max_vis_id)),
                                itogbyday as(
                                    SELECT 
                                        * 
                                    FROM 
                                        serviceback_itogbyday
                                    WHERE
                                        id_org = '{id_org}'
                                        and datestatus= '{formatted_date}'),
                                edititogbyday as(
                                    SELECT
                                        it.datestatus,
                                        it.id_org,
                                        CASE WHEN vis.status = '2' then
                                            it.visit - 1
                                            WHEN '{status}' = '2' THEN
                                            it.visit + 1
                                            ELSE it.visit
                                        END as visit,
                                        CASE WHEN vis.status = '3' then
                                                it.boln - 1
                                            WHEN '{status}' = '3' THEN
                                                it.boln + 1
                                            ELSE it.boln
                                        END as boln,
                                        CASE WHEN vis.status = '4' then
                                            it.otpusk - 1
                                            WHEN '{status}' = '4' THEN
                                            it.otpusk + 1
                                            ELSE it.otpusk
                                        END as otpusk,
                                        CASE WHEN vis.status = '5' or vis.status = '10' then
                                            it.notvisit - 1
                                            WHEN '{status}' = '5' or '{status}' = '10' THEN
                                            it.notvisit + 1
                                            ELSE it.notvisit
                                        END as notvisit
                                    FROM
                                        itogbyday as it
                                    LEFT JOIN
                                        readyvis as vis
                                    ON
                                        it.id_org = vis.id_org),
                                updateday as(
                                    update
                                        serviceback_itogbyday as it
                                    SET
                                        visit = ed.visit,
                                        boln = ed.boln,
                                        otpusk = ed.otpusk,
                                        notvisit = ed.notvisit
                                    FROM
                                        edititogbyday as ed
                                    WHERE
                                        it.id_org = '{id_org}'
                                        and it.datestatus = '{formatted_date}'),
                                itogbymonth as(
                                    SELECT 
                                        * 
                                    FROM 
                                        serviceback_itogbymonth
                                    WHERE
                                        id_org = '{id_org}'
                                        and datestatus=date_trunc('month', '{formatted_date}'::date)),
                                editmonth as(
                                    SELECT
                                        it.datestatus,
                                        it.id_org,
                                        CASE WHEN vis.status = '2' then
                                            it.visit - 1
                                            WHEN '{status}' = '2' THEN
                                            it.visit + 1
                                            ELSE it.visit
                                        END as visit,
                                        CASE WHEN vis.status = '3' then
                                                it.boln - 1
                                            WHEN '{status}' = '3' THEN
                                                it.boln + 1
                                            ELSE it.boln
                                        END as boln,
                                        CASE WHEN vis.status = '4' then
                                            it.otpusk - 1
                                            WHEN '{status}' = '4' THEN
                                            it.otpusk + 1
                                            ELSE it.otpusk
                                        END as otpusk,
                                        CASE WHEN vis.status = '5' or vis.status = '10'  then
                                            it.notvisit - 1
                                            WHEN '{status}' = '5' or '{status}' = '10' THEN
                                            it.notvisit + 1
                                            ELSE it.notvisit
                                        END as notvisit
                                    FROM
                                        itogbymonth as it
                                    LEFT JOIN
                                        readyvis as vis
                                    ON
                                        it.id_org = vis.id_org),
                                updatemonth as(
                                    update
                                        serviceback_itogbymonth as it
                                    SET
                                        visit = ed.visit,
                                        boln = ed.boln,
                                        otpusk = ed.otpusk,
                                        notvisit = ed.notvisit
                                    FROM
                                        editmonth as ed
                                    WHERE
                                        it.id_org = '{id_org}'
                                        and it.datestatus=date_trunc('month', '{formatted_date}'::date))
                                SELECT * FROM itogbyday"""
                        
                        with connection.cursor() as cursor:
                            cursor.execute(query)
                            connection.commit()

                    latest_visit.status = item['status']
                    latest_visit.edited = True
                    latest_visit.username = str(username)
                    latest_visit.timestatus = time_str
                    latest_visit.comments = 'проверен ' + str(username)
                    latest_visit.save()
            else:
                crVisit = Visits.objects.create(id_org = id_org, id_group = item['id_group'],
                                                 iin = iin, datestatus = formatted_date, username = str(username),
                                                 status = item['status'], timestatus = time_str, comments = 'проверен ' + str(username),
                                                  edited = True)

            return HttpResponse('{"status": "Данные ребенка успешно изменены!"}', content_type="application/json", status=200)
        except:
            return HttpResponse('{"status": "Ошибка при редактирования данных ребенка!"}', content_type="application/json", status=500)

# Сервис для посещения ребенка
@api_view(['POST'])
def deletevisit(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    if not request.user.is_staff:
        return HttpResponse('{"status": "Нет доступа к редактированию."}', content_type="application/json", status=401)

    datastr = request.body
    res = json.loads(datastr)

    for item in res:
        if (item['id_org'] == '' or item['id_group'] == '' or
            item['iin'] == '' or item['datestatus'] == ''):
            return HttpResponse('{"status": "Ошибка при редактирования данных ребенка!"}', content_type="application/json", status=500)

        date_obj = datetime.datetime.strptime(item['datestatus'], '%d.%m.%Y')
        formatted_date = date_obj.strftime('%Y-%m-%d')

        try:
            vis = Visits.objects.filter(id_org = item['id_org'], id_group = item['id_group'], iin = item['iin'], datestatus = formatted_date)
            errorfake = False
            for itm in vis:
                if itm.status == '10':
                    errorfake = True
            if not errorfake:
                vis.delete()
                desc = Descriptors.objects.filter(iin = item['iin'], create_date = formatted_date)
                desc.delete()
                return HttpResponse('{"status": "Данные ребенка успешно изменены!"}', content_type="application/json", status=200)
            else:
                return HttpResponse('{"status": "У данного ребенка ложное фото!"}', content_type="application/json", status=400)
        except:
            return HttpResponse('{"status": "Ошибка при редактирования данных ребенка!"}', content_type="application/json", status=500)

# Сервисы, связанные с ребенком
# Сервис получения списка детей в группе для обычных пользователей
@api_view(['GET'])
def childlist(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    if request.user.is_staff:
        id_org = request.GET.get("id_org")
    else:    
        id_org = get_org_id(request.user)

    childname = request.GET.get("childname")
    id_group = request.GET.get("id_group")
    childname = '%' + childname + '%'
    id_group = '%' + id_group + '%'
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""With childs as(
                    SELECT 
                        childs.iin, 
                        childs.name, 
                        childs.id_group,
                        childs.id_org,
                        replace(childs.icon_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as icon_url
                    FROM 
                        serviceback_childs as childs
                    WHERE 
                        (NOT childs.is_delete) 
                        and (childs.id_org LIKE '{id_org}') 
                        and (childs.id_group LIKE '{id_group}')
                        and (upper(name) LIKE upper('{childname}')
                        or (childs.iin LIKE '{childname}'))),
                gr as(
                    SELECT 
                        groups.id_group,
                        groups.id_org,
                        groups.group_name                        
                    FROM
                        serviceback_groups as groups
                    WHERE
                        (groups.id_org, groups.id_group) in (
                                                    SELECT 
                                                        id_org,
                                                        id_group
                                                    FROM
                                                        childs)
                        and (NOT groups.is_delete))
                    SELECT
                        childs.iin, 
                        childs.name, 
                        childs.id_group,
                        childs.id_org,
                        childs.icon_url,
                        gr.group_name
                    FROM
                        childs as childs
                    LEFT JOIN
                        gr
                    ON 
                        childs.id_group = gr.id_group 
                        and childs.id_org = gr.id_org
                    ORDER BY
                        gr.group_name,
                        childs.name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(result):
            lastelement = len(result)
        res = {"list": result[firstelement:lastelement],
               "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res), content_type="application/json")

# Сервис получения списка детей в группе для админа
@api_view(['GET'])
def childlistadmin(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # iin = request.GET.get("iin")
    childname = request.GET.get("childname")

    # iin = '%' + iin + '%'
    childname = '%' + childname + '%'
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""SELECT 
                    childs.iin, 
                    childs.name,
					childs.id_org,
                    childs.id_group,
                    replace(childs.icon_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as icon_url,
                    groups.group_name,
					org.org_name
                FROM 
                    serviceback_childs as childs
                LEFT JOIN 
                    serviceback_groups as groups
                ON 
                    childs.id_group = groups.id_group 
                    and childs.id_org = groups.id_org
				LEFT JOIN 
                    serviceback_organizations as org
                ON 
                    org.id_org = groups.id_org
                WHERE 
                    (NOT childs.is_delete)
					and (childs.iin LIKE '{childname}'
                        or upper(name) LIKE upper('{childname}') 
                        or upper(groups.group_name) LIKE upper('{childname}')
                        or upper(org.org_name) LIKE upper('{childname}')
                        or childs.id_org LIKE upper('{childname}'))
                    and (NOT groups.is_delete)
                ORDER BY
					childs.name"""

    with connection.cursor() as cursor:
        cursor.execute(
            query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(result):
            lastelement = len(result)
        res = {"list": result[firstelement:lastelement],
               "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res), content_type="application/json")

# Сервис получения списка детей в группе
@api_view(['GET'])
def childselect(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    if request.user.is_staff:
        id_org = request.GET.get("id_org")
    else:    
        id_org = get_org_id(request.user)
    id_group = request.GET.get("id_group")
    childname = request.GET.get("childname")
    id_group = '%' + id_group + '%'
    childname = '%' + childname + '%'
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""SELECT 
                    childs.iin, 
                    childs.name, 
                    childs.id_group,
                    replace(childs.icon_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as icon_url,
                    groups.group_name
                FROM 
                    serviceback_childs as childs
                LEFT JOIN 
                    serviceback_groups as groups
                ON 
                    childs.id_group = groups.id_group 
                    and childs.id_org = groups.id_org
                WHERE 
                    (NOT childs.is_delete) 
                    and (childs.id_org = '{id_org}')
                    and (upper(childs.name) LIKE upper('{childname}'))
                    and (childs.id_group NOT LIKE '{id_group}')
                    and (NOT groups.is_delete)
                ORDER BY
                    groups.group_name,
                    childs.name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(result):
            lastelement = len(result)
        res = {"list": result[firstelement:lastelement],
               "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res), content_type="application/json")


# Сервис получения полной информации о ребенке
@api_view(['GET'])
def getchildbyinn(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    iin = request.GET.get("iin")

    query = f"""with ch as(
                    SELECT
                        childs.iin, 
                        childs.name,
		                childs.id_org,
                        to_char(childs.birthday,'dd.mm.yyyy') as birthday, 
                        childs.gender, 
                        childs.id_group,
                        childs.registered,
                        childs.category,                         
                        replace(childs.image_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as image_url
                    FROM 
                        serviceback_childs as childs
		 		    WHERE
                        (NOT childs.is_delete)
                        and (childs.iin = '{iin}')),
                org as(
                    SELECT
                        org.org_name,
                        org.id_org
                    FROM
                        serviceback_organizations as org
                    WHERE
                        org.id_org in(
                                SELECT
                                    id_org
                                FROM
                                    ch
                        )),
	            gr as(
                    SELECT
		 		        gr.id_org,
                        gr.id_group,
                        gr.group_name 
		            FROM
		                serviceback_groups as gr
		            WHERE
                        (NOT gr.is_delete)
                        and (gr.id_group in (SELECT 
                                                id_group
                                            FROM
                                                ch)))
	            SELECT
		            ch.*,
		            gr.group_name,
                    org.org_name
	            FROM 
                    ch as ch
                LEFT JOIN
                    gr as gr
                ON
                    ch.id_group = gr.id_group
                    and ch.id_org = gr.id_org
                LEFT JOIN
                    org as org
                ON
                    ch.id_org = org.id_org"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                    for row in cursor.fetchall()]
    response = {"data": result}
    return HttpResponse(json.dumps(response, indent=2), content_type="application/json")

# сервис для получения фото визита ребенка
@api_view(['GET'])
def childstatusbyiin(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status = 401)

    id_group = request.GET.get("id_group")
    id_org = request.GET.get("id_org")    
    status = request.GET.get("status")
    datestatus = request.GET.get("datestatus")
    iin = request.GET.get("iin")
    
    query = f"""
                with ch as(
                    SELECT
                        child.iin,
                        child.name,
                        replace(child.image_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as image_url
                    FROM
                        serviceback_childs as child
                    WHERE
                        child.iin = '{iin}'),
                    vis as(
                    SELECT
                        vis.iin,
                        replace(vis.image_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as visit_photo
                    FROM
                        serviceback_visits as vis
                    WHERE
                        vis.datestatus = '{datestatus}'
                        and vis.id_org = '{id_org}'
                        and vis.id_group = '{id_group}'
                        and vis.status = '{status}'
                        and vis.iin = '{iin}'
                    ORDER BY
                        vis.id DESC
                    LIMIT 1)
                    SELECT
                        child.*,
                        vis.visit_photo
                    FROM
                        ch as child
                    LEFT JOIN
                        vis as vis
                    ON
                        child.iin = vis.iin"""
    
    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
               for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result), content_type="application/json")    

# Сервис редактирования таблицы "Дети"
@api_view(['POST'])
def childedit(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    
    param = request.GET.get("param")
    id_group = request.GET.get("id_group")
    datastr = request.body
    res = json.loads(datastr)

    if param == 'edit':
        for item in res:
            if request.user.is_staff:
                id_org = item['id_org']
            else:
                id_org = get_org_id(request.user)
            try:
                zapischild = Childs.objects.all()
                zapischild = zapischild.filter(id_org=id_org, id_group=id_group, iin=item['iin'], is_delete=False)
                for zap in zapischild:
                    zap.iin = item['iin']
                    zap.name = item['name']
                    birthday = item['birthday']
                    mass = birthday.split('.')
                    datarojd = datetime.date(
                        int(mass[2]), int(mass[1]), int(mass[0]))
                    zap.birthday = datarojd
                    zap.gender = item['gender']
                    zap.category = item['category']
                    zap.id_group = item['id_group']
                    zap.save()
            except:
                return HttpResponse('{"status": "Ошибка редактирования ребёнка."}', content_type="application/json", status=500)

    if param == 'add':
        for item in res:
            if request.user.is_staff:
                id_org = item['id_org']
            else:
                id_org = get_org_id(request.user)

            check = Childs.objects.filter(id_org=id_org, is_delete=False, iin=item['iin'])
            flagcheck = True
            for itemcheck in check:
                flagcheck = False
            if not flagcheck:
                return HttpResponse('{"status": "Ошибка добавления. Данный ребенок уже зарегистрирован."}', content_type="application/json", status=500)

            zapischild = Childs()
            zapischild.iin = item['iin']
            zapischild.name = item['name']
            zapischild.gender = item['gender']
            zap.category = item['category']
            birthday = item['birthday']
            mass = birthday.split('.')
            datarojd = datetime.date(int(mass[2]), int(mass[1]), int(mass[0]))
            zapischild.birthday = datarojd
            zapischild.id_org = id_org
            zapischild.id_group = item['id_group']
            zapischild.registered = False
            zapischild.is_delete = False
            zapischild.save()

    if param == 'del':
        for item in res:
            if request.user.is_staff:
                id_org = item['id_org']
            else:
                id_org = get_org_id(request.user)

            zapischild = Childs.objects.all()
            zapischild = zapischild.filter(id_org=id_org, iin=item['iin'], is_delete=False)
            for checkitem in zapischild:
                checkitem.is_delete = True
                checkitem.save()

    if param == 'clearface':
        for item in res:
            zapischild = Childs.objects.all()
            zapischild = zapischild.filter(id_org=item['id_org'], iin=item['iin'], is_delete=False)
            for checkitem in zapischild:
                checkitem.registered = False
                checkitem.id_face = ""
                checkitem.save()

            descs = Descriptors.objects.filter(iin = item['iin'])
            for itemdesc in descs:
                itemdesc.delete()

    return HttpResponse('{"status": "Успешно."}', content_type="application/json", status=200)

#Сервисы, связанные с иерархией организаций
# Сервис получения списка иерархии организации
@api_view(['GET'])
def suborg(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    haveorg = request.GET.get("haveorg")
    if haveorg == 'true':
        id_org = request.GET.get("id_org")
    else:
        usrobj = ProfileUser.objects.filter(name=request.user.username)
        for itemorg in usrobj:
            id_org = itemorg.id_org
    
    org_name = request.GET.get("org_name")
    org_name = '%' + org_name + '%'
    id_region = request.GET.get("id_region")
    if (id_region is None):
        id_region = 0
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""with suborg as(
                            SELECT
                                *
                            FROM
                                public.serviceback_suborganizations
                            WHERE
                                id_parent = '{id_org}'),
                    org as(
                            SELECT
                                id_org,
                                org_name,
                                id_obl,
                                id_region
                            FROM
                                public.serviceback_organizations
                            WHERE
                                id_org in(
                                    SELECT
                                        id_child
                                    FROM
                                        suborg)
                                and (upper(org_name) LIKE upper('{org_name}')
                                    or id_org like ('{org_name}'))
                                and case when '{id_region}' = 0 then true 
                                    else id_region = ('{id_region}') end),
					allreg as(
							SELECT
								*
							FROM
								serviceback_regions),
					namereg as(
							SELECT
								id,
								name
							FROM
								allreg
							WHERE
								id in(
									SELECT
										id_region
									FROM
										org)),
					nameobl as(
							SELECT
								id,
								name
							FROM
								allreg
							WHERE
								id in(
									SELECT
										id_obl
									FROM
										org))
                            SELECT
                                suborg.id,
                                suborg.id_parent,
                                suborg.id_child as id_org,
                                org.org_name,
                                org.id_obl,
                                org.id_region,
								namereg.name as name_region,
								nameobl.name as name_obl
                            FROM
                                suborg
                            INNER JOIN
                                org
                            ON
                                suborg.id_child = org.id_org
							LEFT JOIN
								namereg
							ON
								org.id_region = namereg.id
							LEFT JOIN
								nameobl
							ON
								org.id_obl = nameobl.id
                            ORDER BY
                                org.org_name"""

    with connection.cursor() as cursor:
        cursor.execute(
            query)
        columns = [col[0] for col in cursor.description]
        suborg = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(suborg):
            lastelement = len(suborg)
        res = {"list": suborg[firstelement:lastelement],
               "last": lastelement, "total": len(suborg)}

    return HttpResponse(json.dumps(res), content_type="application/json")


# Сервис получения списка неподчиненных организации
@api_view(['GET'])
def nonsuborg(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id_org = request.GET.get("id_org")
    org_name = request.GET.get("org_name")
    org_name = '%' + org_name + '%'
    id_region = request.GET.get("id_region")
    if (id_region is None):
        id_region = 0
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""with suborg as(
                            SELECT
                                *
                            FROM
                                public.serviceback_suborganizations
                            WHERE
                                id_parent = '{id_org}'),
                    org as(
                            SELECT
                                id_org,
                                org_name,
                                id_obl,
                                id_region
                            FROM
                                public.serviceback_organizations
                            WHERE
                                id_org not in(
                                    SELECT
                                        id_child
                                    FROM
                                        suborg)
                                and (upper(org_name) LIKE upper('{org_name}')
                                or id_org LIKE '{org_name}')
                                and case when '{id_region}' = 0 then true 
                                    else id_region = ('{id_region}') end),
					allreg as(
							SELECT
								*
							FROM
								serviceback_regions),
					namereg as(
							SELECT
								id,
								name
							FROM
								allreg
							WHERE
								id in(
									SELECT
										id_region
									FROM
										org)),
					nameobl as(
							SELECT
								id,
								name
							FROM
								allreg
							WHERE
								id in(
									SELECT
										id_obl
									FROM
										org))
                            SELECT
                                org.id_org,
                                org.org_name,
                                org.id_obl,
                                org.id_region,
								namereg.name as name_region,
								nameobl.name as name_obl
                            FROM
                                org
                            LEFT JOIN
								namereg
							ON
								org.id_region = namereg.id
							LEFT JOIN
								nameobl
							ON
								org.id_obl = nameobl.id"""

    with connection.cursor() as cursor:
        cursor.execute(
            query)
        columns = [col[0] for col in cursor.description]
        suborg = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(suborg):
            lastelement = len(suborg)
        res = {"list": suborg[firstelement:lastelement],
               "last": lastelement, "total": len(suborg)}

    return HttpResponse(json.dumps(res), content_type="application/json")


# Сервис добавления в список иерархии
@api_view(['GET'])
def addsuborg(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id_parent = request.GET.get("id_parent")
    id_child = request.GET.get("id_child")

    try:
        sub = SubOrganizations.objects.create(id_parent = id_parent, id_child = id_child)
        return HttpResponse('{"status": "Организация добавлена!"}', content_type="application/json", status=200)
    except:
        return HttpResponse('{"status": "Не удалось добавить организацию!"}', content_type="application/json", status=500)


# Сервис удаления из иерархии организации
@api_view(['GET'])
def delete_suborg(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id = request.GET.get("id")

    try:
        suborg = SubOrganizations.objects.get(id = id)
        suborg.delete()
        return HttpResponse('{"status": "Организация удалена!"}', content_type="application/json", status=200)
    except:
        return HttpResponse('{"status": "Не удалось удалить организацию!"}', content_type="application/json", status=500)

#Сервисы, связанные с воспитателями
# Сервис получения списка воспитателей для обычных пользователей
@api_view(['GET'])
def metodistlist(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    metodist = request.GET.get("metodist")
    metodist = '%' + metodist + '%'
    id_org = get_org_id(request.user)    
    
    page = int(request.GET.get("page"))
    
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""WITH pr as(
                    SELECT
                        profile.name as username,
						profile.id_org
                    FROM
                        serviceback_profileuser as profile
                    WHERE
                        profile.id_org = '{id_org}'
                        and not profile.is_delete
                        and not profile.is_adm_org),
                users as(
                    SELECT
                        users.username,
                        users.first_name,
                        users.email
                    FROM
                        auth_user AS users
                    WHERE
                        users.username in(
                                    SELECT
                                        pr.username
                                    FROM
                                        pr as pr)),
			   full_info as(
			   		SELECT
                        pr.username,
				   		pr.id_org,
				   		users.first_name,
				   		users.email
                    FROM 
                        pr
				    LEFT JOIN
				   		users
				  	ON
				   		pr.username = users.username)
					SELECT
						*
					FROM
						full_info
					WHERE
						CASE 
						WHEN 
							'{metodist}'='%%' THEN TRUE
						ELSE 
							upper(username) like upper('{metodist}')
                    		or upper(first_name) like upper('{metodist}')
						END
					ORDER BY
                        first_name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(result):
            lastelement = len(result)
        res = {"list": result[firstelement:lastelement],
               "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res), content_type="application/json")


# Сервис получения списка воспитателей для админа
@api_view(['GET'])
def metodistlistadmin(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    metodist = request.GET.get("metodist")
    metodist = '%' + metodist + '%'
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""with profile as(
                    SELECT
                        profile.id_org as id_org,
                        profile.name as username
                    FROM
                        serviceback_profileuser as profile
                    WHERE 
                        (profile.is_delete=False)),
	            users as(
                    SELECT
                        users.first_name,
                        users.email,
                        users.username
                    FROM
                        auth_user as users
                    WHERE
                        users.username in (
                                    SELECT
                                        pr.username
                                    FROM
                                        profile as pr)
						and not is_staff),
				pr as(
					SELECT
						users.first_name,
                        users.email,
                        users.username,
						prof.id_org
                    FROM
                        users as users
					LEFT JOIN
						profile as prof
					ON
						prof.username = users.username),
		        org as(
                    SELECT
                        org.id_org as id_org,
                        org.org_name as org_name
                    FROM
                        serviceback_organizations as org
                    WHERE
                        org.id_org in (
                                    SELECT
                                        pr.id_org
                                    FROM
                                        pr as pr)),
                all_rec as(
					SELECT
                        pr.id_org,
                        pr.username,
                        pr.first_name,
                        pr.email,
                        org.org_name
                    FROM
						pr as pr
                    LEFT JOIN
                        org as org
                    ON
                        pr.id_org = org.id_org)
				SELECT
					*
				FROM
					all_rec
				WHERE
                    CASE 
						WHEN 
							'{metodist}'='%%' THEN TRUE
						ELSE 
							upper(first_name) like upper('{metodist}')
                    		or upper(org_name) like upper('{metodist}')
                    		or upper(username) like upper('{metodist}')
                    		or id_org = ('{metodist}')
						END
                ORDER BY
					org_name,
                    first_name"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(result):
            lastelement = len(result)
        res = {"list": result[firstelement:lastelement],
               "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res), content_type="application/json")


# Сервис редактирования таблицы "Воспитатели"
@api_view(['POST'])
def metodistedit(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    
    param = request.GET.get("param")
    datastr = request.body
    res = json.loads(datastr)

    if request.user.is_staff:
        id_org = res['id_org']
    else:
        id_org = get_org_id(request.user)

    if param == 'add':
        try:
            User.objects.create_user(
                username=res['username'],
                password=res['password'],
                email=res['email'],
                first_name=res['first_name'])

            profile = ProfileUser()
            profile.id_org = id_org
            profile.is_adm_org = False
            profile.name = res['username']
            profile.save()

            return HttpResponse('{"status": "Успешно добавлен логин"}', content_type="application/json", status=200)
        except:
            return HttpResponse('{"status": "Ошибка добавления. Данный логин уже зарегистрирован."}', content_type="application/json", status=500)

    if param == 'del':
        profile = ProfileUser.objects.filter(name=res['username'], id_org=id_org)
        if profile.exists():
            for itemprofile in profile:
                itemprofile.is_delete = True
                itemprofile.save()
            try:
                u = User.objects.get(username=res['username'])
                u.delete()
            except User.DoesNotExist:
                return HttpResponse('{"status": "Ошибка удаления пользователя"}', content_type="application/json", status=500)
        else:
            return HttpResponse('{"status": "Пользователь не найден"}', content_type="application/json", status=500)

    if param == 'changepass':
        try:
            profile = ProfileUser.objects.filter(name=res['username'], id_org=id_org)
            if profile.exists():
                u = User.objects.get(username=res['username'])
                u.set_password(res['password'])
                u.save()
            else:
                return HttpResponse('{"status": "Пользователь не найден"}', content_type="application/json", status=500)
        except User.DoesNotExist:
            return HttpResponse('{"status": "Ошибка изменения пароля"}', content_type="application/json", status=500)

    return HttpResponse('{"status": "Успешно."}', content_type="application/json", status=200)

# Сервис импорта файлов данных детей
@api_view(['POST'])
def importfile(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id_org = get_org_id(request.user)
    datastr = request.body
    res = json.loads(datastr)
    fileres = res['file']
    fileres = fileres.replace(
        "data:application/vnd.ms-excel;base64,", "")
    
    excel_data = base64.b64decode(fileres)
    path = basepath + 'temp/' + id_org + '.xls'
    with open(path, 'wb+') as destination:
        destination.write(excel_data)
    workbook = xlrd.open_workbook(path)
    worksheet = workbook.sheet_by_index(0)
    nrows = worksheet.nrows

    try:
        for i in range(1, nrows):
            check = Childs.objects.all()
            check = check.filter(iin=worksheet.cell_value(i, 2))
            flagcheck = True
            for itemcheck in check:
                flagcheck = False
            if flagcheck:
                zapischild = Childs()
                zapischild.iin = worksheet.cell_value(i, 2)
                zapischild.name = worksheet.cell_value(
                    i, 3).replace(' ', '') + ' ' + worksheet.cell_value(i, 4).replace(' ', '') + ' ' + worksheet.cell_value(i, 5).replace(' ', '')

                if zapischild.iin[6] == '5':
                    zapischild.gender = 'm'
                else:
                    zapischild.gender = 'w'
                datarojd = datetime.date(int(
                    '2000') + int(zapischild.iin[:2]), int(zapischild.iin[2:4]), int(zapischild.iin[4:6]))
                zapischild.birthday = datarojd
                zapischild.id_org = id_org
                zapischild.registered = False
                zapischild.is_delete = False
                try:
                    zapisgroup = Groups.objects.get(
                        group_name=worksheet.cell_value(i, 23), id_org=id_org)
                except:
                    zapisgroup = Groups()
                    zapisgroup.group_name = worksheet.cell_value(i, 23)
                    zapisgroup.group_count = 0
                    zapisgroup.group_age = 1
                    zapisgroup.id_org = id_org
                    zapisgroup.save()
                    zapisgroup.id_group = zapisgroup.id
                    zapisgroup.save()
                zapischild.id_group = zapisgroup.id_group
                zapischild.save()

        return HttpResponse('{"status": "Данные детей успешно загружены"}', content_type="application/json", status=200)
    except:
        return HttpResponse('{"status": "Ошибка добавления. Некорректные данные."}', content_type="application/json", status=500)

# Сервис начальной страницы для обычных пользователей
@api_view(['GET'])
def startpage(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    id_org = get_org_id(request.user)

    query = f"""with count_gr as(
                    SELECT
                        count(*) as quantityofgroup 
                    FROM
                        public.serviceback_groups
                    WHERE
                        id_org = '{id_org}' 
                        and is_delete = false),
                count_child as(
                    SELECT 
                        count(*) as quantityofchild
                    FROM
                        public.serviceback_childs
                    WHERE
                        id_org = '{id_org}'
                        and is_delete = false),
                count_metodist as(
                    SELECT
                        count(*) as quantityofmetodist
                    FROM
                        public.serviceback_profileuser
                    WHERE
                        id_org = '{id_org}' 
                        and is_delete = false),
				fullinfo as(
                    SELECT
                        fullname,
                        org_name, 
                        bin, 
                        email, 
                        phonenumber, 
                        adress, 
                        latitude, 
                        longitude,
						cast(case when type_org = 'pr' then 'Частный'
							 else 'Государственный' end as text) as type_org,
						case when type_city = 'gor' then 'Город'
							else 'Село' end as type_city,
					    case when type_ecolog = 'normal' then 'Обычный'
						     when type_ecolog = 'rad' then 'В зоне радиации'
							else 'В зоне экологии' end as type_ecolog,
                        count_place
                    FROM
                        public.serviceback_organizations
				    WHERE
                        id_org = '{id_org}')
                SELECT
                    count_gr.quantityofgroup,
                    count_child.quantityofchild,
                    count_metodist.quantityofmetodist,
                    fullinfo.* 
                FROM
                    count_gr,
                    count_child, 
                    count_metodist, 
                    fullinfo"""

    with connection.cursor() as cursor:
        cursor.execute(query, [id_org])
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result), content_type="application/json")

# Сервис начальной страницы для админа
@api_view(['GET'])
def startpageadmin(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    id_org = get_org_id(request.user)
    
    query = f"""with count_gr as(
                    SELECT
                        count(*) as quantityofgroup 
                    FROM
                        public.serviceback_groups
                    WHERE
                        is_delete = false),
                count_child as(
                    SELECT 
                        count(*) as quantityofchild
                    FROM
                        public.serviceback_childs
                    WHERE
                        is_delete = false),
                count_metodist as(
                    SELECT
                        count(*) as quantityofmetodist
                    FROM
                        public.serviceback_profileuser
                    WHERE
                        is_delete = false),
				fullinfo as(
                    SELECT
                        fullname,
                        org_name, 
                        bin, 
                        email, 
                        phonenumber, 
                        adress, 
                        latitude, 
                        longitude 
                    FROM
                        public.serviceback_organizations
                    WHERE 
                        id_org = '{id_org}')
                SELECT
                    count_gr.quantityofgroup,
                    count_child.quantityofchild,
                    count_metodist.quantityofmetodist,
                    fullinfo.* 
                FROM
                    count_gr,
                    count_child, 
                    count_metodist, 
                    fullinfo"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result), content_type="application/json")


# Сервис регистрации новой организации
@api_view(['POST'])
def registration(request):
    
    datastr = request.body
    item = json.loads(datastr)
    neworg = NewOrganizations()
    
    try:
        neworg = NewOrganizations.objects.get(bin = item['bin'])
        return HttpResponse('{"status": "Заявка уже зарегистрирована!"}', content_type="application/json", status=500)
    except:
        neworg.id_obl      = item['id_obl']
        neworg.id_region      = item['id_region']
        neworg.bin      = item['bin']
        neworg.fullname = item['fullname']
        neworg.org_name = item['org_name']
        neworg.phonenumber = item['phonenumber']
        neworg.adress      = item['adress']
        neworg.email      = item['email']
        neworg.password      = item['bin']
        neworg.latitude      = item['latitude']
        neworg.longitude      = item['longitude']
        neworg.datestatus = datetime.date.today()
        neworg.status      = 0
        neworg.count_place = item['count_place']
        neworg.type_org = item['type_org']
        neworg.type_city = item['type_city']
        neworg.type_ecolog = item['type_ecolog']
        neworg.save()
        return HttpResponse('{"status": "Спасибо за регистарцию, в ближайшее время с Вами свяжутся по указанному номеру телефона"}', content_type="application/json", status=200)

# Сервис получения списка областей
@api_view(['GET'])
def getoblasttype(request):
    
    query = f"""SELECT
                    id as id_obl,
                    name as name_obl
                FROM
                    serviceback_regions
                WHERE
                    id_parent = 0
                ORDER BY
                    id_obl"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result), content_type="application/json")

# Сервис получения списка районов
@api_view(['GET'])
def getregiontype(request):

    query = f"""SELECT
                    id as id_region, 
                    name as name_region, 
                    latitude, 
                    longitude 
                FROM 
                    serviceback_regions 
                WHERE 
                    id_parent = '1'"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result), content_type="application/json")

# Сервис получения организации
@api_view(['GET'])
def getorgelement(request):
    
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id_org = request.GET.get("id_org")

    org = Organizations.objects.filter(bin = id_org)
    if org.exists():
        # Получаем массив групп в запросе и также проверяем на существование
        data = list(org.values())
        return HttpResponse(json.dumps(data, indent=1, default=str), content_type="application/json", status=200)
    else: # Если нет организации возвращаем ложь
        return HttpResponse('{"status": "{Организация не найдена}"}', content_type="application/json", status=500)

# Сервис получения списка организаций
@api_view(['GET'])
def getorglist(request):
    
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    id_org = request.GET.get("bin")
    id_org = '%' + id_org + '%'
    org_name = request.GET.get("org_name")
    org_name = '%' + org_name + '%'
    page = int(request.GET.get("page"))
    firstelement = (page-1) * 25
    lastelement = page * 25

    query = f"""with allorg as(
                    SELECT
                        *
                    FROM
                        serviceback_organizations
                    WHERE
                        id_org LIKE '{org_name}'
                        or upper(org_name) like upper('{org_name}')),
                    allidname as(
                        with allid as(
                            SELECT
                                *
                            FROM
                                serviceback_regions),
                        regid as(
                            SELECT
                                *
                            FROM
                                allid
                            WHERE
                                not id_parent = '0'),
                        nameobl as (
                            SELECT
                                *
                            FROM
                                allid
                            WHERE
                                id IN
                                    (SELECT
                                        id_parent
                                    FROM
                                        regid)),
                        allname as(
                            SELECT 
                                regid.*, 
                                nameobl.name as name_obl 
                            FROM
                                regid
                            LEFT JOIN
                                nameobl
                            ON
                                nameobl.id = regid.id_parent)
                        select
                            *
                        FROM
                            allname),
					kones as(
                        SELECT
                            allorg.id,
                            allorg.bin,
                            allorg.org_name,
                            allorg.fullname, 
                            allorg.id_obl as id_obl, 
                            allorg.id_region as id_region,
                            allidname.name as name_region, 
                            allidname.name_obl 
                        FROM
                            allorg 
                        LEFT JOIN
                            allidname 
                        ON
                            allorg.id_region = allidname.id)
					SELECT
                        * 
                    FROM
                        kones"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        if lastelement > len(result):
            lastelement = len(result)
        res = {"list": result[firstelement:lastelement],
               "last": lastelement, "total": len(result)}

    return HttpResponse(json.dumps(res, indent=4, default=str), content_type="application/json")


# Сервис получения списка заявок на регистрацию
@api_view(['GET'])
def requestlist(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    page = int(request.GET.get("page"))
    
    query = """with allrequest as(
                    SELECT
                        *
                    FROM
                        serviceback_neworganizations
                    WHERE
                        status = '0'),
                    allidname as(
                        with allid as(
                            SELECT
                                *
                            FROM
                                serviceback_regions),
                        regid as(
                            SELECT
                                *
                            FROM
                                allid
                            WHERE
                                not id_parent = '0'),
                        nameobl as (
                            SELECT
                                *
                            FROM
                                allid
                            WHERE
                                id IN
                                    (SELECT
                                        id_parent
                                    FROM
                                        regid)),
                        allname as(
                            SELECT 
                                regid.*, 
                                nameobl.name as name_obl 
                            FROM
                                regid
                            LEFT JOIN
                                nameobl
                            ON
                                nameobl.id = regid.id_parent)
                        select
                            *
                        FROM
                            allname),
					kones as(
                        SELECT
                            allrequest.id,
                            allrequest.bin,
                            allrequest.org_name,
                            allrequest.fullname,
                            allrequest.status, 
                            allrequest.id_obl as id_obl, 
                            allrequest.id_region as id_region,
                            allidname.name as name_region, 
                            allidname.name_obl 
                        FROM
                            allrequest 
                        LEFT JOIN
                            allidname 
                        ON
                            allrequest.id_region = allidname.id)
					SELECT
                        * 
                    FROM
                        kones"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result, indent=4, default=str), content_type="application/json")    

#нужно удалить
# Сервис одобрения заявки администратором таймас
@api_view(['GET'])
def successrequest(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    id_request = request.GET.get("id_request")
    action = request.GET.get("action")

    newreq = NewOrganizations.objects.get(id=id_request)
    if action:
        try:
            existorg = Organizations.objects.get(id_org = newreq.bin)
            newreq.status = "0"
            newreq.save()
            return HttpResponse('{"status":"error", "message": "Организация с таким БИН уже зарегистрирована!"}', content_type="application/json", status = 400) 
        except:
            tmp = 1
        
        try:
            existuser = User.objects.get(username = newreq.bin)
            newreq.status = "0"
            newreq.save()
            return HttpResponse('{"status":"error", "message": "Пользователь с таким БИН уже зарегистрирован!"}', content_type="application/json", status = 400) 
        except:
            tmp = 1

        try:
            existprofile = ProfileUser.objects.get(name = newreq.bin, is_delete = False)
            newreq.status = "0"
            newreq.save()
            return HttpResponse('{"status":"error", "message": "Пользователь уже зарегистрирован!"}', content_type="application/json", status = 400) 
        except:
            tmp = 1



        createneworg = Organizations(id_org = newreq.bin, 
                                    org_name = newreq.org_name,
                                    latitude = newreq.latitude,
                                    longitude = newreq.longitude,
                                    create_date = datetime.date.today(),
                                    checkedgps = True,
                                    adress = newreq.adress,
                                    bin = newreq.bin,
                                    email = newreq.email,
                                    fullname = newreq.fullname,
                                    phonenumber = newreq.phonenumber,
                                    worktimestart = datetime.datetime.strptime("08:00:00", '%H:%M:%S').time(),
                                    worktimestop  = datetime.datetime.strptime("19:00:00", '%H:%M:%S').time() 
                                    )
        createneworg.save()

        createnewuser = User.objects.create_user(
                username = newreq.bin,
                password = newreq.password,
                email = newreq.email,
                first_name = newreq.fullname)
        createnewuser.save()
        
        # message = f"Уважаемый {createnewuser.first_name}!\n Вы успешно прошли регистрацию в 'Таймас', для входа пройдите по ссылке 'https://taimas11.qazna24.kz'\n. Ваши данные для входа на сайт: \n Логин: {createnewuser.username}\n Пароль: {createnewuser.username}"

        # createnewuser.email_user('Регистрация', message, from_email=createnewuser.email)

        profile = ProfileUser(id_org = newreq.bin, is_adm_org = True, name = newreq.bin)
        profile.save()

        newreq.status = "1"
        newreq.save()

        return HttpResponse('{"status":"Одобрено"}', content_type="application/json")
    else:
        newreq.status = "2"
        newreq.save()
        return HttpResponse('{"status":"Отказано"}', content_type="application/json")


# Сервис одобрения заявки администратором таймас
@api_view(['POST'])
def notification(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    action = request.GET.get("action")
    datastr = request.body
    res = json.loads(datastr)
    
    for item in res:
        newreq = NewOrganizations.objects.get(id=item['id'])
        if action:
            try:
                existorg = Organizations.objects.get(id_org = item['bin'])
                newreq.status = "0"
                newreq.save()
                return HttpResponse('{"status":"error", "message": "Организация с таким БИН уже зарегистрирована!"}', content_type="application/json", status = 400) 
            except:
                tmp = 1
                
            try:
                existuser = User.objects.get(username = item['bin'])
                newreq.status = "0"
                newreq.save()
                return HttpResponse('{"status":"error", "message": "Пользователь с таким БИН уже зарегистрирован!"}', content_type="application/json", status = 400) 
            except:
                tmp = 1

            try:
                existprofile = ProfileUser.objects.get(name = item['bin'], is_delete = False)
                newreq.status = "0"
                newreq.save()
                return HttpResponse('{"status":"error", "message": "Пользователь уже зарегистрирован!"}', content_type="application/json", status = 400) 
            except:
                tmp = 1

            createneworg = Organizations(id_org = item['bin'], 
                                    org_name = item['org_name'],
                                    latitude = item['latitude'],
                                    longitude = item['longitude'],
                                    id_obl = item['id_obl'],
                                    id_region = item['id_region'],
                                    create_date = datetime.date.today(),
                                    checkedgps = True,
                                    adress = item['adress'],
                                    bin = item['bin'],
                                    email = item['email'],
                                    fullname = item['fullname'],
                                    phonenumber = item['phonenumber'],
                                    # name_obl = item['name_obl'],
                                    # name_region = item['name_region'],
                                    count_place = item['count_place'],
                                    type_org = item['type_org'],
                                    type_city = item['type_city'],
                                    type_ecolog = item['type_ecolog'],
                                    worktimestart = datetime.datetime.strptime("08:00:00", '%H:%M:%S').time(),
                                    worktimestop  = datetime.datetime.strptime("19:00:00", '%H:%M:%S').time())
            createneworg.save()

            createnewuser = User.objects.create_user(
                username = item['bin'],
                password = item['bin'],
                email = item['email'],
                first_name = item['fullname'])
            createnewuser.save()
        
        # message = f"Уважаемый {createnewuser.first_name}!\n Вы успешно прошли регистрацию в 'Таймас', для входа пройдите по ссылке 'https://taimas11.qazna24.kz'\n. Ваши данные для входа на сайт: \n Логин: {createnewuser.username}\n Пароль: {createnewuser.username}"

        # createnewuser.email_user('Регистрация', message, from_email=createnewuser.email)

            profile = ProfileUser(id_org = item['bin'], is_adm_org = True, name = item['bin'])
            profile.save()

            newreq.status = "1"
            newreq.save()

            return HttpResponse('{"status":"Одобрено"}', content_type="application/json")
        else:
            newreq.status = "2"
            newreq.save()
            return HttpResponse('{"status":"Отказано"}', content_type="application/json")


# Сервис редактирования организации
@api_view(['POST'])
def orgedit(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    if not request.user.is_staff:
        return HttpResponse('{"status": "Нет доступа для редактирования."}', content_type="application/json", status=500)

    datastr = request.body
    res = json.loads(datastr)

    for item in res:
        org = Organizations.objects.filter(id=item['id'])
        if org.exists():
            for orgitem in org:
                orgitem.org_name = item['org_name']
                orgitem.id_obl = item['id_obl']
                orgitem.id_region = item['id_region']
                orgitem.latitude = item['latitude']
                orgitem.longitude = item['longitude']
                orgitem.adress = item['adress']
                orgitem.bin = item['bin']
                orgitem.email = item['email']
                orgitem.fullname = item['fullname']
                orgitem.phonenumber = item['phonenumber']
                orgitem.count_place = item['count_place']
                orgitem.type_org = item['type_org']
                orgitem.type_city = item['type_city']
                orgitem.type_ecolog = item['type_ecolog']
                orgitem.checkedgps = item['checkedgps']
                orgitem.save() 
                return HttpResponse('{"status":"Организация успешно отредактирована"}', content_type="application/json", status=200)
        else:
            return HttpResponse('{"status":"Организация не найдена"}', content_type="application/json", status=500)    


@api_view(['GET'])
def sendmail(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)   


    try:
        user = User.objects.get(username='150140003087')
        print(user.email)
        message = f"Уважаемый {user.first_name}!\n Вы успешно прошли регистрацию в 'Таймас', для входа пройдите по ссылке 'https://taimas11.qazna24.kz'\n. Ваши данные для входа на сайт: \n Логин: {user.username}\n Пароль: {user.username}"


        user.email_user('Подтверждение регистрации!', message, from_email=user.email)
        # send_mail('lfsdakjklfdsjlkdsffjklds', "message",
        #                   DEFAULT_FROM_EMAIL, RECIPIENTS_EMAIL)
    except BadHeaderError:
        return HttpResponse('Ошибка в теме письма.')
    return HttpResponse('{"status":"Успешно"}', content_type="application/json", status = 200)


# Сервис получения данных по заявке
@api_view(['GET'])
def requestelement(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    id_request = request.GET.get("id_request")

    query = f"""with allrequest as(
                    SELECT
                        *
                    FROM
                        serviceback_neworganizations
                    WHERE
                        id = '{id_request}'),
                    allidname as(
                        with allid as(
                            SELECT
                                *
                            FROM
                                serviceback_regions),
                        regid as(
                            SELECT
                                *
                            FROM
                                allid 
                            WHERE
                                not id_parent = '0'),
                        nameobl as (
                            SELECT
                                * 
                            FROM 
                                allid 
                            WHERE
                                id IN
                                    (SELECT
                                        id_parent
                                    FROM
                                        regid)),
                        allname as(
                            SELECT
                                regid.*,
                                nameobl.name as name_obl
                            FROM
                                regid
                            LEFT JOIN
                                nameobl
                            ON
                                nameobl.id = regid.id_parent)
                        SELECT
                            *
                        FROM
                            allname)
                    SELECT 
                        allrequest.*, 
                        allidname.name as name_region, 
                        allidname.name_obl 
                    FROM
                        allrequest 
                    LEFT JOIN
                        allidname 
                    ON
                        allrequest.id_region = allidname.id"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
    return HttpResponse(json.dumps(result, indent=4, default=str), content_type="application/json")

# Сервис частичной и полной синхронизации с КазнаДДО
@api_view(['POST'])
def update(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # Структура должна быть как показано ниже
    datajson = """[
    {
        "org_id": 99999,
        "updateorg": true,
        "group_mass": [
            {
                "group_id": "10001716",
                "group_name": "АДМ",
                "group_count": 32,
                "group_age": "2",
                "category": "1",
                "child_mass": [
                    {
                        "child_name": "Акжигитова Индира Сансызбаевна",
                        "child_iin": "730101403130",
                        "child_birthday": "20.02.1992",
                        "child_gender": "w"
                    },
                    {
                        "child_name": "Ашимов Танат Канатович",
                        "child_iin": "910511301086",
                        "child_birthday": "20.02.1992",
                        "child_gender": "m"
                    }
                ]
            },
            {
                "group_id": "10001717",
                "group_name": "Методист",
                "group_count": 32,
                "group_age": "2",
                "category": "1",
                "child_mass": [
                    {
                        "child_name": "Аманжол Бексұлтан Батырханұлы",
                        "child_iin": "940910301873",
                        "child_birthday": "20.02.1992",
                        "child_gender": "m"
                    }
                ]
            }
        ]
    }
]"""

    # Получаем данные с тела запроса
    data = json.loads(request.body)
    for orgitem in data:
        id_org = orgitem['org_id']
        # Получаем имеющиеся данные с БД (таблица организации)
        findorg = Organizations.objects.filter(id_org=id_org)
        #Если True, то очищаем данные у детей
        updateorg = orgitem['updateorg']
        deleteonlychilds = orgitem['deleteonlychilds']

        # Проверяем, есть ли в БД такая организация
        if findorg.exists():
            if updateorg:
                #Ищем детей с id_org
                delchild = Childs.objects.filter(id_org=id_org)
                for itemdel in delchild:
                    itemdel.id_org = ''
                    itemdel.id_group = ''
                    itemdel.save()

            # Получаем массив групп в запросе и также проверяем на существование
            groups = orgitem['group_mass']
            for itemgr in groups:
                
                childs = itemgr['child_mass']
                id_group = itemgr['group_id']
                # Получаем из БД все группы и фильтруем по коду орг и группы
                findgroup = Groups.objects.filter(id_group=id_group, id_org=id_org)                    
                        
                # Проверим, есть ли такая группа
                if findgroup.exists():
                    for gr in findgroup:
                        gr.group_name = itemgr['group_name']
                        gr.group_count = itemgr['group_count']
                        gr.group_age = itemgr['group_age']
                        gr.category = itemgr['category']
                        gr.save()
                    # ChildEdit(id_org, id_group, childs)
                # Если нет в БД, то создаем группу
                else:
                    itemGr = Groups(id_org=id_org, id_group=id_group, group_name=itemgr['group_name'],
                                    group_count = itemgr['group_count'], group_age = itemgr['group_age'],
                                    category = itemgr['category'])
                    itemGr.save()

                if deleteonlychilds:
                   ChildEdit('', '', childs) 
                    
                else:         
                    ChildEdit(id_org, id_group, childs)
                    
                # visitEdit(id_org, id_group, childs, deleteonlychilds)
            
        else: # Если нет организации возвращаем ложь
            return HttpResponse('{"status": "{Организация не зарегистрирована}"}', content_type="application/json", status=500)

    return HttpResponse('{"status": "success"}', content_type="application/json")

# Сервис частичной и полной синхронизации с КазнаДДО
@api_view(['POST'])
def transfer(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # Получаем данные с тела запроса
    data = json.loads(request.body)
    for orgitem in data:
        id_org = orgitem['org_id']
        # Получаем имеющиеся данные с БД (таблица организации)
        findorg = Organizations.objects.filter(id_org=id_org)
        #Если True, то очищаем данные у детей
        
        # Проверяем, есть ли в БД такая организация
        if findorg.exists():
            
            # Получаем массив групп в запросе и также проверяем на существование
            groups = orgitem['group_mass']
            for itemgr in groups:
                
                childs = itemgr['child_mass']
                id_group = itemgr['group_id']

                VisitEdit(id_org, id_group, childs)
                    
                # visitEdit(id_org, id_group, childs, deleteonlychilds)
            
        else: # Если нет организации возвращаем ложь
            return HttpResponse('{"status": "{Организация не зарегистрирована}"}', content_type="application/json", status=500)

    return HttpResponse('{"status": "success"}', content_type="application/json")



def VisitEdit(id_org, id_group, itemchild):
    
    for itemch in itemchild:
        # Получаем детей в БД, фильтруем по орг, группе, и коду детей
        findchild = Childs.objects.all()
        findchild = findchild.filter(iin=itemch['child_iin'])
        
        mass_date = itemch['child_date'].split('.')
        child_date = datetime.date(int(mass_date[2]), int(mass_date[1]), int(mass_date[0]))
        findvisit = Visits.objects.filter(id_org = id_org, iin=itemch['child_iin'], datestatus = child_date)
        if findvisit.exists():
            for vis in findvisit:
                vis.id_group = id_group
                vis.save()

#Сервис редактирования данных ребенка
def ChildEdit(id_org, id_group, itemchild):
    for itemch in itemchild:
        # Получаем детей в БД, фильтруем по орг, группе, и коду детей
        findchild = Childs.objects.filter(iin=itemch['child_iin'])
                
        # если ребенок существует    
        if findchild.exists():
            for zap in findchild:
                zap.name = itemch['child_name']
                birthday = itemch['child_birthday']
                mass = birthday.split('.')
                datarojd = datetime.date(
                int(mass[2]), int(mass[1]), int(mass[0]))
                zap.birthday = datarojd
                zap.gender = itemch['child_gender']
                zap.id_group = id_group
                zap.category = itemch['child_category']
                zap.id_org = id_org
                zap.is_delete = False
                zap.save()   
        # Если нет в БД, то создаем запись
        else:
            mass = itemch['child_birthday'].split('.')
            datarojd = datetime.date(
                       int(mass[2]), int(mass[1]), int(mass[0]))
            itemChld = Childs(
                            id_group=id_group,
                            id_org=id_org,
                            iin=itemch['child_iin'],
                            name=itemch['child_name'],
                            birthday = datarojd,
                            gender = itemch['child_gender'],
                            category = itemch['child_category'],
                            is_delete = False
                            )
            itemChld.save()

#Сервис для передачи данных посещения в КазнаДДО
@api_view(['POST'])
def visitforKazna(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # Структура должна быть как показано ниже
    datajson = """[
                    {
                        "org_id": "910114301692",
                        "day_mass": [
                                "2023-02-22",
                                "2023-02-23"
                                    ]
                    }
                ]"""

    # Получаем данные с тела запроса
    data = json.loads(request.body)
    for orgitem in data:
        id_org = orgitem['org_id']
        # Получаем имеющиеся данные с БД (таблица организации)
        findorg = Organizations.objects.filter(id_org=id_org)
        # Проверяем, есть ли в БД такая организация
        if findorg.exists():
        # Получаем массив групп в запросе и также проверяем на существование
            days = orgitem['day_mass']
            strel = "', '".join(days)
            
            query = f"""
                        with temp_visit as(
                            SELECT
                                *
                            FROM 
                                serviceback_visits
                            WHERE
                                id_org = '{id_org}'
                                and datestatus in ('{strel}')),
                        last_record as(
                            SELECT
                                iin,
                                datestatus,
                                max(timestatus) as timestatus
                            FROM
                                temp_visit
                            GROUP BY
                                iin,
                                datestatus)
                            SELECT 
                                id_org,
                                id_group,
                                iin,
                                datestatus,
                                status
                            FROM 
                                temp_visit
                            WHERE (iin,	timestatus) in (
                                                    SELECT
                                                        iin,
                                                        timestatus
                                                    FROM
                                                        last_record)
                                and status in ('2', '3', '4', '5', '10')"""
            with connection.cursor() as cursorgr:
                cursorgr.execute(query)
                columnsgr = [col[0] for col in cursorgr.description]
                resultgr = [dict(zip(columnsgr, rowgr))
                    for rowgr in cursorgr.fetchall()]

            return HttpResponse(json.dumps(resultgr, indent=4, default=str), content_type="application/json", status=200)
        else: # Если нет организации возвращаем ложь
            return HttpResponse('{"status": "{Организация не зарегистрирована}"}', content_type="application/json", status=500)

#Сервис для передачи данных посещения в КазнаДДО
@api_view(['POST'])
def visitforKaznaFullOrg(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # Структура должна быть как показано ниже
    datajson = """[
                    {
                    "_day": "2023-10-23",
                    "org_mass"[
                        "123456789987",
                        "123321456987"
                    ]
                    }
                ]"""

    # Получаем данные с тела запроса
    data = json.loads(request.body)
    for orgitem in data:
        _date = orgitem['_day']
        org_mass = orgitem['org_mass']
        strel = "', '".join(org_mass)
        # current_date = datetime.date.today()

        try:        
            query = f"""
                    with temp_visit as(
                            SELECT
                                *
                            FROM 
                                serviceback_visits
                            WHERE
                                id_org in ('{strel}')
                                and datestatus = '{_date}'),
                        last_record as(
                            SELECT
                                id_org,
                                iin,
                                max(id) as _id
                            FROM
                                temp_visit
                            GROUP BY
                                id_org,
                                iin)
                            SELECT 
                                id_org,
                                id_group,
                                iin,
                                datestatus,
                                status
                            FROM 
                                temp_visit
                            WHERE (id) in (
                                            SELECT
                                                _id
                                            FROM
                                                last_record)
                                and status in ('2', '3', '4', '5', '10')"""
            with connection.cursor() as cursorgr:
                cursorgr.execute(query)
                columnsgr = [col[0] for col in cursorgr.description]
                resultgr = [dict(zip(columnsgr, rowgr))
                for rowgr in cursorgr.fetchall()]

                return HttpResponse(json.dumps(resultgr, indent=4, default=str), content_type="application/json", status=200)
        except:
            return HttpResponse('{"status": "{Не удалось выполнить операцию}"}', content_type="application/json", status=500)

#нужно удалить
@api_view(['GET'])
def formnovisit(request):

    datestr = str(datetime.datetime.today())
    query = f"""
                with childs as(
                SELECT
                    iin,
                    id_group,
                    id_org
                FROM
                    public.serviceback_childs
                WHERE
                    iin = '920220301670'),
                visit as(
                    SELECT
                        iin,
                        id_group,
                        id_org
                    FROM
                        public.serviceback_visits
                    WHERE
                        datestatus = '{datestr}'
                        and (id_org, id_group, iin)
                                    in (
                                        SELECT
                                            id_org,
                                            id_group,
                                            iin
                                        FROM
                                            childs)
                    ),
                novis as(
                    SELECT
                        iin,
                        id_group,
                        id_org
                    FROM
                        childs
                    WHERE
                        (id_org, id_group, iin)
                                    not in (
                                        SELECT
                                            id_org,
                                            id_group,
                                            iin
                                        FROM
                                            visit)),
                status as(
                    SELECT
                        max(case when weekend = '{datestr}' then
                                (case when name = 'В' then 50
                                    else 55 end)
                            else 5 end) as status from serviceback_weekendday)
                    SELECT 
                        *
                    FROM novis,
                        status"""
    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
        
        date_obj = datetime.datetime.now()
        formatted_date = date_obj.strftime('%Y-%m-%d')

        current_time = datetime.datetime.now().time()
        time_str = current_time.strftime("%H:%M:%S")            

        for item in result:
            crVisit = Visits.objects.create(id_org = item['id_org'], id_group = item['id_group'],
                                                 iin = item['iin'], datestatus = formatted_date,
                                                 status = item['status'], timestatus = time_str, edited = False)

    return HttpResponse('{"status": "успешно"}', content_type="application/json", status=200)


def format_date(date_obj):
    months = {
        1: 'январь',
        2: 'февраль',
        3: 'март',
        4: 'апрель',
        5: 'май',
        6: 'июнь',
        7: 'июль',
        8: 'август',
        9: 'сентябрь',
        10: 'октябрь',
        11: 'ноябрь',
        12: 'декабрь'
    }
    month_name = months[date_obj.month]
    year = str(date_obj.year)
    return f"{month_name} {year}"

# Сервис получения табеля в формате эксель
@api_view(['GET'])
def generetexlstabel(request):
    
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    # id_org = get_org_id(request.user)
    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')

    id_org    = request.GET.get("id_org")
    id_group = request.GET.get("id_group")
    datenachalo = request.GET.get("datenachalo")
    datekonec = request.GET.get("datekonec")

    org_name = Organizations.objects.get(id_org=id_org).org_name
    if id_group == '':
        group_name = '__________________________________________'
        group_category = '__________________________________________'
        metodist_name = '____________________________________________'
    else:
        group_name = Groups.objects.get(id_group=id_group).group_name
        group_category = Groups.objects.get(id_group=id_group).category
        if group_category == '':
            group_category = '__________________________________________'
        elif group_category == 'gor10':
            group_category = '10,5 часовой'
        elif group_category == 'gor9':
            group_category = '9 часовой'
        elif group_category == 'gorp':
            group_category = 'Неполный день'
        elif group_category == 'gors':
            group_category = 'Санаторный'
        elif group_category == 'gork':
            group_category = 'Коррекционный'

        username = Groups.objects.get(id_group=id_group).username
        metodist_name = User.objects.get(username=username).first_name

    date_object = datetime.datetime.strptime(datenachalo, '%d.%m.%Y')
    monthandyear = format_date(date_object)
    amountday = int(datekonec[:2]) # количество дней в месяце
    raznicadnei = 31 - amountday   # разница дней с 31

    id_group = '%' + id_group + '%'

    conn = ps.connect(
        host="192.168.5.23",
        database="facedb",
        user="postgres",
        password="1"
    )

    sql_query = f"""with weekday as(SELECT * FROM serviceback_weekendday WHERE weekend>='{datenachalo}' and weekend <= '{datekonec}'),

childs as(
                    SELECT
                        child.name as childname,
                        child.image_url as image,
						child.iin as iin,
						child.id_group,
						child.id_org
                    FROM
                        serviceback_childs as child 
                    WHERE
                        child.id_org = '{id_org}'
                        and child.id_group like '{id_group}'
						and not child.is_delete),
	 visits as(
                    SELECT
                        visits.id_org,
		 				visits.id_group,
		 				visits.iin,
		 				visits.datestatus,
		 				visits.timestatus,
		 				visits.status,
		 				visits.edited
                    FROM
                        serviceback_visits as visits
                    WHERE 
                        visits.id_org = '{id_org}'
						and visits.id_group like '{id_group}'
                        and visits.datestatus >= '{datenachalo}'
                        and visits.datestatus <= '{datekonec}'
						and visits.datestatus not in (SELECT weekend FROM weekday)
					
					ORDER BY
						datestatus),
	visit as (SELECT visits.id_org, visits.id_group, visits.iin, visits.datestatus, max(visits.timestatus) as timestatus
			 	FROM visits
			 	GROUP BY visits.id_org, visits.id_group, visits.iin, visits.datestatus),
weekchild as (SELECT child.id_org, child.id_group, child.iin, w.weekend as datestatus, '7' as status
				FROM childs as child, weekday as w
				GROUP BY child.id_org, child.id_group, child.iin, w.weekend),
un as(
			 select vis.id_org, vis.id_group, vis.iin, vis.datestatus, vis.status, vis.edited
				from visits as vis
				WHERE (vis.id_org, vis.id_group, vis.iin, vis.datestatus, vis.timestatus) in
						(SELECT id_org, id_group, iin, datestatus, timestatus FROM visit)
UNION
			select w.id_org, w.id_group, w.iin, w.datestatus, w.status, false as edited
				from weekchild as w),
gr as(
                    SELECT
                        grr.group_name as grname,
						grr.id_group as id_group
                    FROM
                        serviceback_groups as grr
                    WHERE grr.id_org = '{id_org}'
                        and grr.id_group in (
										SELECT
											un.id_group
										FROM
											un)),
vt as(
                    SELECT
                        un.datestatus, un.status, un.edited,
						childs.iin as iin,
                        childs.childname,
                        childs.image,
						childs.id_group,
                        gr.grname
                    FROM
                        childs 
                    LEFT JOIN gr
                    ON
                        childs.id_group = gr.id_group
                    LEFT JOIN
                        un on un.iin = childs.iin),
vtt as(SELECT id_group, iin, datestatus, status, childname, image, grname,
	   (case when status = '2' or status = '7' then 1 else 0 end) as visch,
		(case when status = '3' or status = '4' then 1 else 0 end) as otherreason	   
FROM vt 
order by iin, datestatus),
vttt as (SELECT id_group, iin, (cast(date_part('day', timestamp '{datekonec}') as integer) - sum(visch)) as novisch, sum(visch) as visch, sum(otherreason) as other FROM vtt
GROUP BY id_group, iin),
tmp_vis as(
	SELECT vtt.id_group, vtt.iin, vtt.datestatus, cast(vtt.status as integer) as status, vtt.childname, vtt.grname, vttt.novisch, vttt.other
		from vtt as vtt
		LEFT JOIN vttt as vttt
		ON vtt.id_group = vttt.id_group and vtt.iin = vttt.iin),
tmp_final as(SELECT
                    iin as iin, 
                    max(childname) as childname,
					max(novisch) as novisch,
					max(other) as other,
                    max(case when date_part('day', datestatus) = 1 then status
                        else 0 end) as "day_1",	
                    max(case when date_part('day', datestatus) = 2 then status
                        else 0 end) as "day_2",
                    max(case when date_part('day', datestatus) = 3 then status
                        else 0 end) as "day_3",
                    max(case when date_part('day', datestatus) = 4 then status
                        else 0 end) as "day_4",
                    max(case when date_part('day', datestatus) = 5 then status
                        else 0 end) as "day_5",
                    max(case when date_part('day', datestatus) = 6 then status
                        else 0 end) as "day_6",
                    max(case when date_part('day', datestatus) = 7 then status
                        else 0 end) as "day_7",
                    max(case when date_part('day', datestatus) = 8 then status
                        else 0 end) as "day_8",
                    max(case when date_part('day', datestatus) = 9 then status
                        else 0 end) as "day_9",	
                    max(case when date_part('day', datestatus) = 10 then status
                        else 0 end) as "day_10",
					max(case when date_part('day', datestatus) = 11 then status
                        else 0 end) as "day_11",	
                    max(case when date_part('day', datestatus) = 12 then status
                        else 0 end) as "day_12",
                    max(case when date_part('day', datestatus) = 13 then status
                        else 0 end) as "day_13",
                    max(case when date_part('day', datestatus) = 14 then status
                        else 0 end) as "day_14",
                    max(case when date_part('day', datestatus) = 15 then status
                        else 0 end) as "day_15",
                    max(case when date_part('day', datestatus) = 16 then status
                        else 0 end) as "day_16",
                    max(case when date_part('day', datestatus) = 17 then status
                        else 0 end) as "day_17",
                    max(case when date_part('day', datestatus) = 18 then status
                        else 0 end) as "day_18",
                    max(case when date_part('day', datestatus) = 19 then status
                        else 0 end) as "day_19",	
                    max(case when date_part('day', datestatus) = 20 then status
                        else 0 end) as "day_20",
					max(case when date_part('day', datestatus) = 21 then status
                        else 0 end) as "day_21",	
                    max(case when date_part('day', datestatus) = 22 then status
                        else 0 end) as "day_22",
                    max(case when date_part('day', datestatus) = 23 then status
                        else 0 end) as "day_23",
                    max(case when date_part('day', datestatus) = 24 then status
                        else 0 end) as "day_24",
                    max(case when date_part('day', datestatus) = 25 then status
                        else 0 end) as "day_25",
                    max(case when date_part('day', datestatus) = 26 then status
                        else 0 end) as "day_26",
                    max(case when date_part('day', datestatus) = 27 then status
                        else 0 end) as "day_27",
                    max(case when date_part('day', datestatus) = 28 then status
                        else 0 end) as "day_28",
                    max(case when date_part('day', datestatus) = 29 then status
                        else 0 end) as "day_29",	
                    max(case when date_part('day', datestatus) = 30 then status
                        else 0 end) as "day_30",
					max(case when date_part('day', datestatus) = 31 then status
                        else 0 end) as "day_31" from tmp_vis
                GROUP BY
                    id_group,
                    iin),
tmp_union as(
	SELECT childname, iin, novisch, other,
	   (case when day_1 = 7 then 'В'
	   		when day_1 = 2 then ''
			else 'Н' end) as day_1,
		(case when day_2 = 7 then 'В'
	   		when day_2 = 2 then ''
			else 'Н' end) as day_2,
		(case when day_3 = 7 then 'В'
	   		when day_3 = 2 then ''
			else 'Н' end) as day_3,
		(case when day_4 = 7 then 'В'
	   		when day_4 = 2 then ''
			else 'Н' end) as day_4,
		(case when day_5 = 7 then 'В'
	   		when day_5 = 2 then ''
			else 'Н' end) as day_5,
		(case when day_6 = 7 then 'В'
	   		when day_6 = 2 then ''
			else 'Н' end)  as day_6,
		(case when day_7 = 7 then 'В'
	   		when day_7 = 2 then ''
			else 'Н' end) as day_7,
		(case when day_8 = 7 then 'В'
	   		when day_8 = 2 then ''
			else 'Н' end) as day_8,
		(case when day_9 = 7 then 'В'
	   		when day_9 = 2 then ''
			else 'Н' end)  as day_9,
		(case when day_10 = 7 then 'В'
	   		when day_10 = 2 then ''
			else 'Н' end) as day_10,
		(case when day_11 = 7 then 'В'
	   		when day_11 = 2 then ''
			else 'Н' end) as day_11,
		(case when day_12 = 7 then 'В'
	   		when day_12 = 2 then ''
			else 'Н' end) as day_12,
		(case when day_13 = 7 then 'В'
	   		when day_13 = 2 then ''
			else 'Н' end) as day_13,
		(case when day_14 = 7 then 'В'
	   		when day_14 = 2 then ''
			else 'Н' end) as day_14,
		(case when day_15 = 7 then 'В'
	   		when day_15 = 2 then ''
			else 'Н' end) as day_15,
		(case when day_16 = 7 then 'В'
	   		when day_16 = 2 then ''
			else 'Н' end) as day_16,
		(case when day_17 = 7 then 'В'
	   		when day_17 = 2 then ''
			else 'Н' end) as day_17,
		(case when day_18 = 7 then 'В'
	   		when day_18 = 2 then ''
			else 'Н' end) as day_18,
		(case when day_19 = 7 then 'В'
	   		when day_19 = 2 then ''
			else 'Н' end) as day_19,
		(case when day_20 = 7 then 'В'
	   		when day_20 = 2 then ''
			else 'Н' end) as day_20,
		(case when day_21 = 7 then 'В'
	   		when day_21 = 2 then ''
			else 'Н' end) as day_21,
		(case when day_22 = 7 then 'В'
	   		when day_22 = 2 then ''
			else 'Н' end) as day_22,
		(case when day_23 = 7 then 'В'
	   		when day_23 = 2 then ''
			else 'Н' end) as day_23,
		(case when day_24 = 7 then 'В'
	   		when day_24 = 2 then ''
			else 'Н' end) as day_24,
		(case when day_25 = 7 then 'В'
	   		when day_25 = 2 then ''
			else 'Н' end) as day_25,
		(case when day_26 = 7 then 'В'
	   		when day_26 = 2 then ''
			else 'Н' end) as day_26,
		(case when day_27 = 7 then 'В'
	   		when day_27 = 2 then ''
			else 'Н' end) as day_27,
		(case when day_28 = 7 then 'В'
	   		when day_28 = 2 then ''
			else 'Н' end) as day_28,
		(case when day_29 = 7 then 'В'
	   		when day_29 = 2 then ''
			else 'Н' end) as day_29,
		(case when day_30 = 7 then 'В'
	   		when day_30 = 2 then ''
			else 'Н' end) as day_30,
		(case when day_31 = 7 then 'В'
	   		when day_31 = 2 then ''
			else 'Н' end) as day_31
from tmp_final),
tmp_vt as(select 
	   cast(2 as integer) as poryadok,
       max(CAST(' ' as text)) AS numberstr,
       max(CAST(0 as integer)) AS numberint,
	   max(CAST('Всего отсутствует детей' as text)) as childname,
	   max(CAST(' ' as text)) as iin,
       cast(sum(novisch) as text) as novisch,
	   cast(sum(other) as text) as other,
	   cast(sum(case when day_1='Н' then 1 else 0 end)  as text) as itog_1,
	   cast(sum(case when day_2='Н' then 1 else 0 end) as text) as itog_2,
	   cast(sum(case when day_3='Н' then 1 else 0 end) as text) as itog_3,
	   cast(sum(case when day_4='Н' then 1 else 0 end) as text) as itog_4,
	   cast(sum(case when day_5='Н' then 1 else 0 end) as text) as itog_5,
	   cast(sum(case when day_6='Н' then 1 else 0 end) as text) as itog_6,
	   cast(sum(case when day_7='Н' then 1 else 0 end) as text) as itog_7,
	   cast(sum(case when day_8='Н' then 1 else 0 end) as text) as itog_8,
	   cast(sum(case when day_9='Н' then 1 else 0 end) as text) as itog_9,
	   cast(sum(case when day_10='Н' then 1 else 0 end) as text) as itog_10,
	   cast(sum(case when day_11='Н' then 1 else 0 end) as text) as itog_11,
	   cast(sum(case when day_12='Н' then 1 else 0 end) as text) as itog_12,
	   cast(sum(case when day_13='Н' then 1 else 0 end) as text) as itog_13,
	   cast(sum(case when day_14='Н' then 1 else 0 end) as text) as itog_14,
	   cast(sum(case when day_15='Н' then 1 else 0 end) as text) as itog_15,
	   cast(sum(case when day_16='Н' then 1 else 0 end) as text) as itog_16,
	   cast(sum(case when day_17='Н' then 1 else 0 end) as text) as itog_17,
	   cast(sum(case when day_18='Н' then 1 else 0 end) as text) as itog_18,
	   cast(sum(case when day_19='Н' then 1 else 0 end) as text) as itog_19,
	   cast(sum(case when day_20='Н' then 1 else 0 end) as text) as itog_20,
	   cast(sum(case when day_21='Н' then 1 else 0 end) as text) as itog_21,
	   cast(sum(case when day_22='Н' then 1 else 0 end) as text) as itog_22,
	   cast(sum(case when day_23='Н' then 1 else 0 end) as text) as itog_23,
	   cast(sum(case when day_24='Н' then 1 else 0 end) as text) as itog_24,
	   cast(sum(case when day_25='Н' then 1 else 0 end) as text) as itog_25,
	   cast(sum(case when day_26='Н' then 1 else 0 end) as text) as itog_26,
	   cast(sum(case when day_27='Н' then 1 else 0 end) as text) as itog_27,
	   cast(sum(case when day_28='Н' then 1 else 0 end) as text) as itog_28,
	   cast(sum(case when day_29='Н' then 1 else 0 end) as text) as itog_29,
	   cast(sum(case when day_30='Н' then 1 else 0 end) as text) as itog_30,
	   cast(sum(case when day_31='Н' then 1 else 0 end) as text) as itog_31
from tmp_union
UNION
SELECT  
        cast(1 as integer) as poryadok,
        CAST(ROW_NUMBER() OVER (ORDER BY childname) as text) AS numberstr,
        CAST(ROW_NUMBER() OVER (ORDER BY childname) as integer) AS numberint,
        childname,
		iin,
		cast(novisch as text) as novisch,
		cast(other as text) as other,
		day_1,
		day_2,
		day_3,
		day_4,
		day_5,
		day_6,
		day_7,
		day_8,
		day_9,
		day_10,
		day_11,
		day_12,
		day_13,
		day_14,
		day_15,
		day_16,
		day_17,
		day_18,
		day_19,
		day_20,
		day_21,
		day_22,
		day_23,
		day_24,
		day_25,
		day_26,
		day_27,
		day_28,
		day_29,
		day_30,
		day_31
FROM tmp_union
ORDER BY poryadok, numberint)
SELECT 
       numberstr as numberstr,
       childname as "ФИО ребенка",
	   iin as "ИИН ребенка",
	   itog_1 as "1",
	   itog_2 as "2",
	   itog_3 as "3",
	   itog_4 as "4",
	   itog_5 as "5",
	   itog_6 as "6",
	   itog_7 as "7",
	   itog_8 as "8",
	   itog_9 as "9",
	   itog_10 as "10",
	   itog_11 as "11",
	   itog_12 as "12",
	   itog_13 as "13",
	   itog_14 as "14",
	   itog_15 as "15",
	   itog_16 as "16",
	   itog_17 as "17",
	   itog_18 as "18",
	   itog_19 as "19",
	   itog_20 as "20",
	   itog_21 as "21",
	   itog_22 as "22",
	   itog_23 as "23",
	   itog_24 as "24",
	   itog_25 as "25",
	   itog_26 as "26",
	   itog_27 as "27",
	   itog_28 as "28",
	   itog_29 as "29",
	   itog_30 as "30",
	   itog_31 as "31",
	   novisch as "Всего",
	   other as "По причине",
       CAST(' ' as text) AS "Причина"
FROM tmp_vt"""
    
    with connection.cursor() as cursor:
        cursor.execute(sql_query)
        rows = cursor.fetchall()

    # Создаем новый Excel-файл
    wb = Workbook()
    ws = wb.active

    border = Border(left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'))

    # Записываем данные в Excel-файл
    ws.merge_cells('AB2:AE2')
    ws['AB2'] = 'Приложение 85'
    ws['AB2'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('AB3:AE3')
    ws['AB3'] = 'к приказу и.о. Министра финансов'
    ws['AB3'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('AB4:AE4')
    ws['AB4'] = 'Республики Казахстан'
    ws['AB4'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('AB5:AE5')
    ws['AB5'] = 'от 2 августа 2011 года № 390'
    ws['AB5'] = 'Форма № 305'
    ws['AB5'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('B2:S2')
    ws['B2'].font = Font(u='single')
    ws['B2'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B2'] = org_name

    ws.merge_cells('B3:G3')
    ws['B3'].font = Font(size=9)
    ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B3'] = 'Наименование государственного учреждения (централизованной бухгалтерии)'

    ws.merge_cells('K6:O6')
    ws['K6'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K6'].font = Font(bold=True)
    ws['K6'] = 'Табель'

    ws.merge_cells('K7:O7')
    ws['K7'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K7'].font = Font(bold=True)
    ws['K7'] = 'учета посещаемости детей'
    
    ws.merge_cells('K8:O8')
    ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
    ws['K8'].font = Font(bold=True)
    ws['K8'] = 'за ' + monthandyear + ' г.'

    ws.merge_cells('B9:E9')
    ws['B9'].font = Font(bold=True)
    ws['B9'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B9'] = 'Группа                       ' + group_name

    ws.merge_cells('B10:E10')
    ws['B10'].font = Font(bold=True)
    ws['B10'].alignment = Alignment(horizontal="left", vertical="center")
    ws['B10'] = 'Режим группы      ' + group_category

    ws.row_dimensions[14].height = 60

    # Объединение ячеек
    ws.merge_cells('A13:A14')
    ws.column_dimensions['A'].width = 7
    ws['A13'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A13'].border = border
    ws['A14'].border = border
    ws['A13'].font = Font(bold=True)
    ws['A13'] = '№ п/п'

    ws.merge_cells('B13:B14')
    ws.column_dimensions['B'].width = 40
    ws['B13'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B13'].border = border
    ws['B14'].border = border
    ws['B13'].font = Font(bold=True)
    ws['B13'] = 'ФИО ребенка'

    ws.merge_cells('C13:C14')
    ws.column_dimensions['C'].width = 20
    ws['C13'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C13'].border = border
    ws['C14'].border = border
    ws['C13'].font = Font(bold=True)
    ws['C13'] = 'ИИН ребенка'    

    # вычисляем кол-во дней в месяце для строки Дни
    if amountday == 28:
        ws.merge_cells('D13:AE13')
    elif amountday == 29:
        ws.merge_cells('D13:AF13')
    elif amountday == 30:
        ws.merge_cells('D13:AG13')
    else:
        ws.merge_cells('D13:AH13')
    ws['D13'].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(4, amountday + 4):
        cell = ws.cell(row=13, column=col)
        cell.border = border
    ws['D13'].font = Font(bold=True)
    ws['D13'] = 'Дни'

    NY = ''
    if amountday == 28:
        ws.merge_cells('AF13:AG13')
        NY = 'AF13'
        Reas1 = 'AH13'
        Reas2 = 'AH14'
    elif amountday == 29:
        ws.merge_cells('AG13:AH13')
        NY = 'AG13'
        Reas1 = 'AI13'
        Reas2 = 'AI14'
    elif amountday == 30:
        ws.merge_cells('AH13:AI13')
        NY = 'AH13'
        Reas1 = 'AJ13'
        Reas2 = 'AJ14'
    else:
        ws.merge_cells('AI13:AJ13')
        NY = 'AI13'
        Reas1 = 'AK13'
        Reas2 = 'AK14'

    ws[NY].alignment = Alignment(horizontal="center", vertical="center")
    for col in range(amountday + 4, amountday + 6):
        cell = ws.cell(row=13, column=col)
        cell.border = border
    ws[NY].font = Font(bold=True)
    ws[NY] = 'Пропущено дней'

    cell_below = ws[NY].offset(row=1, column=0)
    cell_below.alignment = Alignment(horizontal="center", vertical="center")
    cell_below.border = border
    cell_below.font = Font(bold=True)
    cell_below.value = 'Всего'

    cell_below = ws[NY].offset(row=1, column=1)
    cell_below.alignment = Alignment(horizontal="center", vertical="distributed")
    cell_below.border = border
    cell_below.font = Font(bold=True)
    cell_below.value = 'В том числе засчитываемых'

    ws.merge_cells(Reas1 + ':' + Reas2)
    ws.column_dimensions[Reas1[:2]].width = 40
    ws[Reas1].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[Reas1].border = border
    ws[Reas2].border = border
    ws[Reas1].font = Font(bold=True)
    ws[Reas1].value = 'Причины \n непосещения \n (основание)'
    # cell_below.value = 'Причины непосещения (основание)'


    # вычисляем буквы от А до AH и т.д. с Аски кода
    cols = []
    for i in range(1, amountday + 1):
        if i <= 26:
            cols.append(string.ascii_uppercase[i-1])
        else:
            cols.append('A' + string.ascii_uppercase[i-27])

    # выводим нумерацию дней от 1 до 30 или 31 или 29
    for col_idx, col_name in enumerate(cols):
        col_num = col_idx + 4
        celll = ws.cell(row=14, column=col_num, value=col_idx+1)
        celll.border = border
        celll.alignment = Alignment(horizontal="center", vertical="center")
    
    # выводим значения из запрооса
    for i, row in enumerate(rows):
        for col in range(1, amountday + 4):
            cel = ws.cell(row=i+15, column=col, value=row[col-1])
            if col != 2: 
                cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = border
    
    # выводим значения из запрооса
    for i, row in enumerate(rows):
        for col in range(35, 38):
            cel = ws.cell(row=i+15, column=col-raznicadnei, value=row[col-1])
            cel.alignment = Alignment(horizontal="center", vertical="center")
            cel.border = border
    

    ws['B12'].offset(row=len(rows) + 4, column=0).value = 'Воспитатель группы      '
    ws['B12'].offset(row=len(rows) + 4, column=0).alignment = Alignment(horizontal="right", vertical="distributed")
    ws['B12'].offset(row=len(rows) + 4, column=0).font = Font(bold=True)


    ws.merge_cells(ws['B12'].offset(row=len(rows) + 4, column=1).coordinate + ':' +  ws['G12'].offset(row=len(rows) + 4, column=0).coordinate)
    ws[ws['B12'].offset(row=len(rows) + 4, column=1).coordinate].value = metodist_name
    ws[ws['B12'].offset(row=len(rows) + 4, column=1).coordinate].alignment = Alignment(horizontal="left", vertical="distributed")
    ws[ws['B12'].offset(row=len(rows) + 4, column=1).coordinate].font = Font(bold=True)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Табель посещаемости.xlsx'
    wb.save(response)
    return response

@api_view(['GET'])
def formfordashsumm(request):
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    query=f"""WITH org as(
            SELECT
                id_org,
                org_name,
                (case when type_org = 'pr' then 'Частный'
                    else 'Государственный' end) as type_org,
                id_obl,
                id_region,
				type_city,
				type_ecolog
            FROM
                serviceback_organizations),
        allreg as(
                SELECT
                    *
                FROM
                    serviceback_regions),
		allprice as(
                SELECT
                    *
                FROM
                    serviceback_priceservice),
	namereg as(
                SELECT
                    reg.id as reg_id,
                    reg.name as reg_name
                FROM
                    allreg as reg
                WHERE
                    id in(
                        SELECT
                            org.id_region
                        FROM
                            org as org)),
        nameobl as(
                SELECT
                    obl.id as obl_id,
                    obl.name as obl_name
                FROM
                    allreg as obl
                WHERE
                    id in(
                        SELECT
                            org.id_obl
                        FROM
                            org as org)),
        full_info as(		
                SELECT 
                    org.id_org as idd_org,
                    org.org_name,
                    org.type_org,
                    org.id_obl,
                    org.id_region,
					org.type_city,
					org.type_ecolog,
                    nameobl.obl_name,
                    namereg.reg_name
                FROM 
                    org as org
                LEFT JOIN
                    nameobl as nameobl
                ON
                    id_obl = obl_id
                LEFT JOIN
                    namereg as namereg
                ON
                    id_region = reg_id),
        c_gr as(
            SELECT 
                gr.id_org,
				gr.id_group,
				gr.group_name,
                gr.category,
                gr.group_count,
				org.id_obl,
				org.type_city,
				org.type_ecolog
            FROM
                serviceback_groups as gr
			LEFT JOIN
				full_info as org
			ON
				gr.id_org = org.idd_org),
	grinfo as(
			SELECT 
                gr.id_org,
				gr.id_group,
				gr.group_name,
                gr.category,
                gr.group_count,
				gr.type_city,
				gr.type_ecolog,
				pr.price
            FROM
                c_gr as gr
			LEFT JOIN
				allprice as pr
			ON
				gr.id_obl = pr.obl_id
				and gr.category = pr.category
				and gr.type_city = pr.type_city
				and gr.type_ecolog = pr.type_ecolog),
    count_week as(
            SELECT 
                date_trunc('month', weekend) as mon,
                sum(1) as count_w
            FROM 
                serviceback_weekendday
            GROUP BY
                mon),
    count_workday as(
            SELECT 
                mon,
                extract(DAY FROM (date_trunc('MONTH', mon::date) + INTERVAL '1 MONTH' - INTERVAL '1 DAY')) - count_w as work_day
            FROM 
                count_week),
	plangr as(
			SELECT
				gr.id_org, 
				gr.id_group,
				sum(cast(case when mon = '2023-01-01' then gr.group_count * work_day else 0 end as integer)) as plan_1,
                sum(cast(case when mon = '2023-02-01' then gr.group_count * work_day else 0 end as integer)) as plan_2,
                sum(cast(case when mon = '2023-03-01' then gr.group_count * work_day else 0 end as integer)) as plan_3,
                sum(cast(case when mon = '2023-04-01' then gr.group_count * work_day else 0 end as integer)) as plan_4,
                sum(cast(case when mon = '2023-05-01' then gr.group_count * work_day else 0 end as integer)) as plan_5,
				sum(cast(case when mon = '2023-06-01' then gr.group_count * work_day else 0 end as integer)) as plan_6,
                sum(cast(case when mon = '2023-07-01' then gr.group_count * work_day else 0 end as integer)) as plan_7,
                sum(cast(case when mon = '2023-08-01' then gr.group_count * work_day else 0 end as integer)) as plan_8,
                sum(cast(case when mon = '2023-09-01' then gr.group_count * work_day else 0 end as integer)) as plan_9,
                sum(cast(case when mon = '2023-10-01' then gr.group_count * work_day else 0 end as integer)) as plan_10,
                sum(cast(case when mon = '2023-11-01' then gr.group_count * work_day else 0 end as integer)) as plan_11,
                sum(cast(case when mon = '2023-12-01' then gr.group_count * work_day else 0 end as integer)) as plan_12
			FROM 
				grinfo as gr,
				count_workday
			GROUP BY
				gr.id_org, id_group),
	planandprice as(			
			SELECT
				gr.id_org,
				gr.id_group,
				(gr.plan_1 * pr.price)/1000 as plan_jan,
				(gr.plan_2 * pr.price)/1000 as plan_feb,
				(gr.plan_3 * pr.price)/1000 as plan_march,
				(gr.plan_4 * pr.price)/1000 as plan_april,
				(gr.plan_5 * pr.price)/1000 as plan_may,
				(gr.plan_6 * pr.price)/1000 as plan_june,
				(gr.plan_7 * pr.price)/1000 as plan_july,
				(gr.plan_8 * pr.price)/1000 as plan_aug,
				(gr.plan_9 * pr.price)/1000 as plan_sept,
				(gr.plan_10 * pr.price)/1000 as plan_oct,
				(gr.plan_11 * pr.price)/1000 as plan_novem,
				(gr.plan_12 * pr.price)/1000 as plan_decem,
				pr.price,
				pr.group_name
			FROM
				plangr as gr
			LEFT JOIN
				grinfo as pr
			ON
				gr.id_org = pr.id_org
				and gr.id_group = pr.id_group),
	all_visit as(
            SELECT
                iin,
                datestatus,
                id_org,
				id_group,
                timestatus,
                status
            FROM
                serviceback_visits),
	visit as(
            SELECT
                id_org,
                iin,
                datestatus,
                max(timestatus) as timestatus
            FROM
                all_visit
            GROUP BY
                id_org,
                iin,
                datestatus),
    count_visit as(
            SELECT
                id_org,
				id_group,
                iin,
                datestatus,
                timestatus,
                status
            FROM
                all_visit
            WHERE 
                ((id_org,
                iin,
                datestatus,
                timestatus) IN (
                    SELECT
                        id_org,
                        iin,
                        datestatus,
                        timestatus
                    FROM
                        visit)
                and status='2' or status='3' or status='4')),
	count_all as(
            SELECT
                id_org,
				id_group,
                cast(date_trunc('month', datestatus) as date) as date_vis,
                sum(case when status='2' then 1 else 0 end) as count_v,
                sum(case when status='3' then 1 else 0 end) as count_b,
                sum(case when status='4' then 1 else 0 end) as count_o
            FROM
                count_visit
            GROUP BY
                id_org,
				id_group,
                date_vis),
	countandprice as(			
			SELECT
				co.id_org,
				co.id_group,
				co.date_vis,
				(co.count_v * pr.price)/1000 as count_v,
				(co.count_b * pr.price)/1000 as count_b,
				(co.count_o * pr.price)/1000 as count_o,
				pr.price
			FROM 
				count_all as co
			LEFT JOIN
				grinfo as pr
			ON
				co.id_org = pr.id_org
				and co.id_group = pr.id_group),
	all_count as(
            SELECT
                id_org as id_org,
				id_group as id_group,
                sum(case when date_vis = '2023-01-01' then count_v else 0 end) as v_jan,
                sum(case when date_vis = '2023-01-01' then count_b else 0 end) as b_jan,
                sum(case when date_vis = '2023-01-01' then count_o else 0 end) as o_jan,
                sum(case when date_vis = '2023-02-01' then count_v else 0 end) as v_feb,
                sum(case when date_vis = '2023-02-01' then count_b else 0 end) as b_feb,
                sum(case when date_vis = '2023-02-01' then count_o else 0 end) as o_feb,
                sum(case when date_vis = '2023-03-01' then count_v else 0 end) as v_march,
                sum(case when date_vis = '2023-03-01' then count_b else 0 end) as b_march,
                sum(case when date_vis = '2023-03-01' then count_o else 0 end) as o_march,
                sum(case when date_vis = '2023-04-01' then count_v else 0 end) as v_april,
                sum(case when date_vis = '2023-04-01' then count_b else 0 end) as b_april,
                sum(case when date_vis = '2023-04-01' then count_o else 0 end) as o_april,
                sum(case when date_vis = '2023-05-01' then count_v else 0 end) as v_may,
                sum(case when date_vis = '2023-05-01' then count_b else 0 end) as b_may,
                sum(case when date_vis = '2023-05-01' then count_o else 0 end) as o_may,
                sum(case when date_vis = '2023-06-01' then count_v else 0 end) as v_june,
                sum(case when date_vis = '2023-06-01' then count_b else 0 end) as b_june,
                sum(case when date_vis = '2023-06-01' then count_o else 0 end) as o_june,
                sum(case when date_vis = '2023-07-01' then count_v else 0 end) as v_july,
                sum(case when date_vis = '2023-07-01' then count_b else 0 end) as b_july,
                sum(case when date_vis = '2023-07-01' then count_o else 0 end) as o_july,
                sum(case when date_vis = '2023-08-01' then count_v else 0 end) as v_aug,
                sum(case when date_vis = '2023-08-01' then count_b else 0 end) as b_aug,
                sum(case when date_vis = '2023-08-01' then count_o else 0 end) as o_aug,
                sum(case when date_vis = '2023-09-01' then count_v else 0 end) as v_sept,
                sum(case when date_vis = '2023-09-01' then count_b else 0 end) as b_sept,
                sum(case when date_vis = '2023-09-01' then count_o else 0 end) as o_sept,
                sum(case when date_vis = '2023-10-01' then count_v else 0 end) as v_oct,
                sum(case when date_vis = '2023-10-01' then count_b else 0 end) as b_oct,
                sum(case when date_vis = '2023-10-01' then count_o else 0 end) as o_oct,
                sum(case when date_vis = '2023-11-01' then count_v else 0 end) as v_novem,
                sum(case when date_vis = '2023-11-01' then count_b else 0 end) as b_novem,
                sum(case when date_vis = '2023-11-01' then count_o else 0 end) as o_novem,
                sum(case when date_vis = '2023-12-01' then count_v else 0 end) as v_decem,
                sum(case when date_vis = '2023-12-01' then count_b else 0 end) as b_decem,
                sum(case when date_vis = '2023-12-01' then count_o else 0 end) as o_decem
            FROM
                countandprice
            GROUP BY 
                id_org,
				id_group)
			SELECT
				f.id_obl,
				(case when f.id_obl is NULL then 1 else 0 end) as por_obl,
				max(f.obl_name) as obl_name,
				f.id_region,
				(case when f.id_region is NULL then 1 else 0 end) as por_reg,
				max(f.reg_name) as reg_name,
				max(f.type_org) as type_name,
				(case when f.id_region is NULL then 0
					 when f.type_org='Частный' then f.id_region+200
					 else f.id_region+100 end) as type_id,
				allc.id_org as id_org,
				max(f.org_name) as org_name,
				allc.id_group as id_group,
				max(gr.group_name) as group_name,
				sum(1) as count_gr,
            	sum(cast(countgr.group_count as integer)) as plan_count,
            	sum(cast(countgr.group_count as integer)) as fact_count,
				sum(cast(gr.plan_jan as integer)) as plan_1,
				sum(cast(allc.v_jan as integer)) as v_1,
				sum(cast(allc.b_jan as integer)) as b_1,
				sum(cast(allc.o_jan as integer)) as o_1,
				sum(cast((gr.plan_jan - allc.v_jan - allc.b_jan - allc.o_jan) as integer)) as pr_1,
				sum(cast(gr.plan_feb as integer)) as plan_2,
				sum(cast(allc.v_feb as integer)) as v_2,
				sum(cast(allc.b_feb as integer)) as b_2,
				sum(cast(allc.o_feb as integer)) as o_2,
				sum(cast((gr.plan_feb - allc.v_feb - allc.b_feb - allc.o_feb) as integer)) as pr_2,
				sum(cast(gr.plan_march as integer)) as plan_3,
				sum(cast(allc.v_march as integer)) as v_3,
				sum(cast(allc.b_march as integer)) as b_3,
				sum(cast(allc.o_march as integer)) as o_3,
				sum(cast((gr.plan_march - allc.v_march - allc.b_march - allc.o_march) as integer)) as pr_3,
				sum(cast(gr.plan_april as integer)) as plan_4,
				sum(cast(allc.v_april as integer)) as v_4,
				sum(cast(allc.b_april as integer)) as b_4,
				sum(cast(allc.o_april as integer)) as o_4,
				sum(cast((gr.plan_april - allc.v_april - allc.b_april - allc.o_april) as integer)) as pr_4,
				sum(cast(gr.plan_may as integer)) as plan_5,
				sum(cast(allc.v_may as integer)) as v_5,
				sum(cast(allc.b_may as integer)) as b_5,
				sum(cast(allc.o_may as integer)) as o_5,
				sum(cast((gr.plan_may - allc.v_may - allc.b_may - allc.o_may) as integer)) as pr_5,
				sum(cast(gr.plan_june as integer)) as plan_6,
				sum(cast(allc.v_june as integer)) as v_6,
				sum(cast(allc.b_june as integer)) as b_6,
				sum(cast(allc.o_june as integer)) as o_6,
				sum(cast((gr.plan_june - allc.v_june - allc.b_june - allc.o_june) as integer)) as pr_6,
				sum(cast(gr.plan_july as integer)) as plan_7,
				sum(cast(allc.v_july as integer)) as v_7,
				sum(cast(allc.b_july as integer)) as b_7,
				sum(cast(allc.o_july as integer)) as o_7,
				sum(cast((gr.plan_july - allc.v_july - allc.b_july - allc.o_july) as integer)) as pr_7,
				sum(cast(gr.plan_aug as integer)) as plan_8,
				sum(cast(allc.v_aug as integer)) as v_8,
				sum(cast(allc.b_aug as integer)) as b_9,
				sum(cast(allc.o_aug as integer)) as o_9,
				sum(cast((gr.plan_aug - allc.v_aug - allc.b_aug - allc.o_aug) as integer)) as pr_8,
				sum(cast(gr.plan_sept as integer)) as plan_9,
				sum(cast(allc.v_sept as integer)) as v_9,
				sum(cast(allc.b_sept as integer)) as b_9,
				sum(cast(allc.o_sept as integer)) as o_9,
				sum(cast((gr.plan_sept - allc.v_sept - allc.b_sept - allc.o_sept) as integer)) as pr_9,
				sum(cast(gr.plan_oct as integer)) as plan_10,
				sum(cast(allc.v_oct as integer)) as v_10,
				sum(cast(allc.b_oct as integer)) as b_10,
				sum(cast(allc.o_oct as integer)) as o_10,
				sum(cast((gr.plan_oct - allc.v_oct - allc.b_oct - allc.o_oct) as integer)) as pr_10,
				sum(cast(gr.plan_novem as integer)) as plan_11,
				sum(cast(allc.v_novem as integer)) as v_11,
				sum(cast(allc.b_novem as integer)) as b_11,
				sum(cast(allc.o_novem as integer)) as o_11,
				sum(cast((gr.plan_novem - allc.v_novem - allc.b_novem - allc.o_novem) as integer)) as pr_11,
				sum(cast(gr.plan_decem as integer)) as plan_12,
				sum(cast(allc.v_decem as integer)) as v_12,
				sum(cast(allc.b_decem as integer)) as b_12,
				sum(cast(allc.o_decem as integer)) as o_12,
				sum(cast((gr.plan_decem - allc.v_decem - allc.b_decem - allc.o_decem) as integer)) as pr_12
			FROM
				all_count as allc
			INNER JOIN
				planandprice as gr
			ON
				allc.id_org = gr.id_org
				and allc.id_group = gr.id_group
			LEFT JOIN
				full_info as f
			ON
				allc.id_org = f.idd_org
			LEFT JOIN
				c_gr as countgr
			ON
				allc.id_org = countgr.id_org
				and allc.id_group = countgr.id_group
			GROUP BY
			GROUPING SETS ((f.id_obl), (id_region), (type_id), (allc.id_org), ((allc.id_group)))			
			ORDER BY por_obl, por_reg, type_id"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    json_data = []

    for row in result:
        context = {}
        id_obl = row[0]
        obl_name = row[2]
        id_region = row[3]
        reg_name =  row[5]
        type_name = row[6]
        type_id = row[7]
        id_org = row[8]
        org_name = row[9]
        id_group = row[10]
        group_name = row[11]
        count_gr = row[12]
        plan_count = row[13]
        fact_count = row[14]
        for i in range(1,13):
            ii = (14 + i) + (4*(i-1))            
            context['plan_{}'.format(i)] = row[ii]
            context['v_{}'.format(i)] = row[ii+1]
            context['b_{}'.format(i)] = row[ii+2]
            context['o_{}'.format(i)] = row[ii+3]
            context['pr_{}'.format(i)] = row[ii+4]
         

        if id_obl is not None and id_obl not in json_data:
            obl_data = {
                'data': {
                    'sub': 1,
                    'id_obl': id_obl,
                    'obl_name': obl_name,
                    'count_gr': count_gr,
                    'plan_count': plan_count,
                    'fact_count': fact_count
                },
                'children': []
            }

            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    obl_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    obl_data['data'][v_key] = context[v_key]
                if b_key in context:
                    obl_data['data'][b_key] = context[b_key]
                if o_key in context:
                    obl_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    obl_data['data'][pr_key] = context[pr_key]
            json_data.append(obl_data)

        if id_region is not None:
            region_data = {
                'data': {
                            'sub': 2,
                            'id_obl': id_region,
                            'obl_name': reg_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        },
                'children': []
            }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    region_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    region_data['data'][v_key] = context[v_key]
                if b_key in context:
                    region_data['data'][b_key] = context[b_key]
                if o_key in context:
                    region_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    region_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    obl['children'].append(region_data)
                    break

        if type_id is not None:
            type_data = {
                'data': {
                            'sub': 3,
                            'id_obl': type_id,
                            'obl_name': type_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        },
                'children': []
            }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    type_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    type_data['data'][v_key] = context[v_key]
                if b_key in context:
                    type_data['data'][b_key] = context[b_key]
                if o_key in context:
                    type_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    type_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    for region in obl['children']:
                        if region['data']['obl_name'] == reg_name:
                            region['children'].append(type_data)
                            break

        if id_org is not None:
            org_data = {
                        'data': {
                            'sub': 4,
                            'id_obl': id_org,
                            'obl_name': org_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        },
                        'children': []
                    }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    org_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    org_data['data'][v_key] = context[v_key]
                if b_key in context:
                    org_data['data'][b_key] = context[b_key]
                if o_key in context:
                    org_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    org_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    for region in obl['children']:
                        if region['data']['obl_name'] == reg_name:
                            for type_o in region['children']:
                                if type_o['data']['obl_name'] == type_name:
                                    type_o['children'].append(org_data)
                                    break


        if id_group is not None:
            group_data = {
                        'data': {
                            'sub': 5,
                            'id_obl': id_group,
                            'obl_name': group_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        }
                    }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    group_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    group_data['data'][v_key] = context[v_key]
                if b_key in context:
                    group_data['data'][b_key] = context[b_key]
                if o_key in context:
                    group_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    group_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    for region in obl['children']:
                        if region['data']['obl_name'] == reg_name:
                            for type_o in region['children']:
                                if type_o['data']['obl_name'] == type_name:
                                    for org in type_o['children']:
                                        if org['data']['obl_name'] == org_name:
                                            org['children'].append(group_data)
                                            break       

    json_result = json.dumps(json_data, indent=4, default=str)

    return HttpResponse(json_result, content_type="application/json")

@api_view(['GET'])
def getpriceservice(request):
    
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    query =f"""
        SELECT 
            id, 
            name 
        FROM
            serviceback_regions
        WHERE
            id_parent=0
        """

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result, indent=4, sort_keys=True, default=str), content_type="application/json")

@api_view(['GET'])
def getpriceobl(request):
    
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    id_obl    = request.GET.get("id_obl")

    query =f"""
                with allprice as(
                        SELECT
                            *
                        FROM
                            serviceback_priceservice
                        WHERE
                            obl_id = '{id_obl}'),
                tmp_nepoln as(
                        SELECT
                            max(cast(1 as integer)) as poryadok,
                            max(cast('gorp' as text)) as category,
                            max(cast('Группы с неполным днем пребывания,  классы предшкольной подготовки при общеобразовательной школе' as text)) as name_price,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then price else 0 end) as norm_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then id else 0 end) as id_norm_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then price else 0 end) as norm_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then id else 0 end) as id_norm_selo,                            
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then price else 0 end) as eco_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then id else 0 end) as id_eco_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then price else 0 end) as eco_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then id else 0 end) as id_eco_selo,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then price else 0 end) as rad_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then id else 0 end) as id_rad_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then price else 0 end) as rad_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then id else 0 end) as id_rad_selo
                        FROM
                            allprice
                        WHERE
                            category='gorp'),
                tmp_gor9 as(
                        SELECT
                            max(cast(2 as integer)) as poryadok,
                            max(cast('gor9' as text)) as category,
                            max(cast('Группа с 9-часовым режимом пребывания' as text)) as name_price,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then price else 0 end) as norm_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then id else 0 end) as id_norm_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then price else 0 end) as norm_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then id else 0 end) as id_norm_selo,                            
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then price else 0 end) as eco_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then id else 0 end) as id_eco_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then price else 0 end) as eco_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then id else 0 end) as id_eco_selo,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then price else 0 end) as rad_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then id else 0 end) as id_rad_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then price else 0 end) as rad_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then id else 0 end) as id_rad_selo
                        FROM
                            allprice
                        WHERE
                            category='gor9'),
                tmp_gor10 as(
                        SELECT
                            max(cast(3 as integer)) as poryadok,
                            max(cast('gor10' as text)) as category,
                            max(cast('Группа с 10,5-часовым режимом пребывания' as text)) as name_price,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then price else 0 end) as norm_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then id else 0 end) as id_norm_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then price else 0 end) as norm_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then id else 0 end) as id_norm_selo,                            
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then price else 0 end) as eco_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then id else 0 end) as id_eco_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then price else 0 end) as eco_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then id else 0 end) as id_eco_selo,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then price else 0 end) as rad_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then id else 0 end) as id_rad_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then price else 0 end) as rad_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then id else 0 end) as id_rad_selo
                        FROM
                            allprice
                        WHERE
                            category='gor10'),
                tmp_gors as(
                        SELECT
                            max(cast(4 as integer)) as poryadok,
                            max(cast('gors' as text)) as category,
                            max(cast('Санаторная группа' as text)) as name_price,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then price else 0 end) as norm_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then id else 0 end) as id_norm_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then price else 0 end) as norm_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then id else 0 end) as id_norm_selo,                            
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then price else 0 end) as eco_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then id else 0 end) as id_eco_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then price else 0 end) as eco_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then id else 0 end) as id_eco_selo,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then price else 0 end) as rad_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then id else 0 end) as id_rad_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then price else 0 end) as rad_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then id else 0 end) as id_rad_selo
                        FROM
                            allprice
                        WHERE
                            category='gors'),
                tmp_gork as(
                        SELECT
                            max(cast(5 as integer)) as poryadok,
                            max(cast('gork' as text)) as category,
                            max(cast('Коррекционная группа' as text)) as name_price,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then price else 0 end) as norm_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'normal' then id else 0 end) as id_norm_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then price else 0 end) as norm_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'normal' then id else 0 end) as id_norm_selo,                            
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then price else 0 end) as eco_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'eco' then id else 0 end) as id_eco_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then price else 0 end) as eco_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'eco' then id else 0 end) as id_eco_selo,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then price else 0 end) as rad_gor,
                            max(case when type_city = 'gor' and type_ecolog = 'rad' then id else 0 end) as id_rad_gor,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then price else 0 end) as rad_selo,
                            max(case when type_city = 'selo' and type_ecolog = 'rad' then id else 0 end) as id_rad_selo
                        FROM
                            allprice
                        WHERE
                            category='gork')
                select * FROM tmp_gor9
                union
                select * FROM tmp_nepoln
                union
                select * FROM tmp_gor10
                union
                select * FROM tmp_gors
                union
                select * FROM tmp_gork
                order by poryadok
        """

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

    return HttpResponse(json.dumps(result, indent=4, sort_keys=True, default=str), content_type="application/json")


@api_view(['POST'])
def setpriceobl(request):
    
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    datastr = request.body
    res = json.loads(datastr)

    id_obl = res['id_obl']
    price  = res['result']
    for item in price:
        try:
            price_g_normal = PriceService.objects.get(id=item['id_norm_gor'])
            price_g_normal.price = item['norm_gor']
            price_g_normal.save()
        except:
            a=0
            # new_g_normal = PriceService()
            # new_g_normal.obl = id_obl
            # new_g_normal.category = 'gorp'
            # new_g_normal.type_city = 'gor'
            # new_g_normal.type_ecolog = 'normal'
            # new_g_normal.price = item['norm_gor']
            # new_g_normal.save()

        try:
            price_s_normal = PriceService.objects.get(id=item['id_norm_selo'])
            price_s_normal.price = item['norm_selo']
            price_s_normal.save()
        except:
            a=0
            # new_s_normal = PriceService()
            # new_s_normal.obl = id_obl
            # new_s_normal.category = 'gorp'
            # new_s_normal.type_city = 'selo'
            # new_s_normal.type_ecolog = 'normal'
            # new_s_normal.price = item['norm_selo']
            # new_s_normal.save()
        
        try:
            price_g_eco = PriceService.objects.get(id=item['id_eco_gor'])
            price_g_eco.price = item['eco_gor']
            price_g_eco.save()
        except:
            a=0
            # new_g_eco = PriceService()
            # new_g_normal.obl = id_obl
            # new_g_normal.category = 'gorp'
            # new_g_normal.type_city = 'gor'
            # new_g_normal.type_ecolog = 'normal'
            # new_g_normal.price = item['norm_gor']
            # new_g_normal.save()
        
        try:
            price_s_eco = PriceService.objects.get(id=item['id_eco_selo'])
            price_s_eco.price = item['eco_selo']
            price_s_eco.save()
        except:
            a=0
        
        try:
            price_g_rad = PriceService.objects.get(id=item['id_rad_gor'])
            price_g_rad.price = item['rad_gor']
            price_g_rad.save()
        except:
            a=0
        
        try:
            price_s_rad = PriceService.objects.get(id=item['id_rad_selo'])
            price_s_rad.price = item['rad_selo']
            price_s_rad.save()
        except:
            a=0
        
        return HttpResponse('{"status": "Данные успешно сохранены"}', content_type="application/json", status=200)

@api_view(['GET'])
def formfordash(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    id_org = request.GET.get("id_org")
    id_region = request.GET.get("id_region")
    current_date = datetime.date.today()

    query = f"""WITH org as(
            SELECT
                id_org,
                org_name,
                (case when type_org = 'pr' then 'Частный'
                    else 'Государственный' end) as type_org,
                id_obl,
                id_region
            FROM
                serviceback_organizations
			where
				id_obl = '1'
                and case when '{id_region}' = 0 then true
					else id_region = '{id_region}' end
				and case when '{id_org}' = '' then true
					else id_org = '{id_org}' end
				and not id_org = '910114301692'),
        allreg as(
                SELECT
                    *
                FROM
                    serviceback_regions),
        namereg as(
                SELECT
                    reg.id as reg_id,
                    reg.name as reg_name
                FROM
                    allreg as reg
                WHERE
                    id in(
                        SELECT
                            id_region
                        FROM
                            org)),
        nameobl as(
                SELECT
                    obl.id as obl_id,
                    obl.name as obl_name
                FROM
                    allreg as obl
                WHERE
                    id in(
                        SELECT
                            id_obl
                        FROM
                            org)),
        full_info as(		
                SELECT 
                    org.id_org,
                    org.org_name,
                    org.type_org,
                    org.id_obl,
                    org.id_region,
                    nameobl.obl_name,
                    namereg.reg_name
                FROM 
                    org as org
                LEFT JOIN
                    nameobl as nameobl
                ON
                    org.id_obl = nameobl.obl_id
                LEFT JOIN
                    namereg as namereg
                ON
                    org.id_region = namereg.reg_id),
        count_gr as(
            SELECT 
                id_org,
                COUNT(*) as count_gr,
                sum(group_count) as plan_count
            FROM
                serviceback_groups
			WHERE
				id_org in(
                        SELECT
                            id_org
                        FROM
                            org)
				and not is_delete
            GROUP BY
                id_org),
		count_week as(
            SELECT 
                date_trunc('month', weekend) as mon,
                sum(1) as count_w
            FROM 
                serviceback_weekendday
			WHERE
				weekend <= '{current_date}'
            GROUP BY
                mon),
        count_workday as(
            SELECT 
                mon,
                case when not date_trunc('month', '{current_date}'::date) = date_trunc('MONTH', mon::date)
						then extract(DAY FROM (date_trunc('MONTH', mon::date) + INTERVAL '1 MONTH' - INTERVAL '1 DAY')) - count_w
					else date_part('day', '{current_date}'::date) - count_w  end as work_day
            FROM 
                count_week),
		groupandplan as(
            SELECT 
                countt.id_org,
                max(countt.count_gr) as count_gr,
                max(countt.plan_count) as plan_count,
                sum(cast(case when w.mon = '2023-01-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_jan,
                sum(cast(case when w.mon = '2023-02-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_feb,
                sum(cast(case when w.mon = '2023-03-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_march,
                sum(cast(case when w.mon = '2023-04-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_april,
                sum(cast(case when w.mon = '2023-05-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_may,
                sum(cast(case when w.mon = '2023-06-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_june,
                sum(cast(case when w.mon = '2023-07-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_july,
                sum(cast(case when w.mon = '2023-08-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_aug,
                sum(cast(case when w.mon = '2023-09-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_sept,
                sum(cast(case when w.mon = '2023-10-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_oct,
                sum(cast(case when w.mon = '2023-11-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_novem,
                sum(cast(case when w.mon = '2023-12-01' then countt.plan_count * w.work_day else 0 end as numeric)) as plan_decem
            FROM 
                count_gr as countt,
				count_workday as w
            GROUP BY
                id_org),
	visit as(
           SELECT
              id as _id,
              id_org,
              id_group,
              iin,
			  datestatus,
              status
           FROM
              serviceback_visits
           WHERE
              (id_org) in (
                      SELECT
                         id_org
                      FROM
                         org)
              and datestatus = '{current_date}'),
   max_vis_id as(
          SELECT
			  max(_id) as _id,
			  iin
          FROM
          	  visit
          GROUP BY
              iin),
   vis_finall as(
         SELECT
             *
         FROM
             visit
         WHERE
             _id in(
                SELECT
                   _id
                FROM
                   max_vis_id)),
                                    tmp_finall as(
                                            SELECT 
                                                id_org,
												datestatus,
                                                CASE
                                                    WHEN status = '2'  THEN 1
                                                    ELSE 0
                                                END as visited,
                                                CASE
                                                    WHEN status = '3'  THEN 1
                                                    ELSE 0
                                                END as bolnich,
                                                CASE
                                                    WHEN status = '4' THEN 1
                                                    ELSE 0
                                                END as otpusk,
                                                status as common
                                            FROM 
                                                vis_finall),
                                        tmp_fin as(
                                            SELECT
                                                id_org,
												datestatus,
                                                sum(visited) as vis, 
                                                sum(bolnich) as boln, 
                                                sum(otpusk) as otp,
                                                count(common) as common
                                            FROM
                                                tmp_finall
                                            GROUP BY
                                                id_org,
												datestatus),
										tmp_all_rec as(                                            
											SELECT
                                                fin.id_org,
												date_trunc('month', fin.datestatus::date) as datestatus,												
                                                fin.vis as visited,
                                                fin.boln as bolnich,
                                                fin.otp as otpusk,
                                                fin.common - fin.vis - fin.boln - fin.otp as not_vis
                                            FROM
                                                tmp_fin as fin),
   		itogbym as(
			SELECT
				*
			FROM
				serviceback_itogbymonth
			WHERE
				id_org in(
                        SELECT
                            id_org
                        FROM
                            org)),
		addtoday as(
			SELECT
				i.id_org,
				i.datestatus,
				CASE WHEN tmp.visited is null THEN i.visit
					 ELSE tmp.visited + i.visit
				END as visit,
				CASE WHEN tmp.bolnich is null THEN i.boln
					 ELSE tmp.bolnich + i.boln
				END as boln,
				CASE WHEN tmp.otpusk is null THEN i.otpusk
					 ELSE tmp.otpusk + i.otpusk
				END as otpusk,
				CASE WHEN tmp.not_vis is null THEN i.notvisit
					 ELSE tmp.not_vis + i.notvisit
				END as notvisit
			FROM
				itogbym as i
			LEFT JOIN
				tmp_all_rec as tmp
			ON
				i.id_org = tmp.id_org
				and i.datestatus = tmp.datestatus),
        all_count as(
            SELECT
                id_org,
                sum(case when datestatus = '2023-01-01' then visit else 0 end) as v_1,
                sum(case when datestatus = '2023-01-01' then boln else 0 end) as b_1,
                sum(case when datestatus = '2023-01-01' then otpusk else 0 end) as o_1,
				sum(case when datestatus = '2023-01-01' then notvisit else 0 end) as pr_1,
                sum(case when datestatus = '2023-02-01' then visit else 0 end) as v_2,
                sum(case when datestatus = '2023-02-01' then boln else 0 end) as b_2,
                sum(case when datestatus = '2023-02-01' then otpusk else 0 end) as o_2,
				sum(case when datestatus = '2023-02-01' then notvisit else 0 end) as pr_2,
                sum(case when datestatus = '2023-03-01' then visit else 0 end) as v_3,
                sum(case when datestatus = '2023-03-01' then boln else 0 end) as b_3,
                sum(case when datestatus = '2023-03-01' then otpusk else 0 end) as o_3,
				sum(case when datestatus = '2023-03-01' then notvisit else 0 end) as pr_3,
                sum(case when datestatus = '2023-04-01' then visit else 0 end) as v_4,
                sum(case when datestatus = '2023-04-01' then boln else 0 end) as b_4,
                sum(case when datestatus = '2023-04-01' then otpusk else 0 end) as o_4,
				sum(case when datestatus = '2023-04-01' then notvisit else 0 end) as pr_4,
                sum(case when datestatus = '2023-05-01' then visit else 0 end) as v_5,
                sum(case when datestatus = '2023-05-01' then boln else 0 end) as b_5,
                sum(case when datestatus = '2023-05-01' then otpusk else 0 end) as o_5,
				sum(case when datestatus = '2023-05-01' then notvisit else 0 end) as pr_5,
                sum(case when datestatus = '2023-06-01' then visit else 0 end) as v_6,
                sum(case when datestatus = '2023-06-01' then boln else 0 end) as b_6,
                sum(case when datestatus = '2023-06-01' then otpusk else 0 end) as o_6,
				sum(case when datestatus = '2023-06-01' then notvisit else 0 end) as pr_6,
                sum(case when datestatus = '2023-07-01' then visit else 0 end) as v_7,
                sum(case when datestatus = '2023-07-01' then boln else 0 end) as b_7,
                sum(case when datestatus = '2023-07-01' then otpusk else 0 end) as o_7,
				sum(case when datestatus = '2023-07-01' then notvisit else 0 end) as pr_7,
                sum(case when datestatus = '2023-08-01' then visit else 0 end) as v_8,
                sum(case when datestatus = '2023-08-01' then boln else 0 end) as b_8,
                sum(case when datestatus = '2023-08-01' then otpusk else 0 end) as o_8,
				sum(case when datestatus = '2023-08-01' then notvisit else 0 end) as pr_8,
                sum(case when datestatus = '2023-09-01' then visit else 0 end) as v_9,
                sum(case when datestatus = '2023-09-01' then boln else 0 end) as b_9,
                sum(case when datestatus = '2023-09-01' then otpusk else 0 end) as o_9,
				sum(case when datestatus = '2023-09-01' then notvisit else 0 end) as pr_9,
                sum(case when datestatus = '2023-10-01' then visit else 0 end) as v_10,
                sum(case when datestatus = '2023-10-01' then boln else 0 end) as b_10,
                sum(case when datestatus = '2023-10-01' then otpusk else 0 end) as o_10,
				sum(case when datestatus = '2023-10-01' then notvisit else 0 end) as pr_10,
                sum(case when datestatus = '2023-11-01' then visit else 0 end) as v_11,
                sum(case when datestatus = '2023-11-01' then boln else 0 end) as b_11,
                sum(case when datestatus = '2023-11-01' then otpusk else 0 end) as o_11,
				sum(case when datestatus = '2023-11-01' then notvisit else 0 end) as pr_11,
                sum(case when datestatus = '2023-12-01' then visit else 0 end) as v_12,
                sum(case when datestatus = '2023-12-01' then boln else 0 end) as b_12,
                sum(case when datestatus = '2023-12-01' then otpusk else 0 end) as o_12,
				sum(case when datestatus = '2023-12-01' then notvisit else 0 end) as pr_12
            FROM
                addtoday
			GROUP BY
				id_org),
	tmp_finalll as(
			SELECT
				f.id_obl,
				(case when f.id_obl is NULL then 1 else 0 end) as por_obl,
				max(f.obl_name) as obl_name,
				f.id_region,
				(case when f.id_region is NULL then 1 else 0 end) as por_reg,
				max(f.reg_name) as reg_name,
				allc.id_org as id_org,
				max(f.org_name) as org_name,
				sum(cast(gr.count_gr as integer)) as count_gr,
				sum(cast(gr.plan_count as integer)) as plan_count,
				sum(cast(gr.plan_count as integer)) as fact_count,
				sum(cast(gr.plan_jan as integer)) as plan_1,
				sum(cast(allc.v_1 as integer)) as v_1,
				sum(cast(allc.b_1 as integer)) as b_1,
				sum(cast(allc.o_1 as integer)) as o_1,
				sum(cast(allc.pr_1 as integer)) as pr_1,
				sum(cast(gr.plan_feb as integer)) as plan_2,
				sum(cast(allc.v_2 as integer)) as v_2,
				sum(cast(allc.b_2 as integer)) as b_2,
				sum(cast(allc.o_2 as integer)) as o_2,
				sum(cast(allc.pr_2 as integer)) as pr_2,
				sum(cast(gr.plan_march as integer)) as plan_3,
				sum(cast(allc.v_3 as integer)) as v_3,
				sum(cast(allc.b_3 as integer)) as b_3,
				sum(cast(allc.o_3 as integer)) as o_3,
				sum(cast(allc.pr_3 as integer)) as pr_3,
				sum(cast(gr.plan_april as integer)) as plan_4,
				sum(cast(allc.v_4 as integer)) as v_4,
				sum(cast(allc.b_4 as integer)) as b_4,
				sum(cast(allc.o_4 as integer)) as o_4,
				sum(cast(allc.pr_4 as integer)) as pr_4,
				sum(cast(gr.plan_may as integer)) as plan_5,
				sum(cast(allc.v_5 as integer)) as v_5,
				sum(cast(allc.b_5 as integer)) as b_5,
				sum(cast(allc.o_5 as integer)) as o_5,
				sum(cast(allc.pr_5 as integer)) as pr_5,
				sum(cast(gr.plan_june as integer)) as plan_6,
				sum(cast(allc.v_6 as integer)) as v_6,
				sum(cast(allc.b_6 as integer)) as b_6,
				sum(cast(allc.o_6 as integer)) as o_6,
				sum(cast(allc.pr_6 as integer)) as pr_6,
				sum(cast(gr.plan_july as integer)) as plan_7,
				sum(cast(allc.v_7 as integer)) as v_7,
				sum(cast(allc.b_7 as integer)) as b_7,
				sum(cast(allc.o_7 as integer)) as o_7,
				sum(cast(allc.pr_7 as integer)) as pr_7,
				sum(cast(gr.plan_aug as integer)) as plan_8,
				sum(cast(allc.v_8 as integer)) as v_8,
				sum(cast(allc.b_8 as integer)) as b_8,
				sum(cast(allc.o_8 as integer)) as o_8,
				sum(cast(allc.pr_8 as integer)) as pr_8,
				sum(cast(gr.plan_sept as integer)) as plan_9,
				sum(cast(allc.v_9 as integer)) as v_9,
				sum(cast(allc.b_9 as integer)) as b_9,
				sum(cast(allc.o_9 as integer)) as o_9,
				sum(cast(allc.pr_9 as integer)) as pr_9,
				sum(cast(gr.plan_oct as integer)) as plan_10,
				sum(cast(allc.v_10 as integer)) as v_10,
				sum(cast(allc.b_10 as integer)) as b_10,
				sum(cast(allc.o_10 as integer)) as o_10,
				sum(cast(allc.pr_10 as integer)) as pr_10,
				sum(cast(gr.plan_novem as integer)) as plan_11,
				sum(cast(allc.v_11 as integer)) as v_11,
				sum(cast(allc.b_11 as integer)) as b_11,
				sum(cast(allc.o_11 as integer)) as o_11,
				sum(cast(allc.pr_11 as integer)) as pr_11,
				sum(cast(gr.plan_decem as integer)) as plan_12,
				sum(cast(allc.v_12 as integer)) as v_12,
				sum(cast(allc.b_12 as integer)) as b_12,
				sum(cast(allc.o_12 as integer)) as o_12,
				sum(cast(allc.pr_12 as integer)) as pr_12,
				max(f.type_org) as type_name,
				(case when f.id_region is NULL then 0
					 when f.type_org='Частный' then f.id_region+200
					 else f.id_region+100 end) as type_id
			FROM
				all_count as allc
			INNER JOIN
				groupandplan as gr
			ON
				allc.id_org = gr.id_org
			LEFT JOIN
				full_info as f
			ON
				allc.id_org = f.id_org
			GROUP BY
			GROUPING SETS ((f.id_obl), (f.id_region), (type_id), (allc.id_org))
			ORDER BY por_obl, por_reg, type_id)
			SELECT * FROM tmp_finalll"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        results = cursor.fetchall()

    json_data = []

    for row in results:
        context = {}
        id_obl = row[0]
        obl_name = row[2]
        id_region = row[3]
        reg_name =  row[5]
        id_org = row[6]
        org_name = row[7]
        count_gr = row[8]
        plan_count = row[9]
        fact_count = row[10]
        for i in range(1,13):
            ii = (10 + i) + (4*(i-1))            
            context['plan_{}'.format(i)] = row[ii]
            context['v_{}'.format(i)] = row[ii+1]
            context['b_{}'.format(i)] = row[ii+2]
            context['o_{}'.format(i)] = row[ii+3]
            context['pr_{}'.format(i)] = row[ii+4]
        type_name = row[ii+5]
        type_id = row[ii+6] 

        if id_obl is not None and id_obl not in json_data:
            obl_data = {
                'data': {
                    'sub': 1,
                    'id_obl': id_obl,
                    'obl_name': obl_name,
                    'count_gr': count_gr,
                    'plan_count': plan_count,
                    'fact_count': fact_count
                },
                'children': []
            }

            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    obl_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    obl_data['data'][v_key] = context[v_key]
                if b_key in context:
                    obl_data['data'][b_key] = context[b_key]
                if o_key in context:
                    obl_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    obl_data['data'][pr_key] = context[pr_key]
            json_data.append(obl_data)

        if id_region is not None:
            region_data = {
                'data': {
                            'sub': 2,
                            'id_obl': id_region,
                            'obl_name': reg_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        },
                'children': []
            }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    region_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    region_data['data'][v_key] = context[v_key]
                if b_key in context:
                    region_data['data'][b_key] = context[b_key]
                if o_key in context:
                    region_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    region_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    obl['children'].append(region_data)
                    break

        if type_id is not None:
            type_data = {
                'data': {
                            'sub': 3,
                            'id_obl': type_id,
                            'obl_name': type_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        },
                'children': []
            }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    type_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    type_data['data'][v_key] = context[v_key]
                if b_key in context:
                    type_data['data'][b_key] = context[b_key]
                if o_key in context:
                    type_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    type_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    for region in obl['children']:
                        if region['data']['obl_name'] == reg_name:
                            region['children'].append(type_data)
                            break

        if id_org is not None:
            org_data = {
                        'data': {
                            'sub': 4,
                            'id_obl': id_org,
                            'obl_name': org_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        }
                    }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    org_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    org_data['data'][v_key] = context[v_key]
                if b_key in context:
                    org_data['data'][b_key] = context[b_key]
                if o_key in context:
                    org_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    org_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    for region in obl['children']:
                        if region['data']['obl_name'] == reg_name:
                            for type_o in region['children']:
                                if type_o['data']['obl_name'] == type_name:
                                    type_o['children'].append(org_data)
                                    break
               

    json_result = json.dumps(json_data, indent=4, default=str)

    return HttpResponse(json_result, content_type="application/json")

# старый метод
@api_view(['GET'])
def formfordash_1(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    id_org = request.GET.get("id_org")
    id_org = '%' + id_org + '%'
    id_obl = request.GET.get("id_obl")
    id_region = request.GET.get("id_region")

    query = f"""WITH org as(
            SELECT
                id_org,
                org_name,
                (case when type_org = 'pr' then 'Частный'
                    else 'Государственный' end) as type_org,
                id_obl,
                id_region
            FROM
                serviceback_organizations
			where
				case when '{id_obl}' = 0 then true
					else id_obl = '{id_obl}' end
				and case when '{id_region}' = 0 then true
					else id_region = '{id_region}' end
				and case when '{id_org}' = '' then true
					else id_org like '{id_org}' end),
        allreg as(
                SELECT
                    *
                FROM
                    serviceback_regions),
        namereg as(
                SELECT
                    reg.id as reg_id,
                    reg.name as reg_name
                FROM
                    allreg as reg
                WHERE
                    id in(
                        SELECT
                            id_region
                        FROM
                            org)),
        nameobl as(
                SELECT
                    obl.id as obl_id,
                    obl.name as obl_name
                FROM
                    allreg as obl
                WHERE
                    id in(
                        SELECT
                            id_obl
                        FROM
                            org)),
        full_info as(		
                SELECT 
                    id_org as idd_org,
                    org_name,
                    type_org,
                    id_obl,
                    id_region,
                    obl_name,
                    reg_name
                FROM 
                    org
                LEFT JOIN
                    nameobl
                ON
                    id_obl = obl_id
                LEFT JOIN
                    namereg
                ON
                    id_region = reg_id),
        count_gr as(
            SELECT 
                id_org,
                COUNT(*) as count_gr,
                sum(group_count) as plan_count
            FROM
                serviceback_groups
			WHERE
				id_org in(
                        SELECT
                            id_org
                        FROM
                            org)
            GROUP BY
                id_org),
        count_week as(
            SELECT 
                date_trunc('month', weekend) as mon,
                sum(1) as count_w
            FROM 
                serviceback_weekendday
			WHERE
				weekend <= current_date
            GROUP BY
                mon),
        count_workday as(
            SELECT 
                mon,
                case when not date_trunc('month', current_date::date) = date_trunc('MONTH', mon::date)
						then extract(DAY FROM (date_trunc('MONTH', mon::date) + INTERVAL '1 MONTH' - INTERVAL '1 DAY')) - count_w
					else date_part('day', current_date) - count_w  end as work_day
            FROM 
                count_week),
		groupandplan as(
            SELECT 
                id_org as id_orgplan,
                max(count_gr) as count_gr,
                max(plan_count) as plan_count,
                sum(cast(case when mon = '2023-01-01' then plan_count * work_day else 0 end as numeric)) as plan_jan,
                sum(cast(case when mon = '2023-02-01' then plan_count * work_day else 0 end as numeric)) as plan_feb,
                sum(cast(case when mon = '2023-03-01' then plan_count * work_day else 0 end as numeric)) as plan_march,
                sum(cast(case when mon = '2023-04-01' then plan_count * work_day else 0 end as numeric)) as plan_april,
                sum(cast(case when mon = '2023-05-01' then plan_count * work_day else 0 end as numeric)) as plan_may,
                sum(cast(case when mon = '2023-06-01' then plan_count * work_day else 0 end as numeric)) as plan_june,
                sum(cast(case when mon = '2023-07-01' then plan_count * work_day else 0 end as numeric)) as plan_july,
                sum(cast(case when mon = '2023-08-01' then plan_count * work_day else 0 end as numeric)) as plan_aug,
                sum(cast(case when mon = '2023-09-01' then plan_count * work_day else 0 end as numeric)) as plan_sept,
                sum(cast(case when mon = '2023-10-01' then plan_count * work_day else 0 end as numeric)) as plan_oct,
                sum(cast(case when mon = '2023-11-01' then plan_count * work_day else 0 end as numeric)) as plan_novem,
                sum(cast(case when mon = '2023-12-01' then plan_count * work_day else 0 end as numeric)) as plan_decem
            FROM 
                count_gr, count_workday
            GROUP BY
                id_org),
        all_visit as(
            SELECT
                iin,
                datestatus,
                id_org,
                timestatus,
                status
            FROM
                serviceback_visits
			where
				id_org in(
                        SELECT
                            id_org
                        FROM
                            org)),
        visit as(
            SELECT
                id_org,
                iin,
                datestatus,
                max(timestatus) as timestatus
            FROM
                all_visit
            GROUP BY
                id_org,
                iin,
                datestatus),
        count_visit as(
            SELECT
                id_org,
                iin,
                datestatus,
                timestatus,
                status
            FROM
                all_visit
            WHERE 
                ((id_org,
                iin,
                datestatus,
                timestatus) IN (
                    SELECT
                        id_org,
                        iin,
                        datestatus,
                        timestatus
                    FROM
                        visit)
                and status='2' or status='3' or status='4')),
        count_all as(
            SELECT
                id_org,
                cast(date_trunc('month', datestatus) as date) as date_vis,
                sum(case when status='2' then 1 else 0 end) as count_v,
                sum(case when status='3' then 1 else 0 end) as count_b,
                sum(case when status='4' then 1 else 0 end) as count_o
            FROM
                count_visit
            GROUP BY
                id_org,
                date_vis),
        all_count as(
            SELECT
                id_org as id_orgs,
                sum(case when date_vis = '2023-01-01' then count_v else 0 end) as v_jan,
                sum(case when date_vis = '2023-01-01' then count_b else 0 end) as b_jan,
                sum(case when date_vis = '2023-01-01' then count_o else 0 end) as o_jan,
                sum(case when date_vis = '2023-02-01' then count_v else 0 end) as v_feb,
                sum(case when date_vis = '2023-02-01' then count_b else 0 end) as b_feb,
                sum(case when date_vis = '2023-02-01' then count_o else 0 end) as o_feb,
                sum(case when date_vis = '2023-03-01' then count_v else 0 end) as v_march,
                sum(case when date_vis = '2023-03-01' then count_b else 0 end) as b_march,
                sum(case when date_vis = '2023-03-01' then count_o else 0 end) as o_march,
                sum(case when date_vis = '2023-04-01' then count_v else 0 end) as v_april,
                sum(case when date_vis = '2023-04-01' then count_b else 0 end) as b_april,
                sum(case when date_vis = '2023-04-01' then count_o else 0 end) as o_april,
                sum(case when date_vis = '2023-05-01' then count_v else 0 end) as v_may,
                sum(case when date_vis = '2023-05-01' then count_b else 0 end) as b_may,
                sum(case when date_vis = '2023-05-01' then count_o else 0 end) as o_may,
                sum(case when date_vis = '2023-06-01' then count_v else 0 end) as v_june,
                sum(case when date_vis = '2023-06-01' then count_b else 0 end) as b_june,
                sum(case when date_vis = '2023-06-01' then count_o else 0 end) as o_june,
                sum(case when date_vis = '2023-07-01' then count_v else 0 end) as v_july,
                sum(case when date_vis = '2023-07-01' then count_b else 0 end) as b_july,
                sum(case when date_vis = '2023-07-01' then count_o else 0 end) as o_july,
                sum(case when date_vis = '2023-08-01' then count_v else 0 end) as v_aug,
                sum(case when date_vis = '2023-08-01' then count_b else 0 end) as b_aug,
                sum(case when date_vis = '2023-08-01' then count_o else 0 end) as o_aug,
                sum(case when date_vis = '2023-09-01' then count_v else 0 end) as v_sept,
                sum(case when date_vis = '2023-09-01' then count_b else 0 end) as b_sept,
                sum(case when date_vis = '2023-09-01' then count_o else 0 end) as o_sept,
                sum(case when date_vis = '2023-10-01' then count_v else 0 end) as v_oct,
                sum(case when date_vis = '2023-10-01' then count_b else 0 end) as b_oct,
                sum(case when date_vis = '2023-10-01' then count_o else 0 end) as o_oct,
                sum(case when date_vis = '2023-11-01' then count_v else 0 end) as v_novem,
                sum(case when date_vis = '2023-11-01' then count_b else 0 end) as b_novem,
                sum(case when date_vis = '2023-11-01' then count_o else 0 end) as o_novem,
                sum(case when date_vis = '2023-12-01' then count_v else 0 end) as v_decem,
                sum(case when date_vis = '2023-12-01' then count_b else 0 end) as b_decem,
                sum(case when date_vis = '2023-12-01' then count_o else 0 end) as o_decem
            FROM
                count_all
            GROUP BY 
                id_org)
        SELECT
            f.id_obl,
            (case when f.id_obl is NULL then 1 else 0 end) as por_obl,
			max(f.obl_name) as obl_name,
            f.id_region,
            (case when f.id_region is NULL then 1 else 0 end) as por_reg,
            max(f.reg_name) as reg_name,
			allc.id_orgs as id_org,
            max(f.org_name) as org_name,
            sum(cast(gr.count_gr as integer)) as count_gr,
            sum(cast(gr.plan_count as integer)) as plan_count,
            sum(cast(gr.plan_count as integer)) as fact_count,
            sum(cast(gr.plan_jan as integer)) as plan_1,
            sum(cast(allc.v_jan as integer)) as v_1,
            sum(cast(allc.b_jan as integer)) as b_1,
            sum(cast(allc.o_jan as integer)) as o_1,
            sum(cast((gr.plan_jan - allc.v_jan - allc.b_jan - allc.o_jan) as integer)) as pr_1,
            sum(cast(gr.plan_feb as integer)) as plan_2,
            sum(cast(allc.v_feb as integer)) as v_2,
            sum(cast(allc.b_feb as integer)) as b_2,
            sum(cast(allc.o_feb as integer)) as o_2,
            sum(cast((gr.plan_feb - allc.v_feb - allc.b_feb - allc.o_feb) as integer)) as pr_2,
            sum(cast(gr.plan_march as integer)) as plan_3,
            sum(cast(allc.v_march as integer)) as v_3,
            sum(cast(allc.b_march as integer)) as b_3,
            sum(cast(allc.o_march as integer)) as o_3,
            sum(cast((gr.plan_march - allc.v_march - allc.b_march - allc.o_march) as integer)) as pr_3,
            sum(cast(gr.plan_april as integer)) as plan_4,
            sum(cast(allc.v_april as integer)) as v_4,
            sum(cast(allc.b_april as integer)) as b_4,
            sum(cast(allc.o_april as integer)) as o_4,
            sum(cast((gr.plan_april - allc.v_april - allc.b_april - allc.o_april) as integer)) as pr_4,
            sum(cast(gr.plan_may as integer)) as plan_5,
            sum(cast(allc.v_may as integer)) as v_5,
            sum(cast(allc.b_may as integer)) as b_5,
            sum(cast(allc.o_may as integer)) as o_5,
            sum(cast((gr.plan_may - allc.v_may - allc.b_may - allc.o_may) as integer)) as pr_5,
            sum(cast(gr.plan_june as integer)) as plan_6,
            sum(cast(allc.v_june as integer)) as v_6,
            sum(cast(allc.b_june as integer)) as b_6,
            sum(cast(allc.o_june as integer)) as o_6,
            sum(cast((gr.plan_june - allc.v_june - allc.b_june - allc.o_june) as integer)) as pr_6,
            sum(cast(gr.plan_july as integer)) as plan_7,
            sum(cast(allc.v_july as integer)) as v_7,
            sum(cast(allc.b_july as integer)) as b_7,
            sum(cast(allc.o_july as integer)) as o_7,
            sum(cast((gr.plan_july - allc.v_july - allc.b_july - allc.o_july) as integer)) as pr_7,
            sum(cast(gr.plan_aug as integer)) as plan_8,
            sum(cast(allc.v_aug as integer)) as v_8,
            sum(cast(allc.b_aug as integer)) as b_9,
            sum(cast(allc.o_aug as integer)) as o_9,
            sum(cast((gr.plan_aug - allc.v_aug - allc.b_aug - allc.o_aug) as integer)) as pr_8,
            sum(cast(gr.plan_sept as integer)) as plan_9,
            sum(cast(allc.v_sept as integer)) as v_9,
            sum(cast(allc.b_sept as integer)) as b_9,
            sum(cast(allc.o_sept as integer)) as o_9,
            sum(cast((gr.plan_sept - allc.v_sept - allc.b_sept - allc.o_sept) as integer)) as pr_9,
            sum(cast(gr.plan_oct as integer)) as plan_10,
            sum(cast(allc.v_oct as integer)) as v_10,
            sum(cast(allc.b_oct as integer)) as b_10,
            sum(cast(allc.o_oct as integer)) as o_10,
            sum(cast((gr.plan_oct - allc.v_oct - allc.b_oct - allc.o_oct) as integer)) as pr_10,
            sum(cast(gr.plan_novem as integer)) as plan_11,
            sum(cast(allc.v_novem as integer)) as v_11,
            sum(cast(allc.b_novem as integer)) as b_11,
            sum(cast(allc.o_novem as integer)) as o_11,
            sum(cast((gr.plan_novem - allc.v_novem - allc.b_novem - allc.o_novem) as integer)) as pr_11,
            sum(cast(gr.plan_decem as integer)) as plan_12,
            sum(cast(allc.v_decem as integer)) as v_12,
            sum(cast(allc.b_decem as integer)) as b_12,
            sum(cast(allc.o_decem as integer)) as o_12,
            sum(cast((gr.plan_decem - allc.v_decem - allc.b_decem - allc.o_decem) as integer)) as pr_12,
			max(f.type_org) as type_name,
			(case when f.id_region is NULL then 0
				 when f.type_org='Частный' then f.id_region+200
				 else f.id_region+100 end) as type_id
        FROM
            all_count as allc
        INNER JOIN
            groupandplan as gr
        ON
            id_orgs = id_orgplan
        LEFT JOIN
            full_info as f
        ON
            id_orgs = idd_org
		GROUP BY
		GROUPING SETS ((id_obl), (id_region), (type_id), (id_org))
        ORDER BY por_obl, por_reg, type_id"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        results = cursor.fetchall()

    json_data = []

    for row in results:
        context = {}
        id_obl = row[0]
        obl_name = row[2]
        id_region = row[3]
        reg_name =  row[5]
        id_org = row[6]
        org_name = row[7]
        count_gr = row[8]
        plan_count = row[9]
        fact_count = row[10]
        for i in range(1,13):
            ii = (10 + i) + (4*(i-1))            
            context['plan_{}'.format(i)] = row[ii]
            context['v_{}'.format(i)] = row[ii+1]
            context['b_{}'.format(i)] = row[ii+2]
            context['o_{}'.format(i)] = row[ii+3]
            context['pr_{}'.format(i)] = row[ii+4]
        type_name = row[ii+5]
        type_id = row[ii+6] 

        if id_obl is not None and id_obl not in json_data:
            obl_data = {
                'data': {
                    'sub': 1,
                    'id_obl': id_obl,
                    'obl_name': obl_name,
                    'count_gr': count_gr,
                    'plan_count': plan_count,
                    'fact_count': fact_count
                },
                'children': []
            }

            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    obl_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    obl_data['data'][v_key] = context[v_key]
                if b_key in context:
                    obl_data['data'][b_key] = context[b_key]
                if o_key in context:
                    obl_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    obl_data['data'][pr_key] = context[pr_key]
            json_data.append(obl_data)

        if id_region is not None:
            region_data = {
                'data': {
                            'sub': 2,
                            'id_obl': id_region,
                            'obl_name': reg_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        },
                'children': []
            }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    region_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    region_data['data'][v_key] = context[v_key]
                if b_key in context:
                    region_data['data'][b_key] = context[b_key]
                if o_key in context:
                    region_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    region_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    obl['children'].append(region_data)
                    break

        if type_id is not None:
            type_data = {
                'data': {
                            'sub': 3,
                            'id_obl': type_id,
                            'obl_name': type_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        },
                'children': []
            }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    type_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    type_data['data'][v_key] = context[v_key]
                if b_key in context:
                    type_data['data'][b_key] = context[b_key]
                if o_key in context:
                    type_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    type_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    for region in obl['children']:
                        if region['data']['obl_name'] == reg_name:
                            region['children'].append(type_data)
                            break

        if id_org is not None:
            org_data = {
                        'data': {
                            'sub': 4,
                            'id_obl': id_org,
                            'obl_name': org_name,
                            'count_gr': count_gr,
                            'plan_count': plan_count,
                            'fact_count': fact_count
                        }
                    }
            for i in range(1, 13):
                plan_key = 'plan_{}'.format(i)
                v_key = 'v_{}'.format(i)
                b_key = 'b_{}'.format(i)
                o_key = 'o_{}'.format(i)
                pr_key = 'pr_{}'.format(i)
                if plan_key in context:
                    org_data['data'][plan_key] = context[plan_key]
                if v_key in context:
                    org_data['data'][v_key] = context[v_key]
                if b_key in context:
                    org_data['data'][b_key] = context[b_key]
                if o_key in context:
                    org_data['data'][o_key] = context[o_key]
                if pr_key in context:
                    org_data['data'][pr_key] = context[pr_key]
            for obl in json_data:
                if obl['data']['obl_name'] == obl_name:
                    for region in obl['children']:
                        if region['data']['obl_name'] == reg_name:
                            for type_o in region['children']:
                                if type_o['data']['obl_name'] == type_name:
                                    type_o['children'].append(org_data)
                                    break
               

    json_result = json.dumps(json_data, indent=4, default=str)

    return HttpResponse(json_result, content_type="application/json")

#Сервис для дневного плана по регионам
@api_view(['GET'])
def getStatusRegion(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    id_region = int(request.GET.get("id_region"))
    period = request.GET.get("period")

    if datetime.datetime.now().strftime("%d.%m.%Y") == period: 
        # для текущей даты
        query = f"""
                    WITH org as(
						SELECT
							id_org,
							org_name,
							id_obl,
							id_region
						FROM
							serviceback_organizations
						WHERE
							CASE WHEN '{id_region}' = 0 THEN TRUE
								ELSE id_region = '{id_region}'
							END
						and not id_org = '910114301692'),
				name_reg as(
						SELECT
							reg.id as id_region,
							reg.name as reg_name
						FROM
							serviceback_regions as reg
						WHERE
							reg.id in (
										SELECT
											id_region
										FROM
											org)),
				org_full_info as(
						SELECT
							org.*,
							reg.reg_name
						FROM
							org
						LEFT JOIN
							name_reg as reg
						ON
							org.id_region = reg.id_region),
				child_info as(
                         SELECT
							 id_org,
							 id_group,
                             registered,
                             iin
                         FROM
                             serviceback_childs
                         WHERE
                             id_org in (
									SELECT
										id_org
									FROM
										org)
							and not is_delete),
                    visit as(
                        SELECT
                            id as _id,
                            id_org,
                            id_group,
                            iin,
                            status,
                            comments
                        FROM
                            serviceback_visits
                        WHERE
                            (id_org,
                            id_group,
                            iin) in (
                                SELECT
                                    id_org,
                                    id_group,
                                    iin
                                FROM
                                    child_info)
                            and datestatus = '{period}'),
                        max_vis_id as(
                                SELECT
                                    max(_id) as _id,
                                    iin
                                FROM
                                    visit
                                GROUP BY
                                    iin),
                        vis_finall as(
                                SELECT
                                    *
                                FROM
                                    visit
                                WHERE
                                    _id in(
                                        SELECT
                                            _id
                                        FROM
                                            max_vis_id)),
                        full_info as(
                                SELECT					
                                    vis.status,
                            		vis.comments as com,
                                    ch.*
                                FROM
                                	child_info as ch
                                LEFT JOIN
									vis_finall as vis                                    
                                ON
                                    vis.iin = ch.iin),
                        tmp_full as(
                                SELECT
                                    id_org,
                                    id_group,
                                    iin,
                                    CASE 
                                        WHEN status is NULL and NOT registered THEN '0'
										WHEN status is NULL and registered THEN '1'
                                        WHEN status = '10' and com IS NULL THEN '11'
                                        ELSE status
                                    END as status
                                FROM
                                    full_info),
						tmp_finall as(
                                SELECT 
                                    id_group, 
									id_org,
									CASE
										WHEN status = '0'  THEN 1
										ELSE 0
									END as notregister,
									CASE
										WHEN status = '1' THEN 1
										ELSE 0
									END as notscanned,
									CASE
										WHEN status = '2'  THEN 1
										ELSE 0
									END as visited,
									CASE
										WHEN status = '3'  THEN 1
										ELSE 0
									END as bolnich,
									CASE
										WHEN status = '4' THEN 1
										ELSE 0
									END as otpusk,
                                    CASE
										WHEN status = '5' THEN 1
										ELSE 0
									END as notvisited,
									CASE
										WHEN status = '10' THEN 1
										ELSE 0
									END as fake,
									CASE
										WHEN status = '11' THEN 1
										ELSE 0
									END as checkphoto,
									status as common
                                FROM 
                                    tmp_full),
							tmp_fin as(
								SELECT
									id_org, 
									sum(notregister) as notreg, 
									sum(notscanned) as notscanned,
                                    sum(notvisited) as notvis, 
									sum(visited) as vis, 
									sum(bolnich) as boln, 
									sum(otpusk) as otp,
									sum(fake) as fake,
									sum(checkphoto) as check,
									count(common) as common
								FROM
									tmp_finall
								GROUP BY
									id_org)
								SELECT
									fin.id_org,
									org.org_name,
									org.id_obl,
									org.id_region,
									org.reg_name as name_region,
									fin.vis as status_2,
									fin.boln as status_3,
									fin.otp as status_4,
									fin.fake as status_10,
									fin.common as all_childs,
									fin.common - fin.vis - fin.boln - fin.otp - fin.fake as not_vis
								FROM
									tmp_fin as fin
								LEFT JOIN
									org_full_info as org
								ON
									fin.id_org = org.id_org"""

        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row))
                for row in cursor.fetchall()]

        return HttpResponse(json.dumps(result), content_type="application/json")

    else:
        query = f"""WITH org as(
						SELECT
							id_org,
							org_name,
							id_obl,
							id_region
						FROM
							serviceback_organizations
						WHERE
							CASE WHEN '{id_region}' = 0 THEN TRUE
								ELSE id_region = '{id_region}'
							END
						and not id_org = '910114301692'),
				name_reg as(
						SELECT
							reg.id as id_region,
							reg.name as reg_name
						FROM
							serviceback_regions as reg
						WHERE
							reg.id in (
										SELECT
											id_region
										FROM
											org)),
				org_full_info as(
						SELECT
							org.*,
							reg.reg_name
						FROM
							org
						LEFT JOIN
							name_reg as reg
						ON
							org.id_region = reg.id_region),
				visit as(
                        SELECT
                            id as _id,
                            id_org,
                            id_group,
                            iin,
                            status,
                            comments as com
                        FROM
                            serviceback_visits
                        WHERE
                            id_org in (
									SELECT
										id_org
									FROM
										org)
                            and datestatus = '{period}'),
                        max_vis_id as(
                                SELECT
                                    max(_id) as _id,
                                    iin
                                FROM
                                    visit
                                GROUP BY
                                    iin),
                        vis_finall as(
                                SELECT
                                    *
                                FROM
                                    visit
                                WHERE
                                    _id in(
                                        SELECT
                                            _id
                                        FROM
                                            max_vis_id)),
                        child_info as(
                                SELECT
                                    registered,
                                    iin
                                FROM
                                    serviceback_childs
                                WHERE
                                    iin in(
                                        SELECT
                                            iin
                                        FROM
                                            vis_finall)),
                        full_info as(
                                SELECT					
                                    vis.*,
                                    ch.registered
                                FROM
                                    vis_finall as vis
                                LEFT JOIN
                                    child_info as ch
                                ON
                                    vis.iin = ch.iin),
                        tmp_full as(
                                SELECT
                                    id_org,
                                    id_group,
                                    iin,
                                    CASE 
                                        WHEN status = '1' and NOT registered THEN '0'
                                        WHEN status = '10' and com IS NULL THEN '11'
                                        ELSE status
                                    END as status
                                FROM
                                    full_info),
						tmp_finall as(
                                SELECT 
                                    id_group, 
									id_org,
									CASE
										WHEN status = '0'  THEN 1
										ELSE 0
									END as notregister,
									CASE
										WHEN status = '1' THEN 1
										ELSE 0
									END as notscanned,
									CASE
										WHEN status = '2'  THEN 1
										ELSE 0
									END as visited,
									CASE
										WHEN status = '3'  THEN 1
										ELSE 0
									END as bolnich,
									CASE
										WHEN status = '4' THEN 1
										ELSE 0
									END as otpusk,
                                    CASE
										WHEN status = '5' THEN 1
										ELSE 0
									END as notvisited,
									CASE
										WHEN status = '10' THEN 1
										ELSE 0
									END as fake,
									CASE
										WHEN status = '11' THEN 1
										ELSE 0
									END as checkphoto,
									status as common
                                FROM 
                                    tmp_full),
							tmp_fin as(
								SELECT
									id_org, 
									sum(notregister) as notreg, 
									sum(notscanned) as notscanned,
                                    sum(notvisited) as notvis, 
									sum(visited) as vis, 
									sum(bolnich) as boln, 
									sum(otpusk) as otp,
									sum(fake) as fake,
									sum(checkphoto) as check,
									count(common) as common
								FROM
									tmp_finall
								GROUP BY
									id_org)
								SELECT
									fin.id_org,
									org.org_name,
									org.id_obl,
									org.id_region,
									org.reg_name as name_region,
									fin.vis as status_2,
									fin.boln as status_3,
									fin.otp as status_4,
									fin.fake as status_10,
									fin.common as all_childs,
									fin.common - fin.vis - fin.boln - fin.otp - fin.fake as not_vis
								FROM
									tmp_fin as fin
								LEFT JOIN
									org_full_info as org
								ON
									fin.id_org = org.id_org"""

        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            result = [dict(zip(columns, row))
                for row in cursor.fetchall()]

        return HttpResponse(json.dumps(result), content_type="application/json")

#Сервис для дневного плана по регионам
# старый метод
@api_view(['GET'])
def getStatusRegion_1(request):

    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)
    
    id_region = int(request.GET.get("id_region"))
    period = request.GET.get("period")

    query = f"""
        WITH all_vis as(
            SELECT 
                id_org, 
                iin,
                status,
				timestatus
            FROM
                serviceback_visits 
            WHERE
                datestatus = '{period}'),
		visit as(
			SELECT
				iin,
				max(timestatus) as timestatus
			FROM
				all_vis
			GROUP BY
				iin),
		all_visit as(
			SELECT 
                id_org, 
                iin,
                status
            FROM
                all_vis 
            WHERE
				(iin, timestatus) in
					(SELECT
						iin,
						timestatus
					FROM
						visit)),
        all_childs as(
            SELECT
                id_org,
                count(iin) as child_count
            FROM
                serviceback_childs
            WHERE
                is_delete = false
            GROUP BY
                id_org),
        name_org as(
            SELECT
                id_org,
                org_name,
                id_obl,
                id_region
            FROM
                serviceback_organizations
            WHERE
                id_org in
                    (SELECT
                        id_org
                    FROM
                        all_childs)),
        name_reg as(
            SELECT
                n.id_org,
                n.org_name,
                n.id_obl,
                n.id_region,
                reg.name as reg_name
            FROM
                name_org as n
            LEFT JOIN
                serviceback_regions as reg
            ON
                n.id_region = reg.id),
        tmp_status as(
            SELECT
                id_org,
                sum(CASE WHEN status='2' then 1 else 0 end) as status_2,
                sum(CASE WHEN status='3' then 1 else 0 end) as status_3,
                sum(CASE WHEN status='4' then 1 else 0 end) as status_4,
                sum(CASE WHEN status='10' then 1 else 0 end) as status_10
            FROM
                all_visit
            GROUP BY id_org),
        tbl_status as(
            SELECT
                ch.id_org as id_org,
                CASE WHEN tmp.status_2 IS NULL THEN 0 ELSE tmp.status_2 END as status_2,
                CASE WHEN tmp.status_3 IS NULL THEN 0 ELSE tmp.status_3 END  as status_3,
                CASE WHEN tmp.status_4 IS NULL THEN 0 ELSE tmp.status_4 END  as status_4,
                CASE WHEN tmp.status_10 IS NULL THEN 0 ELSE tmp.status_10 END  as status_10,
                ch.child_count as child_count
            FROM
                tmp_status as tmp
            RIGHT JOIN
                all_childs as ch
            ON
                tmp.id_org = ch.id_org),
        finall_status as(
            SELECT
                tmp.id_org as id_org,
                reg.org_name as org_name,
                reg.id_obl as id_obl,
                reg.id_region as id_region,
                reg.reg_name as name_region,
                tmp.status_2 as status_2,
                tmp.status_3 as status_3,
                tmp.status_4 as status_4,
                tmp.status_10 as status_10,
                tmp.child_count as all_childs,
                tmp.child_count - tmp.status_2 - tmp.status_3 - tmp.status_4 - tmp.status_10 as not_vis
            FROM
                tbl_status as tmp
            INNER JOIN
                name_reg as reg
            ON
                tmp.id_org = reg.id_org
            WHERE
                CASE WHEN '{id_region}' = 0 then true
                    else id_region = '{id_region}' end)
            SELECT 
                * 
            FROM 
                finall_status
            ORDER BY
                id_region"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]
    
    return HttpResponse(json.dumps(result), content_type="application/json")


#Сервис получения других фото ребенка для фэйковые данные
@api_view(['GET'])
def getotherphoto(request):
    
    if request.user.is_anonymous:
        return HttpResponse('{"status": "Ошибка авторизации."}', content_type="application/json", status=401)

    username = request.user
    iin = request.GET.get("iin")
    
    query = f"""SELECT
                    id, 
                    iin, 
                    replace(image_url, 'FilesArchiv', 'https://face11.qazna24.kz/media') as image_url 
                FROM 
                    serviceback_descriptors
                WHERE 
                    iin = '{iin}'
                    and not create_date = current_date
                ORDER BY
                    random()
                LIMIT 4"""

    with connection.cursor() as cursorgr:
        cursorgr.execute(query)
        columnsgr = [col[0] for col in cursorgr.description]
        resultgr = [dict(zip(columnsgr, rowgr))
               for rowgr in cursorgr.fetchall()]

    return HttpResponse(json.dumps(resultgr), content_type="application/json")











